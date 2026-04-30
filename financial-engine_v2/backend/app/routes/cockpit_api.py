from __future__ import annotations

import asyncio
import base64
import csv
import hashlib
import html
import json
import logging
import math
import os
import re
import subprocess
import sys
import threading
import time
import uuid
from concurrent.futures import ThreadPoolExecutor, as_completed
from io import BytesIO
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from collections.abc import Callable
from typing import IO, Any, AsyncGenerator, Literal
from urllib.parse import urlparse

import httpx
from fastapi import APIRouter, HTTPException, Request
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, ConfigDict, Field

from app.core.config import PROJECT_ROOT, settings
from app.core.db import SessionLocal
from app.models.documents import Document
from app.providers.market_price_provider import MarketPriceProvider, MarketPriceProviderError
from app.services.cockpit_service import CockpitService
from app.services.llamacpp_runtime import (
    is_manual_fallback_llm_model,
    resolve_llm_runtime_config,
)
from app.services.marketplace_browser_profile import (
    check_marketplace_browser_health,
    marketplace_scan_health_allows_execution,
)
from app.services.marketplace_benchmark_service import (
    MarketplaceBenchmarkService,
    REVIEW_STATUSES,
)
from app.services.marketplace_mission_service import (
    MarketplaceMissionError,
    MarketplaceMissionNotFound,
    MarketplaceMissionService,
)
from app.services.marketplace_price_intelligence import MarketplacePriceIntelligenceService
from app.services.marketplace_requirement_preparation import (
    marketplace_candidate_contexts,
    marketplace_candidate_products_payload,
    marketplace_requirement_profile,
    prepare_requirement_driven_mission,
)
from app.services.marketplace_scanner import MarketplaceScanCancelled, MarketplaceScanner
from app.services.router_state import get_extraction_activity_snapshot
from app.services.structured_chunking import simple_chunk
from cockpit.core.config import (
    compute_effective_cockpit_config,
    effective_anthropic_api_key,
    load_env,
)
from cockpit.core.conversation_commands import derive_conversational_command
from cockpit.integrations.qual_context_bootstrap import context_enabled

router = APIRouter()
logger = logging.getLogger(__name__)
_MARKETPLACE_SCHEDULER_LOCK = threading.Lock()
_MARKETPLACE_SCHEDULER_STARTED = False
_MARKETPLACE_SCHEDULER_INTERVAL_SECONDS = 60

# How long the SSE chat stream may remain silent before emitting a keepalive
# comment. Module-level so tests can monkey-patch it to a smaller value.
SSE_KEEPALIVE_INTERVAL_SECONDS = 10.0


@dataclass
class QueuedActionJobRuntime:
    job_id: str
    action_id: str
    started_at: str
    stop_event: threading.Event = field(default_factory=threading.Event)
    process: subprocess.Popen[str] | None = None
    terminal_status: str | None = None
    _lock: threading.Lock = field(default_factory=threading.Lock, repr=False)

    def set_process(self, proc: subprocess.Popen[str] | None) -> None:
        with self._lock:
            self.process = proc

    def get_process(self) -> subprocess.Popen[str] | None:
        with self._lock:
            return self.process

    def record_terminal(self, status: str) -> bool:
        with self._lock:
            if self.terminal_status is not None:
                return False
            self.terminal_status = status
            return True


_queued_action_jobs: dict[str, QueuedActionJobRuntime] = {}
_queued_action_jobs_lock = threading.Lock()


def _register_queued_action_job(runtime: QueuedActionJobRuntime) -> None:
    with _queued_action_jobs_lock:
        _queued_action_jobs[runtime.job_id] = runtime


def _get_queued_action_job(job_id: str) -> QueuedActionJobRuntime | None:
    with _queued_action_jobs_lock:
        return _queued_action_jobs.get(job_id)


def _forget_queued_action_job(job_id: str) -> None:
    with _queued_action_jobs_lock:
        _queued_action_jobs.pop(job_id, None)


def _get_backend_job_tracker() -> Any | None:
    try:
        from app.services.job_tracker import get_tracker

        return get_tracker()
    except Exception as exc:
        logger.debug("Backend job tracker unavailable: %s", exc)
        return None


def _best_effort_tracker_call(method: str, *args: Any, **kwargs: Any) -> None:
    tracker = _get_backend_job_tracker()
    if tracker is None:
        return
    try:
        getattr(tracker, method)(*args, **kwargs)
    except Exception as exc:
        logger.debug("Backend tracker %s failed for %s: %s", method, kwargs.get("job_id"), exc)


def _action_job_ticker(args: dict[str, Any]) -> str | None:
    raw = str(args.get("ticker") or args.get("tickers") or "").strip()
    if not raw:
        return None
    return raw.split(",")[0].strip().upper() or None


def _persist_action_job_row(
    service: CockpitService,
    *,
    job_id: str,
    action_id: str,
    args: dict[str, Any],
    started_at: str,
    status: str,
    stdout_path: Path,
    stderr_path: Path,
    exit_code: int | None = None,
    ended_at: str | None = None,
    progress_stage: str | None = None,
    progress_pct: float | None = None,
) -> None:
    existing = service.state_store.get_job(job_id)
    if existing is not None:
        if progress_stage is None:
            progress_stage = existing.get("progress_stage")
        if progress_pct is None:
            progress_pct = existing.get("progress_pct")
    service.state_store.add_job(
        {
            "job_id": job_id,
            "action_id": action_id,
            "args": args,
            "started_at": started_at,
            "ended_at": ended_at,
            "status": status,
            "exit_code": exit_code,
            "stdout_path": str(stdout_path),
            "stderr_path": str(stderr_path),
            "artifacts": [],
            "progress_stage": progress_stage,
            "progress_pct": progress_pct,
        }
    )

_ACTION_JOB_PROCS: dict[str, subprocess.Popen[str]] = {}
_ACTION_JOB_PROC_LOCK = threading.Lock()
_ACTION_JOB_CANCEL_REQUESTS: set[str] = set()


# ---------------------------------------------------------------------------
# Pydantic response models
# ---------------------------------------------------------------------------


class ServiceHealthItem(BaseModel):
    name: str
    status: str  # "healthy" | "degraded" | "down" | "unknown"
    endpoint: str | None = None
    response_time_ms: float | None = None
    error: str | None = None
    details: dict[str, Any] | None = None


class AggregatedHealthResponse(BaseModel):
    status: str  # "healthy" | "degraded" | "down"
    services: list[ServiceHealthItem] = Field(default_factory=list)


class CockpitConfigResponse(BaseModel):
    class ExtractionActivityRun(BaseModel):
        token: str
        run_id: str | None = None
        document_id: str | None = None
        requested_method: str | None = None
        strict_method: bool | None = None
        ticker: str | None = None
        title: str | None = None
        expires_at: float | None = None
        expires_in_seconds: int | None = None

    llm_model: str | None = None
    llm_endpoint: str | None = None
    anthropic_key_configured: bool = False
    extraction_active: bool = False
    extraction_activity_source: str | None = None
    extraction_activity_expires_in_seconds: int | None = None
    extraction_active_runs: list[ExtractionActivityRun] = Field(default_factory=list)
    extract_model: str | None = None
    embed_model: str | None = None
    routing_policy: str | None = None
    backend_url: str | None = None
    profile: str | None = None
    features: dict[str, bool] = Field(default_factory=dict)
    python_version: str | None = None
    git_branch: str | None = None
    data_root: str | None = None


class ModelInfo(BaseModel):
    id: str
    filename: str
    size_gb: float
    quantization: str | None = None
    available: bool = True
    manual_fallback: bool = False


class ModelGroup(BaseModel):
    location: str
    label: str
    models: list[ModelInfo] = Field(default_factory=list)


def _pick_preferred_loaded_model_id(server_models: dict[str, dict[str, Any]]) -> str | None:
    """Prefer native stack models over manual-fallback IDs (e.g. gpt-oss) when several are loaded."""
    loaded = sorted(
        mid
        for mid, inf in server_models.items()
        if str(inf.get("status") or "") == "loaded"
    )
    if not loaded:
        return None
    primary = [m for m in loaded if not is_manual_fallback_llm_model(m)]
    return primary[0] if primary else loaded[0]


def _hoist_manual_fallback_model_groups(groups: list[ModelGroup]) -> list[ModelGroup]:
    """Move opt-in / fallback models into a dedicated trailing group for the settings UI."""
    bucket: list[ModelInfo] = []
    out: list[ModelGroup] = []
    for g in groups:
        kept: list[ModelInfo] = []
        for m in g.models:
            is_fb = is_manual_fallback_llm_model(f"{m.id} {m.filename}")
            entry = ModelInfo(
                id=m.id,
                filename=m.filename,
                size_gb=m.size_gb,
                quantization=m.quantization,
                available=m.available,
                manual_fallback=is_fb,
            )
            if is_fb:
                bucket.append(entry)
            else:
                kept.append(entry)
        if kept:
            out.append(ModelGroup(location=g.location, label=g.label, models=kept))
    if bucket:
        out.append(
            ModelGroup(
                location="manual_fallback",
                label="Manual fallback (opt-in)",
                models=bucket,
            )
        )
    return out


class AvailableModelsResponse(BaseModel):
    groups: list[ModelGroup] = Field(default_factory=list)
    active_model: str | None = None


class ModelLoadRequest(BaseModel):
    model_config = ConfigDict(protected_namespaces=())
    model_id: str | None = None


class ModelLoadResponse(BaseModel):
    ok: bool
    requested_model: str
    resolved_model: str | None = None
    runtime_url: str | None = None
    already_loaded: bool = False
    message: str


class QueueStatusResponse(BaseModel):
    pending: int = 0
    active: int = 0
    completed: int = 0
    failed: int = 0


class CockpitHoldingRecord(BaseModel):
    holding_id: str
    ticker: str
    account_label: str | None = None
    market_exchange: str | None = None
    thesis_bucket: str | None = None
    status: str | None = None
    quantity: float | None = None
    avg_cost: float | None = None
    cost_currency: str | None = None
    opened_at: str | None = None
    updated_at: str | None = None
    note: str | None = None
    current_price: float | None = None
    price_currency: str | None = None
    price_as_of: str | None = None
    market_value: float | None = None
    unrealized_pnl: float | None = None
    valuation_warning: str | None = None


class CockpitHoldingListResponse(BaseModel):
    items: list[CockpitHoldingRecord] = Field(default_factory=list)


class CockpitHoldingCreateRequest(BaseModel):
    ticker: str
    account_label: str | None = None
    market_exchange: str | None = None
    thesis_bucket: str | None = None
    quantity: float | None = None
    avg_cost: float | None = None
    cost_currency: str | None = None
    opened_at: str | None = None
    note: str | None = None


class CockpitHoldingUpdateRequest(BaseModel):
    ticker: str | None = None
    account_label: str | None = None
    market_exchange: str | None = None
    thesis_bucket: str | None = None
    status: str | None = None
    quantity: float | None = None
    avg_cost: float | None = None
    cost_currency: str | None = None
    opened_at: str | None = None
    note: str | None = None


class CockpitHoldingMutationResponse(BaseModel):
    ok: bool
    holding_id: str


class CockpitChatAttachmentUploadRequest(BaseModel):
    filename: str
    content_base64: str
    mime_type: str | None = None
    csv_profile: Literal["auto", "holdings", "trades"] = "auto"
    csv_strict: bool = False


class CockpitChatAttachmentUploadResponse(BaseModel):
    ok: bool = True
    file_kind: Literal["holdings_csv", "strategy_pdf"]
    message: str
    imported_count: int = 0
    skipped_count: int = 0
    errors: list[str] = Field(default_factory=list)
    source_id: str | None = None
    source_kind: Literal["ephemeral", "concat", "primary"] | None = None
    chunks_staged: int = 0
    key_points: list[str] = Field(default_factory=list)


_CSV_HEADER_NORMALIZE_RE = re.compile(r"[^a-z0-9]+")
_CSV_TICKER_RE = re.compile(r"^[A-Za-z][A-Za-z0-9.\-]{1,11}$")
_EXCHANGE_ALIASES = {
    "ASX": "ASX",
    "NASDAQ": "NASDAQ",
    "NAS": "NASDAQ",
    "NYSE": "NYSE",
    "LSE": "LSE",
    "TSX": "TSX",
    "HKEX": "HKEX",
    "HKSE": "HKEX",
}
_CSV_BUY_SIDE_ALIASES = {"buy", "b", "long", "bot", "purchase", "add"}
_CSV_SELL_SIDE_ALIASES = {"sell", "s", "short", "sld", "dispose", "reduce"}
_CSV_HOLDINGS_REQUIRED_COLUMNS = {
    "ticker": ("ticker", "symbol", "asx", "code", "security"),
    "quantity": ("quantity", "qty", "shares", "units", "holding"),
}
_CSV_TRADE_REQUIRED_COLUMNS = {
    "ticker": ("ticker", "symbol", "asx", "code", "security"),
    "side": ("side", "trade_side", "action", "buy_sell", "transaction_type"),
    "quantity": ("quantity", "qty", "shares", "units"),
}
_CSV_TRADE_PRICE_COLUMNS = (
    "price",
    "trade_price",
    "execution_price",
    "fill_price",
    "avg_price",
)
_CSV_TRADE_AMOUNT_COLUMNS = ("amount", "total_value", "notional", "gross_value")
_HOLDINGS_PRICE_CACHE_TTL_SECONDS = 90.0
_HOLDINGS_PRICE_CACHE_LOCK = threading.Lock()
_HOLDINGS_PRICE_CACHE: dict[tuple[str, str | None], tuple[float, dict[str, Any] | None]] = {}
_HOLDINGS_PRICE_EXCHANGE_FALLBACKS = ("ASX", "NASDAQ", "NYSE")


def _normalize_csv_header(value: str) -> str:
    return _CSV_HEADER_NORMALIZE_RE.sub("_", str(value or "").strip().lower()).strip("_")


def _normalize_market_exchange(value: str | None) -> str | None:
    raw = str(value or "").strip().upper()
    if not raw:
        return None
    key = _normalize_csv_header(raw).replace("_", "")
    mapped = _EXCHANGE_ALIASES.get(raw) or _EXCHANGE_ALIASES.get(key)
    return mapped or raw


def _coerce_optional_float(value: str | None) -> float | None:
    cleaned = str(value or "").strip()
    if not cleaned:
        return None
    cleaned = cleaned.replace(",", "")
    cleaned = cleaned.replace("$", "")
    if cleaned.startswith("(") and cleaned.endswith(")"):
        cleaned = f"-{cleaned[1:-1]}"
    return float(cleaned)


def _parse_csv_rows(
    content_text: str,
) -> tuple[list[dict[str, str | None]], dict[str, str]]:
    if not str(content_text or "").strip():
        raise ValueError("CSV file is empty")

    sample = content_text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.DictReader(content_text.splitlines(), dialect=dialect)
    if not reader.fieldnames:
        raise ValueError("CSV header row is missing")

    headers_by_key: dict[str, str] = {}
    for header in reader.fieldnames:
        if header is None:
            continue
        key = _normalize_csv_header(header)
        if key and key not in headers_by_key:
            headers_by_key[key] = header

    rows = [dict(row) for row in reader]
    return rows, headers_by_key


def _parse_xlsx_rows(
    content_bytes: bytes,
) -> tuple[list[dict[str, str | None]], dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError(f"XLSX parser unavailable: {exc}") from exc

    workbook = None
    try:
        workbook = load_workbook(
            filename=BytesIO(content_bytes),
            data_only=True,
            read_only=True,
        )
        worksheet = workbook.active
        raw_headers: list[str] = []
        headers_by_key: dict[str, str] = {}
        rows: list[dict[str, str | None]] = []

        header_found = False
        for excel_row in worksheet.iter_rows(values_only=True):
            normalized_row = [str(value).strip() if value is not None else "" for value in excel_row]
            if not header_found:
                if not any(normalized_row):
                    continue
                header_found = True
                for index, header in enumerate(normalized_row, start=1):
                    final_header = header or f"column_{index}"
                    raw_headers.append(final_header)
                    key = _normalize_csv_header(final_header)
                    if key and key not in headers_by_key:
                        headers_by_key[key] = final_header
                continue

            if not any(normalized_row):
                continue

            row: dict[str, str | None] = {}
            for idx, header in enumerate(raw_headers):
                text = normalized_row[idx] if idx < len(normalized_row) else ""
                row[header] = text or None
            rows.append(row)

        if not header_found:
            raise ValueError("XLSX header row is missing")
        return rows, headers_by_key
    except ValueError:
        raise
    except Exception as exc:
        raise ValueError(f"failed to read XLSX content: {exc}") from exc
    finally:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                pass


def _csv_cell(
    row: dict[str, str | None],
    headers_by_key: dict[str, str],
    *aliases: str,
) -> str:
    for alias in aliases:
        column = headers_by_key.get(alias)
        if column is None:
            continue
        value = row.get(column)
        if value is None:
            continue
        text = str(value).strip()
        if text:
            return text
    return ""


def _has_any_csv_column(headers_by_key: dict[str, str], aliases: tuple[str, ...]) -> bool:
    return any(alias in headers_by_key for alias in aliases)


def _validate_csv_required_columns(
    headers_by_key: dict[str, str],
    required_columns: dict[str, tuple[str, ...]],
    *,
    profile_name: str,
) -> None:
    missing = [
        label
        for label, aliases in required_columns.items()
        if not _has_any_csv_column(headers_by_key, aliases)
    ]
    if missing:
        joined = ", ".join(sorted(missing))
        raise ValueError(
            f"CSV schema mismatch for {profile_name} profile; missing columns: {joined}"
        )


def _detect_csv_profile(headers_by_key: dict[str, str]) -> Literal["holdings", "trades"]:
    has_trade_side = _has_any_csv_column(
        headers_by_key,
        ("side", "trade_side", "action", "buy_sell", "transaction_type"),
    )
    has_trade_price = _has_any_csv_column(
        headers_by_key,
        _CSV_TRADE_PRICE_COLUMNS + _CSV_TRADE_AMOUNT_COLUMNS,
    )
    has_quantity = _has_any_csv_column(
        headers_by_key,
        ("quantity", "qty", "shares", "units", "holding"),
    )
    if has_trade_side and has_quantity:
        return "trades"
    if has_trade_price and has_quantity and "avg_cost" not in headers_by_key:
        return "trades"
    return "holdings"


def _extract_holdings_rows_from_csv(
    rows: list[dict[str, str | None]],
    headers_by_key: dict[str, str],
    *,
    strict: bool = False,
) -> tuple[list[dict[str, Any]], list[str]]:
    if strict:
        _validate_csv_required_columns(
            headers_by_key,
            _CSV_HOLDINGS_REQUIRED_COLUMNS,
            profile_name="holdings",
        )

    def cell(row: dict[str, str | None], *aliases: str) -> str:
        return _csv_cell(row, headers_by_key, *aliases)

    records: list[dict[str, Any]] = []
    errors: list[str] = []
    for row_number, row in enumerate(rows, start=2):
        ticker = cell(row, "ticker", "symbol", "asx", "code", "security").upper()
        if not ticker:
            errors.append(f"row {row_number}: missing ticker")
            continue
        if not _CSV_TICKER_RE.match(ticker):
            errors.append(f"row {row_number}: invalid ticker '{ticker}'")
            continue

        try:
            quantity = _coerce_optional_float(
                cell(row, "quantity", "qty", "shares", "units", "holding")
            )
        except ValueError:
            errors.append(f"row {row_number}: invalid quantity")
            continue
        try:
            avg_cost = _coerce_optional_float(
                cell(
                    row,
                    "avg_cost",
                    "average_cost",
                    "average_price",
                    "avg_price",
                    "cost",
                    "entry_price",
                )
            )
        except ValueError:
            errors.append(f"row {row_number}: invalid avg_cost")
            continue

        if avg_cost is None and quantity is not None and abs(quantity) > 0:
            try:
                total_cost = _coerce_optional_float(
                    cell(
                        row,
                        "cost_basis",
                        "total_cost",
                        "invested_amount",
                        "invested",
                    )
                )
            except ValueError:
                errors.append(f"row {row_number}: invalid cost_basis")
                continue
            if total_cost is not None:
                avg_cost = total_cost / abs(quantity)
            else:
                try:
                    value = _coerce_optional_float(
                        cell(
                            row,
                            "value",
                            "market_value",
                            "position_value",
                            "current_value",
                        )
                    )
                    capital_gain = _coerce_optional_float(
                        cell(
                            row,
                            "capital_gain",
                            "gain",
                            "unrealized_gain",
                            "pnl",
                            "profit_loss",
                        )
                    )
                except ValueError:
                    errors.append(f"row {row_number}: invalid value/capital_gain")
                    continue
                if value is not None and capital_gain is not None:
                    avg_cost = (value - capital_gain) / abs(quantity)

        cost_currency = cell(row, "cost_currency", "currency", "ccy").upper() or None
        if cost_currency and len(cost_currency) > 8:
            errors.append(f"row {row_number}: invalid cost_currency '{cost_currency}'")
            continue

        records.append(
            {
                "ticker": ticker,
                "account_label": cell(row, "account_label", "account", "broker") or None,
                "market_exchange": _normalize_market_exchange(
                    cell(row, "market_exchange", "exchange", "market", "venue", "market_code")
                ),
                "thesis_bucket": cell(row, "thesis_bucket", "bucket", "strategy") or None,
                "quantity": quantity,
                "avg_cost": avg_cost,
                "cost_currency": cost_currency,
                "opened_at": cell(row, "opened_at", "open_date", "date", "acquired_at")
                or None,
                "note": cell(row, "note", "notes", "comment") or None,
            }
        )

    return records, errors


def _normalize_trade_side(raw_side: str) -> Literal["buy", "sell"] | None:
    normalized = _normalize_csv_header(raw_side).replace("_", "")
    if not normalized:
        return None
    if normalized in _CSV_BUY_SIDE_ALIASES:
        return "buy"
    if normalized in _CSV_SELL_SIDE_ALIASES:
        return "sell"
    return None


def _extract_trade_holdings_rows_from_csv(
    rows: list[dict[str, str | None]],
    headers_by_key: dict[str, str],
    *,
    strict: bool = False,
) -> tuple[list[dict[str, Any]], list[str]]:
    if strict:
        _validate_csv_required_columns(
            headers_by_key,
            _CSV_TRADE_REQUIRED_COLUMNS,
            profile_name="trades",
        )

    trade_rows: list[dict[str, Any]] = []
    errors: list[str] = []
    for row_number, row in enumerate(rows, start=2):
        ticker = _csv_cell(
            row, headers_by_key, "ticker", "symbol", "asx", "code", "security"
        ).upper()
        if not ticker:
            errors.append(f"row {row_number}: missing ticker")
            continue
        if not _CSV_TICKER_RE.match(ticker):
            errors.append(f"row {row_number}: invalid ticker '{ticker}'")
            continue

        raw_side = _csv_cell(
            row,
            headers_by_key,
            "side",
            "trade_side",
            "action",
            "buy_sell",
            "transaction_type",
        )
        side = _normalize_trade_side(raw_side)
        try:
            quantity_value = _coerce_optional_float(
                _csv_cell(row, headers_by_key, "quantity", "qty", "shares", "units")
            )
        except ValueError:
            errors.append(f"row {row_number}: invalid quantity")
            continue
        if quantity_value is None or abs(quantity_value) <= 0.0:
            errors.append(f"row {row_number}: missing quantity")
            continue
        quantity = abs(quantity_value)
        if side is None:
            if strict:
                errors.append(f"row {row_number}: invalid side '{raw_side}'")
                continue
            side = "sell" if quantity_value < 0 else "buy"

        try:
            trade_price = _coerce_optional_float(
                _csv_cell(row, headers_by_key, *_CSV_TRADE_PRICE_COLUMNS)
            )
        except ValueError:
            errors.append(f"row {row_number}: invalid price")
            continue
        try:
            trade_amount = _coerce_optional_float(
                _csv_cell(row, headers_by_key, *_CSV_TRADE_AMOUNT_COLUMNS)
            )
        except ValueError:
            errors.append(f"row {row_number}: invalid amount")
            continue

        if trade_price is None and trade_amount is not None:
            trade_price = abs(trade_amount) / quantity
        if trade_price is not None and trade_price <= 0:
            errors.append(f"row {row_number}: invalid price")
            continue
        if strict and trade_price is None:
            errors.append(
                f"row {row_number}: missing price or amount for strict trades profile"
            )
            continue

        cost_currency = _csv_cell(
            row, headers_by_key, "cost_currency", "currency", "ccy"
        ).upper() or None
        if cost_currency and len(cost_currency) > 8:
            errors.append(f"row {row_number}: invalid cost_currency '{cost_currency}'")
            continue

        trade_rows.append(
            {
                "row_number": row_number,
                "ticker": ticker,
                "side": side,
                "quantity": quantity,
                "trade_price": trade_price,
                "account_label": _csv_cell(
                    row, headers_by_key, "account_label", "account", "broker"
                )
                or None,
                "market_exchange": _normalize_market_exchange(
                    _csv_cell(
                        row,
                        headers_by_key,
                        "market_exchange",
                        "exchange",
                        "market",
                        "venue",
                        "market_code",
                    )
                ),
                "thesis_bucket": _csv_cell(
                    row, headers_by_key, "thesis_bucket", "bucket", "strategy"
                )
                or None,
                "cost_currency": cost_currency,
                "opened_at": _csv_cell(
                    row,
                    headers_by_key,
                    "trade_date",
                    "date",
                    "executed_at",
                    "opened_at",
                    "open_date",
                )
                or None,
                "note": _csv_cell(row, headers_by_key, "note", "notes", "comment") or None,
            }
        )

    aggregates: dict[tuple[str, str | None, str | None, str | None], dict[str, Any]] = {}
    for trade in trade_rows:
        key = (
            str(trade["ticker"]),
            trade["account_label"],
            trade["market_exchange"],
            trade["thesis_bucket"],
        )
        accumulator = aggregates.get(key)
        if accumulator is None:
            accumulator = {
                "ticker": trade["ticker"],
                "account_label": trade["account_label"],
                "market_exchange": trade["market_exchange"],
                "thesis_bucket": trade["thesis_bucket"],
                "quantity": 0.0,
                "cost_basis": 0.0,
                "cost_currency": trade["cost_currency"],
                "opened_at": trade["opened_at"],
                "note": trade["note"],
                "unknown_cost": False,
            }
            aggregates[key] = accumulator
        elif not accumulator.get("cost_currency") and trade["cost_currency"]:
            accumulator["cost_currency"] = trade["cost_currency"]

        side = str(trade["side"])
        quantity = float(trade["quantity"])
        trade_price = trade["trade_price"]
        if side == "buy":
            accumulator["quantity"] += quantity
            if trade_price is None:
                accumulator["unknown_cost"] = True
            else:
                accumulator["cost_basis"] += quantity * float(trade_price)
            continue

        # Selling adjusts quantity and reduces cost basis proportionally.
        current_qty = float(accumulator["quantity"])
        if current_qty <= 0:
            errors.append(
                f"row {trade['row_number']}: sell before buy for {trade['ticker']}"
            )
            continue
        if strict and quantity > current_qty + 1e-9:
            errors.append(
                "row "
                f"{trade['row_number']}: sell quantity exceeds open position for "
                f"{trade['ticker']}"
            )
            continue
        sold_qty = min(quantity, current_qty)
        avg_cost = float(accumulator["cost_basis"]) / current_qty if current_qty > 0 else 0.0
        accumulator["quantity"] = current_qty - sold_qty
        accumulator["cost_basis"] = max(
            0.0, float(accumulator["cost_basis"]) - (avg_cost * sold_qty)
        )
        if float(accumulator["quantity"]) <= 1e-9:
            accumulator["quantity"] = 0.0
            accumulator["cost_basis"] = 0.0

    records: list[dict[str, Any]] = []
    for key in sorted(aggregates):
        item = aggregates[key]
        quantity = float(item["quantity"])
        if quantity <= 1e-9:
            continue
        avg_cost: float | None = None
        if not item["unknown_cost"]:
            avg_cost = float(item["cost_basis"]) / quantity if quantity > 0 else None
        records.append(
            {
                "ticker": item["ticker"],
                "account_label": item["account_label"],
                "market_exchange": item["market_exchange"],
                "thesis_bucket": item["thesis_bucket"],
                "quantity": quantity,
                "avg_cost": avg_cost,
                "cost_currency": item["cost_currency"],
                "opened_at": item["opened_at"],
                "note": item["note"],
            }
        )
    return records, errors


def _extract_pdf_text(content_bytes: bytes) -> str:
    if not content_bytes.startswith(b"%PDF"):
        raise ValueError("uploaded file is not a PDF")

    try:
        import fitz  # type: ignore[import-not-found]
    except Exception as exc:
        raise RuntimeError(f"PDF parser unavailable: {exc}") from exc

    document = None
    try:
        document = fitz.open(stream=content_bytes, filetype="pdf")
        chunks: list[str] = []
        total_chars = 0
        max_chars = 300_000
        for page in document:
            text = str(page.get_text("text") or "")
            text = re.sub(r"\s+\n", "\n", text)
            text = re.sub(r"[ \t]+", " ", text).strip()
            if not text:
                continue
            remaining = max_chars - total_chars
            if remaining <= 0:
                break
            if len(text) > remaining:
                text = text[:remaining]
            if text:
                chunks.append(text)
                total_chars += len(text)
        combined = "\n".join(chunks).strip()
        if not combined:
            raise ValueError("no extractable text found in PDF")
        return combined
    except ValueError:
        raise
    except Exception as exc:
        raise RuntimeError(f"failed to read PDF content: {exc}") from exc
    finally:
        if document is not None:
            try:
                document.close()
            except Exception:
                pass


def _derive_key_points_from_text(text: str, limit: int = 5) -> list[str]:
    points: list[str] = []
    seen: set[str] = set()

    def push(candidate: str) -> None:
        if len(points) >= limit:
            return
        cleaned = re.sub(r"\s+", " ", str(candidate or "")).strip(" -\t")
        if len(cleaned) < 40 or len(cleaned) > 260:
            return
        normalized = cleaned.lower()
        if normalized in seen:
            return
        seen.add(normalized)
        points.append(cleaned)

    for line in text.splitlines():
        if len(points) >= limit:
            break
        push(line)

    if len(points) < limit:
        for sentence in re.split(r"(?<=[.!?])\s+", text[:24000]):
            if len(points) >= limit:
                break
            push(sentence)

    return points


def _stage_uploaded_pdf_chunks(
    *,
    filename: str,
    extracted_text: str,
    published_at: str,
) -> tuple[str, int]:
    stage_dir = Path("~/.tenn/memory/staged_chunks").expanduser().resolve()
    stage_dir.mkdir(parents=True, exist_ok=True)

    digest = hashlib.sha1(
        f"{filename}|{published_at}|{extracted_text}".encode("utf-8")
    ).hexdigest()[:16]
    slug = re.sub(r"[^a-z0-9]+", "-", filename.lower()).strip("-") or "uploaded-pdf"
    source_id = f"market_commentary:{slug}:{digest}"

    chunks = []
    seen: set[str] = set()
    for chunk in simple_chunk(extracted_text, max_chars=1400):
        normalized = re.sub(r"\s+", " ", str(chunk or "")).strip()
        if not normalized:
            continue
        key = normalized.lower()
        if key in seen:
            continue
        seen.add(key)
        chunks.append(normalized)

    if not chunks:
        raise ValueError("no usable text chunks extracted from PDF")

    staged_path = stage_dir / f"{source_id}.jsonl"
    with staged_path.open("w", encoding="utf-8") as handle:
        for index, chunk in enumerate(chunks):
            row = {
                "payload": {
                    "chunk_id": f"{source_id}:{index}",
                    "source_id": source_id,
                    "chunk_index": index,
                    "text": chunk,
                    "source_name": filename,
                    "source_type": "market_commentary",
                    "speaker": "Uploaded PDF",
                    "published_at": published_at,
                }
            }
            handle.write(json.dumps(row, ensure_ascii=False) + "\n")

    return source_id, len(chunks)


def _effective_cockpit_feature_flags(
    effective_cfg: dict[str, Any],
) -> dict[str, bool]:
    web_cfg = effective_cfg.get("web") if isinstance(effective_cfg.get("web"), dict) else {}
    rag_cfg = effective_cfg.get("rag") if isinstance(effective_cfg.get("rag"), dict) else {}
    db_cfg = effective_cfg.get("db") if isinstance(effective_cfg.get("db"), dict) else {}
    qual_cfg = (
        rag_cfg.get("qualitative_context")
        if isinstance(rag_cfg.get("qualitative_context"), dict)
        else {}
    )
    news_cfg = (
        rag_cfg.get("news_context")
        if isinstance(rag_cfg.get("news_context"), dict)
        else {}
    )

    rag_setting = rag_cfg.get("enabled")
    rag_config_enabled = (
        bool(rag_setting)
        if rag_setting is not None
        else bool(settings.enable_embeddings and settings.enable_qdrant)
    )
    rag_path_enabled = bool(rag_cfg.get("backend_search_enabled", False)) or context_enabled(
        qual_cfg, default=False
    ) or context_enabled(news_cfg, default=False)

    return {
        "web_search": bool(web_cfg.get("enabled_default", False)),
        "rag": rag_config_enabled and (
            rag_path_enabled or bool(settings.enable_embeddings and settings.enable_qdrant)
        ),
        "extraction": bool(settings.enable_extraction),
        "session_memory": bool(getattr(settings, "enable_session_memory", True)),
        "db_diagnostics": bool(db_cfg.get("diagnostic_query_enabled", False)),
    }


class IntelPulseStats(BaseModel):
    document_count: int = 0
    extraction_count: int = 0
    signal_count: int = 0
    memory_count: int = 0
    population_index: float = 0.0
    trust_score_avg: float = 0.0
    quarantine_rate: float = 0.0


class IntelPulseStageHealth(BaseModel):
    id: str
    label: str
    health: float
    status: str


class IntelPulseFailure(BaseModel):
    id: str
    entity: str
    type: str
    message: str
    confidence: float
    timestamp: str


class IntelPulseResponse(BaseModel):
    stats: IntelPulseStats
    pipeline: list[IntelPulseStageHealth]
    failures: list[IntelPulseFailure] | None = None


class IntelPulseEntityMetric(BaseModel):
    entity: str
    metrics: dict[
        str, str
    ]  # metric_name -> status (populated, abstain, failed, sparse)


class IntelPulseMatrixResponse(BaseModel):
    stage: str
    entities: list[IntelPulseEntityMetric]


# ---------------------------------------------------------------------------
# Helper: normalize chat sources for cockpit UI
# ---------------------------------------------------------------------------


def _safe_source_score(value: Any) -> float:
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return 0.0
    return numeric if math.isfinite(numeric) else 0.0


def _clean_source_text(value: Any, *, max_chars: int = 280) -> str | None:
    text = re.sub(r"\s+", " ", str(value or "").strip())
    if not text:
        return None
    if len(text) <= max_chars:
        return text
    return text[: max_chars - 3].rstrip() + "..."


def _summarize_scalar_fields(raw: dict[str, Any], *, max_items: int = 4) -> str | None:
    bits: list[str] = []
    for key, value in raw.items():
        if isinstance(value, (dict, list, tuple, set)) or value is None:
            continue
        label = str(key).strip().replace("_", " ")
        text = str(value).strip()
        if not label or not text:
            continue
        bits.append(f"{label}: {text}")
        if len(bits) >= max_items:
            break
    if not bits:
        return None
    return _clean_source_text("; ".join(bits))


def _normalize_source_item(
    raw: dict[str, Any],
    *,
    default_title: str = "Source",
    kind: str = "context",
) -> dict[str, Any] | None:
    if not isinstance(raw, dict):
        return None

    title = str(
        raw.get("title")
        or raw.get("source_name")
        or raw.get("source")
        or raw.get("file")
        or raw.get("document_id")
        or default_title
    ).strip()
    url = str(raw.get("url") or raw.get("source_url") or "").strip()
    document_id = str(raw.get("document_id") or raw.get("source_document_id") or "").strip()
    source_id = str(raw.get("source_id") or raw.get("chunk_id") or "").strip()
    path = str(raw.get("path") or raw.get("pdf_path") or raw.get("file") or "").strip()
    doc_type = str(
        raw.get("doc_type")
        or raw.get("doc_class")
        or raw.get("source_corpus")
        or raw.get("corpus")
        or ""
    ).strip()
    published_at = str(raw.get("published_at") or "").strip()
    snippet = _clean_source_text(
        raw.get("snippet")
        or raw.get("text")
        or raw.get("excerpt")
        or raw.get("content")
        or raw.get("claim")
    )

    if not title and not url and not snippet and not document_id and not path:
        return None

    return {
        "title": title or default_title,
        "score": _safe_source_score(
            raw.get("score") or raw.get("final_score") or raw.get("semantic_score")
        ),
        "url": url or None,
        "snippet": snippet,
        "published_at": published_at or None,
        "document_id": document_id or None,
        "source_id": source_id or None,
        "doc_type": doc_type or None,
        "path": path or None,
        "kind": kind,
    }


def _append_source_item(
    items: list[dict[str, Any]],
    seen: set[str],
    raw: dict[str, Any],
    *,
    default_title: str = "Source",
    kind: str = "context",
    limit: int = 10,
) -> None:
    if len(items) >= limit:
        return

    item = _normalize_source_item(raw, default_title=default_title, kind=kind)
    if item is None:
        return

    dedupe_key = next(
        (
            str(candidate).strip().lower()
            for candidate in (
                item.get("url"),
                item.get("source_id"),
                item.get("document_id"),
                item.get("title"),
            )
            if str(candidate or "").strip()
        ),
        "",
    )
    if not dedupe_key or dedupe_key in seen:
        return

    seen.add(dedupe_key)
    items.append(item)


def _decode_truncated_tool_result(result: dict[str, Any]) -> dict[str, Any]:
    """Best-effort decode for historical truncated tool payload envelopes.

    Older tool truncation stored a JSON string under ``data`` which made source
    extraction impossible. Newer truncation keeps structured fields, so this is
    primarily a backward-compatible decoder.
    """
    if not isinstance(result, dict):
        return {}
    if not result.get("_truncated"):
        return result
    data = result.get("data")
    if not isinstance(data, str):
        return result

    parsed: dict[str, Any] | None = None
    try:
        raw = json.loads(data)
        if isinstance(raw, dict):
            parsed = raw
    except (TypeError, ValueError, json.JSONDecodeError):
        # Try parsing up to the last complete object brace.
        last_brace = data.rfind("}")
        if last_brace > 1:
            candidate = data[: last_brace + 1]
            try:
                raw = json.loads(candidate)
                if isinstance(raw, dict):
                    parsed = raw
            except (TypeError, ValueError, json.JSONDecodeError):
                parsed = None

    if parsed is None:
        return result

    merged = dict(parsed)
    for key, value in result.items():
        if key in {"data"}:
            continue
        if key not in merged or merged.get(key) in (None, "", [], {}):
            merged[key] = value
    return merged


def _dict_rows(value: Any) -> list[dict[str, Any]]:
    return [row for row in value if isinstance(row, dict)] if isinstance(value, list) else []


def _source_payloads_for_evidence(ev: dict[str, Any], details: dict[str, Any]) -> list[dict[str, Any]]:
    payloads: list[dict[str, Any]] = []
    for candidate in (details, ev.get("result")):
        if isinstance(candidate, dict):
            payloads.append(candidate)
            for nested_key in ("financial_truth", "backend"):
                nested = candidate.get(nested_key)
                if isinstance(nested, dict):
                    payloads.append(nested)
    return payloads


def _append_financial_payload_sources(
    items: list[dict[str, Any]],
    seen: set[str],
    payload: dict[str, Any],
) -> None:
    for row in _dict_rows(payload.get("announcement_context")):
        _append_source_item(
            items,
            seen,
            row,
            default_title="Announcement excerpt",
            kind="document",
        )

    for row in _dict_rows(payload.get("docs")):
        _append_source_item(
            items,
            seen,
            row,
            default_title="Financial document",
            kind="document",
        )

    financial_rows = _dict_rows(payload.get("financials"))
    snapshot = payload.get("latest_financial_snapshot")
    if isinstance(snapshot, dict):
        financial_rows.append(snapshot)

    for row in financial_rows:
        ticker = str(row.get("ticker") or "").strip().upper()
        period_type = str(row.get("period_type") or "").strip()
        period_end = str(row.get("period_end") or "").strip()
        title = " ".join(part for part in (ticker or "Financials", period_type, period_end) if part)
        metric_bits = []
        for metric in (
            "revenue",
            "ebit",
            "np_attributable",
            "operating_cf",
            "cash_end",
            "net_debt",
            "shares_outstanding",
        ):
            value = row.get(metric)
            if value not in (None, ""):
                metric_bits.append(f"{metric}: {value}")
            if len(metric_bits) >= 4:
                break
        _append_source_item(
            items,
            seen,
            {
                **row,
                "title": title,
                "document_id": row.get("source_document_id"),
                "source_id": (
                    row.get("source_id")
                    or row.get("source_document_id")
                    or f"financials:{ticker or 'unknown'}:{period_end or 'unknown'}:{period_type or 'period'}"
                ),
                "published_at": period_end or row.get("published_at"),
                "doc_type": period_type or row.get("doc_type"),
                "snippet": "; ".join(metric_bits) if metric_bits else None,
            },
            default_title="Financial period",
            kind="document",
        )


def _append_memory_payload_sources(
    items: list[dict[str, Any]],
    seen: set[str],
    payload: dict[str, Any],
    *,
    default_title: str,
    source_prefix: str,
) -> None:
    for row in _dict_rows(payload.get("items")):
        entry_id = row.get("entry_id") or row.get("proposal_id") or row.get("source_id")
        source = str(row.get("source") or default_title).strip()
        signal_type = str(row.get("type") or row.get("entry_type") or row.get("signal") or "").strip()
        title = " ".join(part for part in (source, signal_type) if part) or default_title
        _append_source_item(
            items,
            seen,
            {
                **row,
                "title": title,
                "source_id": row.get("source_id") or (f"{source_prefix}:{entry_id}" if entry_id else None),
                "snippet": row.get("statement") or row.get("summary"),
                "published_at": row.get("updated_at") or row.get("created_at"),
                "score": row.get("active_score") or row.get("confidence"),
            },
            default_title=default_title,
            kind="context",
        )


def _append_youtube_recent_video_sources(
    items: list[dict[str, Any]],
    seen: set[str],
    result: dict[str, Any],
) -> None:
    channel = str(result.get("name") or result.get("channel_name") or "YouTube channel").strip()
    for index, video in enumerate(_dict_rows(result.get("videos")), start=1):
        video_id = str(video.get("video_id") or video.get("id") or "").strip()
        url = str(video.get("webpage_url") or video.get("url") or "").strip()
        if not url and video_id:
            url = f"https://www.youtube.com/watch?v={video_id}"
        duration = video.get("duration_seconds")
        bits = [f"Channel: {channel}"]
        if video_id:
            bits.append(f"Video ID: {video_id}")
        if isinstance(duration, (int, float)) and duration > 0:
            bits.append(f"Duration: {int(round(float(duration) / 60.0))} min")
        scores = video.get("scores") if isinstance(video.get("scores"), dict) else {}
        score = scores.get("overall") if isinstance(scores, dict) else None
        if isinstance(score, (int, float)):
            bits.append(f"Score: {float(score):.2f}")
        _append_source_item(
            items,
            seen,
            {
                "title": video.get("title") or video_id or f"YouTube video {index}",
                "url": url or None,
                "source_id": video.get("source_id") or (f"youtube:{video_id}" if video_id else None),
                "snippet": "; ".join(bits),
                "published_at": video.get("published_at") or video.get("published_date"),
                "score": score,
                "doc_type": "youtube_video",
            },
            default_title="YouTube video",
            kind="web",
        )


def _append_news_no_hit_source(
    items: list[dict[str, Any]],
    seen: set[str],
    result: dict[str, Any],
) -> None:
    query = str(result.get("query") or result.get("normalized_query") or "").strip()
    ticker = str(result.get("ticker") or "").strip().upper()
    freshness = str(result.get("freshness_warning") or "").strip()
    searched = query or ticker or "news"
    bits = [f"No news hits returned for {searched}."]
    if freshness:
        bits.append(freshness)
    _append_source_item(
        items,
        seen,
        {
            "title": f"News search: no hits for {searched}",
            "source_id": f"search_news:no_hits:{searched.lower()}",
            "snippet": " ".join(bits),
            "score": 1.0,
            "doc_type": "operational_no_hit",
        },
        default_title="News search audit",
        kind="context",
    )


def _build_ui_sources(evidence: list[dict[str, Any]] | None) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    seen: set[str] = set()

    for ev in evidence or []:
        if not isinstance(ev, dict):
            continue
        ev_type = str(ev.get("type") or "").strip().lower()
        details_payload = ev.get("details")
        details = details_payload if isinstance(details_payload, dict) else {}

        if ev_type == "local_context":
            qual_context = (
                details.get("qual_context") if isinstance(details.get("qual_context"), dict) else {}
            )
            for hit in qual_context.get("hits", []) if isinstance(qual_context.get("hits"), list) else []:
                if isinstance(hit, dict):
                    _append_source_item(
                        items,
                        seen,
                        hit,
                        default_title="Context source",
                        kind="rag",
                    )

            for row in details.get("docs", []) if isinstance(details.get("docs"), list) else []:
                if isinstance(row, dict):
                    _append_source_item(
                        items,
                        seen,
                        row,
                        default_title="Document",
                        kind="document",
                    )

            for row in details.get("doc_snippets", []) if isinstance(details.get("doc_snippets"), list) else []:
                if isinstance(row, dict):
                    _append_source_item(
                        items,
                        seen,
                        row,
                        default_title="Document excerpt",
                        kind="document",
                    )

            for row in details.get("web_facts", []) if isinstance(details.get("web_facts"), list) else []:
                if isinstance(row, dict):
                    _append_source_item(
                        items,
                        seen,
                        row,
                        default_title="Web fact",
                        kind="web",
                    )

        elif ev_type == "company_dump":
            backend = details.get("backend") if isinstance(details.get("backend"), dict) else {}
            for row in backend.get("docs", []) if isinstance(backend.get("docs"), list) else []:
                if isinstance(row, dict):
                    _append_source_item(
                        items,
                        seen,
                        row,
                        default_title="Company document",
                        kind="document",
                    )

            for row in backend.get("announcement_context", []) if isinstance(backend.get("announcement_context"), list) else []:
                if isinstance(row, dict):
                    _append_source_item(
                        items,
                        seen,
                        row,
                        default_title="Announcement excerpt",
                        kind="document",
                    )

            _append_financial_payload_sources(items, seen, backend)

        elif ev_type in {
            "financial_truth",
            "company_memory",
            "market_memory",
            "user_thesis_memory",
            "orchestrator",
        }:
            for payload in _source_payloads_for_evidence(ev, details):
                _append_financial_payload_sources(items, seen, payload)
                if ev_type == "company_memory" or isinstance(payload.get("company_memory"), dict):
                    memory_payload = (
                        payload.get("company_memory")
                        if isinstance(payload.get("company_memory"), dict)
                        else payload
                    )
                    _append_memory_payload_sources(
                        items,
                        seen,
                        memory_payload,
                        default_title="Company memory",
                        source_prefix="company_memory",
                    )
                if ev_type == "market_memory" or isinstance(payload.get("market_memory"), dict):
                    memory_payload = (
                        payload.get("market_memory")
                        if isinstance(payload.get("market_memory"), dict)
                        else payload
                    )
                    _append_memory_payload_sources(
                        items,
                        seen,
                        memory_payload,
                        default_title="Market memory",
                        source_prefix="market_memory",
                    )
                if ev_type == "user_thesis_memory" or isinstance(payload.get("user_thesis_memory"), dict):
                    memory_payload = (
                        payload.get("user_thesis_memory")
                        if isinstance(payload.get("user_thesis_memory"), dict)
                        else payload
                    )
                    _append_memory_payload_sources(
                        items,
                        seen,
                        memory_payload,
                        default_title="User thesis memory",
                        source_prefix="user_thesis_memory",
                    )

        elif ev_type == "news_search":
            for row in details.get("hits", []) if isinstance(details.get("hits"), list) else []:
                if isinstance(row, dict):
                    _append_source_item(
                        items,
                        seen,
                        row,
                        default_title="News source",
                        kind="news",
                    )
            if not _dict_rows(details.get("hits")):
                _append_news_no_hit_source(items, seen, details)

        elif ev_type in {"news_summary", "article_request"}:
            _append_source_item(
                items,
                seen,
                details,
                default_title="News source",
                kind="news",
            )

        elif ev_type == "web":
            pages = details.get("pages") if isinstance(details.get("pages"), list) else []
            facts = details.get("facts") if isinstance(details.get("facts"), list) else []
            if pages:
                for row in pages:
                    if isinstance(row, dict):
                        _append_source_item(
                            items,
                            seen,
                            row,
                            default_title="Web source",
                            kind="web",
                        )
            elif facts:
                for row in facts:
                    if isinstance(row, dict):
                        _append_source_item(
                            items,
                            seen,
                            row,
                            default_title="Web source",
                            kind="web",
                        )
            else:
                _append_source_item(
                    items,
                    seen,
                    details,
                    default_title="Web source",
                    kind="web",
                )

        elif ev_type == "runtime_clock":
            _append_source_item(
                items,
                seen,
                details,
                default_title="Cockpit runtime clock",
                kind="context",
            )

        elif ev_type == "holdings":
            rows = (
                details_payload
                if isinstance(details_payload, list)
                else (
                    details.get("items")
                    if isinstance(details.get("items"), list)
                    else []
                )
            )
            for row in rows:
                if not isinstance(row, dict):
                    continue
                ticker = str(row.get("ticker") or "").strip().upper()
                account = str(row.get("account_label") or "").strip()
                qty = row.get("quantity")
                avg_cost = row.get("avg_cost")
                bits: list[str] = []
                if account:
                    bits.append(f"Account: {account}")
                if qty is not None:
                    bits.append(f"Quantity: {qty}")
                if avg_cost is not None:
                    bits.append(f"Avg cost: {avg_cost}")
                _append_source_item(
                    items,
                    seen,
                    {
                        **row,
                        "title": f"{ticker or 'Holding'} holding",
                        "source_id": (
                            row.get("holding_id")
                            or (f"holding:{ticker}" if ticker else None)
                        ),
                        "snippet": "; ".join(bits) if bits else None,
                    },
                    default_title="Holding",
                    kind="context",
                )

        elif ev_type in ("market_update_report", "market_update_reports"):
            reports = (
                [details_payload]
                if isinstance(details_payload, dict)
                else (details_payload if isinstance(details_payload, list) else [])
            )
            for report in reports:
                if not isinstance(report, dict):
                    continue
                run_date = str(
                    report.get("report_date") or report.get("run_date") or ""
                ).strip()
                status = str(report.get("status") or "").strip()
                summary = report.get("summary") if isinstance(report.get("summary"), dict) else {}
                movers = summary.get("movers") if isinstance(summary.get("movers"), list) else []
                tickers = summary.get("tickers") if isinstance(summary.get("tickers"), list) else []
                mover_count = report.get("mover_count")
                if mover_count is None and movers:
                    mover_count = len(movers)
                ticker_count = report.get("ticker_count")
                if ticker_count is None and tickers:
                    ticker_count = len(tickers)
                bits: list[str] = []
                if mover_count is not None:
                    bits.append(f"{mover_count} mover(s)")
                if ticker_count is not None:
                    bits.append(f"{ticker_count} ticker(s) scanned")
                if status:
                    bits.append(f"status: {status}")
                _append_source_item(
                    items,
                    seen,
                    {
                        "title": f"Market update {run_date}" if run_date else "Market update",
                        "source_id": f"market_update:{run_date or id(report)}",
                        "score": 1.0,
                        "snippet": "; ".join(bits) if bits else None,
                        "published_at": run_date or None,
                    },
                    default_title="Market update",
                    kind="context",
                )

        elif ev_type == "watchlist":
            rows = (
                details_payload
                if isinstance(details_payload, list)
                else (
                    details.get("items")
                    if isinstance(details.get("items"), list)
                    else []
                )
            )
            for row in rows:
                ticker = ""
                added_at = None
                if isinstance(row, dict):
                    ticker = str(row.get("ticker") or "").strip().upper()
                    added_at = row.get("added_at")
                elif isinstance(row, str):
                    ticker = row.strip().upper()
                if not ticker:
                    continue
                added_text = str(added_at or "").strip()
                _append_source_item(
                    items,
                    seen,
                    {
                        "title": f"{ticker} watchlist",
                        "source_id": f"watchlist:{ticker}",
                        "snippet": (
                            f"Added: {added_text[:10]}" if added_text else "Tracked in watchlist."
                        ),
                    },
                    default_title="Watchlist item",
                    kind="context",
                )

        elif ev.get("tool") and not ev.get("type"):
            # Agent loop evidence format: {tool, arguments, result}
            # Handle search_news and gather_local_context tool results.
            tool_name = str(ev.get("tool") or "")
            result = ev.get("result") or {}
            if isinstance(result, dict):
                result = _decode_truncated_tool_result(result)
                if tool_name == "search_news":
                    hits = _dict_rows(result.get("hits"))
                    for hit in hits:
                        if isinstance(hit, dict):
                            _append_source_item(
                                items,
                                seen,
                                hit,
                                default_title="News article",
                                kind="news",
                            )
                    if not hits:
                        _append_news_no_hit_source(items, seen, result)
                elif tool_name == "search_announcements":
                    for row in result.get("documents") or []:
                        if isinstance(row, dict):
                            _append_source_item(
                                items,
                                seen,
                                row,
                                default_title="Announcement document",
                                kind="document",
                            )
                    for row in result.get("context") or []:
                        if isinstance(row, dict):
                            _append_source_item(
                                items,
                                seen,
                                row,
                                default_title="Announcement excerpt",
                                kind="document",
                            )
                elif tool_name in ("gather_local_context", "query_ticker_data"):
                    for hit in result.get("hits") or result.get("rag_hits") or []:
                        if isinstance(hit, dict):
                            _append_source_item(
                                items,
                                seen,
                                hit,
                                default_title="Context source",
                                kind="rag",
                            )
                    for row in result.get("docs") or []:
                        if isinstance(row, dict):
                            _append_source_item(
                                items,
                                seen,
                                row,
                                default_title="Document",
                                kind="document",
                            )
                elif tool_name == "get_financials":
                    for row in result.get("financials") or []:
                        if isinstance(row, dict):
                            _append_source_item(
                                items,
                                seen,
                                {
                                    **row,
                                    "title": (
                                        f"{row.get('ticker') or 'Financials'} "
                                        f"{row.get('period_type') or ''} "
                                        f"{row.get('period_end') or ''}"
                                    ).strip(),
                                    "document_id": row.get("source_document_id"),
                                    "published_at": row.get("period_end"),
                                    "doc_type": row.get("period_type"),
                                    "snippet": result.get("narrative") or None,
                                },
                                default_title="Financial period",
                                kind="document",
                            )
                elif tool_name == "recall_dossier":
                    for row in result.get("findings") or []:
                        if isinstance(row, dict):
                            _append_source_item(
                                items,
                                seen,
                                {
                                    **row,
                                    "title": (
                                        str(row.get("source") or "").strip()
                                        or str(row.get("category") or "").strip()
                                        or "Dossier finding"
                                    ),
                                    "url": row.get("source_url"),
                                    "snippet": row.get("finding_with_age") or row.get("finding"),
                                    "published_at": row.get("ts"),
                                    "score": row.get("confidence"),
                                },
                                default_title="Dossier finding",
                                kind="context",
                            )
                elif tool_name == "deep_research":
                    research = result.get("research")
                    if isinstance(research, dict):
                        _append_source_item(
                            items,
                            seen,
                            {
                                "title": "Deep research brief",
                                "source_id": f"deep_research:{result.get('ticker') or ''}",
                                "snippet": research.get("summary"),
                                "score": research.get("confidence"),
                            },
                            default_title="Deep research brief",
                            kind="context",
                        )
                elif tool_name == "search_web":
                    for row in result.get("results") or result.get("pages") or []:
                        if isinstance(row, dict):
                            _append_source_item(
                                items,
                                seen,
                                row,
                                default_title="Web result",
                                kind="web",
                            )
                elif tool_name == "fetch_url":
                    _append_source_item(
                        items,
                        seen,
                        result,
                        default_title="Fetched page",
                        kind="web",
                    )
                elif tool_name == "get_data_quality":
                    for row in result.get("recent_failures") or []:
                        if isinstance(row, dict):
                            _append_source_item(
                                items,
                                seen,
                                row,
                                default_title="Extraction failure",
                                kind="document",
                            )
                    for row in result.get("recent_low_conf_rows") or []:
                        if isinstance(row, dict):
                            _append_source_item(
                                items,
                                seen,
                                {
                                    **row,
                                    "title": (
                                        f"{row.get('ticker') or 'Low-confidence'} "
                                        f"{row.get('period_type') or ''} "
                                        f"{row.get('period_end') or ''}"
                                    ).strip(),
                                    "document_id": row.get("source_document_id"),
                                    "published_at": row.get("period_end"),
                                    "score": row.get("confidence_metrics"),
                                    "snippet": (
                                        f"Confidence {row.get('confidence_metrics')}"
                                        if row.get("confidence_metrics") is not None
                                        else None
                                    ),
                                },
                                default_title="Low-confidence financial",
                                kind="context",
                            )
                elif tool_name == "run_analysis":
                    for row in result.get("modules") or []:
                        if not isinstance(row, dict):
                            continue
                        metrics = row.get("metrics") if isinstance(row.get("metrics"), dict) else {}
                        metric_bits = [
                            f"{key}: {value}"
                            for key, value in metrics.items()
                            if value not in (None, "")
                        ]
                        snippet = str(row.get("narrative") or "").strip()
                        if not snippet and metric_bits:
                            snippet = ", ".join(metric_bits)
                        _append_source_item(
                            items,
                            seen,
                            {
                                "title": f"{row.get('module') or 'Analysis'} analysis",
                                "source_id": (
                                    f"analysis:{result.get('ticker') or ''}:{row.get('module') or ''}"
                                ),
                                "snippet": snippet or None,
                                "score": 1.0 if row.get("status") == "complete" else 0.0,
                            },
                            default_title="Analysis result",
                            kind="context",
                        )
                elif tool_name == "get_price":
                    price = result.get("price") if isinstance(result.get("price"), dict) else {}
                    current = price.get("current") if isinstance(price.get("current"), dict) else {}
                    _append_source_item(
                        items,
                        seen,
                        {
                            "title": f"{result.get('ticker') or 'Ticker'} price data",
                            "source_id": (
                                f"price:{result.get('ticker') or ''}:{price.get('range') or ''}:{price.get('interval') or ''}"
                            ),
                            "snippet": (
                                f"Provider: {price.get('provider') or 'unknown'}. "
                                f"Market time: {current.get('market_time') or 'unknown'}."
                            ),
                        },
                        default_title="Price data",
                        kind="context",
                    )
                elif tool_name == "get_price_on_date":
                    _append_source_item(
                        items,
                        seen,
                        {
                            "title": (
                                f"{result.get('ticker') or 'Ticker'} price on "
                                f"{result.get('date') or 'requested date'}"
                            ),
                            "source_id": (
                                f"price_on_date:{result.get('ticker') or ''}:{result.get('date') or ''}"
                            ),
                            "snippet": (
                                f"Open {result.get('open')}, high {result.get('high')}, "
                                f"low {result.get('low')}, close {result.get('close')}."
                            ),
                        },
                        default_title="Historical price",
                        kind="context",
                    )
                elif tool_name == "get_price_range":
                    _append_source_item(
                        items,
                        seen,
                        {
                            "title": (
                                f"{result.get('ticker') or 'Ticker'} price range "
                                f"{result.get('start_date') or ''} to {result.get('end_date') or ''}"
                            ).strip(),
                            "source_id": (
                                f"price_range:{result.get('ticker') or ''}:{result.get('start_date') or ''}:{result.get('end_date') or ''}"
                            ),
                            "snippet": f"{result.get('data_points') or 0} price observations returned.",
                        },
                        default_title="Price range",
                        kind="context",
                    )
                elif tool_name == "search_social":
                    for row in result.get("stories") or result.get("results") or []:
                        if isinstance(row, dict):
                            _append_source_item(
                                items,
                                seen,
                                row,
                                default_title="Social result",
                                kind="web",
                            )
                elif tool_name == "get_watchlist_alerts":
                    alerts = result.get("alerts") if isinstance(result.get("alerts"), list) else []
                    for row in alerts:
                        if isinstance(row, dict):
                            _append_source_item(
                                items,
                                seen,
                                {
                                    **row,
                                    "title": (
                                        f"{row.get('ticker') or 'Watchlist'} "
                                        f"{row.get('type') or 'alert'}"
                                    ).strip(),
                                    "source_id": row.get("id"),
                                    "snippet": row.get("message"),
                                    "published_at": row.get("ts"),
                                },
                                default_title="Watchlist alert",
                                kind="context",
                            )
                    if not alerts:
                        ticker = str(result.get("ticker") or "watchlist").strip().upper()
                        since_hours = result.get("since_hours")
                        snippet = (
                            f"No alerts returned in the last {since_hours} hour(s)."
                            if isinstance(since_hours, (int, float))
                            else "No alerts returned for the current watchlist scan."
                        )
                        _append_source_item(
                            items,
                            seen,
                            {
                                "title": f"{ticker or 'Watchlist'} alerts",
                                "source_id": f"watchlist_alerts:{ticker or 'all'}",
                                "snippet": snippet,
                            },
                            default_title="Watchlist alerts",
                            kind="context",
                        )
                elif tool_name == "tv_screener":
                    market = str(result.get("market") or "").strip().upper()
                    screener_results = (
                        result.get("results") if isinstance(result.get("results"), list) else []
                    )
                    for index, row in enumerate(screener_results):
                        if not isinstance(row, dict):
                            continue
                        symbol = str(
                            row.get("symbol")
                            or row.get("ticker")
                            or row.get("code")
                            or row.get("name")
                            or ""
                        ).strip()
                        _append_source_item(
                            items,
                            seen,
                            {
                                **row,
                                "title": (
                                    f"{symbol or 'Market mover'}"
                                    f"{f' ({market})' if market else ''}"
                                ),
                                "source_id": f"tv_screener:{market}:{symbol or index}",
                                "snippet": _summarize_scalar_fields(row, max_items=5),
                            },
                            default_title="TradingView screener",
                            kind="context",
                        )
                    if not screener_results:
                        count = result.get("count")
                        snippet = (
                            f"Screener returned {count} rows."
                            if isinstance(count, (int, float))
                            else "Screener returned no rows for this query."
                        )
                        _append_source_item(
                            items,
                            seen,
                            {
                                "title": (
                                    f"TradingView screener ({market})"
                                    if market
                                    else "TradingView screener"
                                ),
                                "source_id": f"tv_screener:{market or 'unknown'}",
                                "snippet": snippet,
                            },
                            default_title="TradingView screener",
                            kind="context",
                        )
                elif tool_name == "get_tv_indicators":
                    ticker = str(result.get("ticker") or "").strip().upper()
                    exchange = str(result.get("exchange") or "").strip().upper()
                    indicators = (
                        result.get("indicators")
                        if isinstance(result.get("indicators"), dict)
                        else {}
                    )
                    indicator_bits: list[str] = []
                    for name, value in indicators.items():
                        if isinstance(value, dict):
                            err = str(value.get("error") or "").strip()
                            if err:
                                indicator_bits.append(f"{name}: error ({err})")
                        elif value is not None:
                            indicator_bits.append(f"{name}: {value}")
                        if len(indicator_bits) >= 6:
                            break
                    _append_source_item(
                        items,
                        seen,
                        {
                            "title": f"{exchange + ':' if exchange else ''}{ticker or 'Ticker'} indicators",
                            "source_id": f"tv_indicators:{exchange}:{ticker}",
                            "snippet": _clean_source_text("; ".join(indicator_bits)),
                        },
                        default_title="TradingView indicators",
                        kind="context",
                    )
                elif tool_name == "check_youtube_channel_recent_videos":
                    _append_youtube_recent_video_sources(items, seen, result)

        if len(items) >= 10:
            break

    return items


_NON_SUBSTANTIVE_CHAT_MESSAGE_RE = re.compile(
    r"^\s*(?:"
    r"/[a-z_][\w-]*.*|"
    r"hi|hello|hey|yo|sup|"
    r"good (?:morning|afternoon|evening)|"
    r"thanks|thank you|"
    r"ok(?:ay)?|yes|no|sure|cool|continue|go on|"
    r"help(?: me)?|"
    r"what can you do\??|"
    r"show sources|show source|sources\??"
    r")\s*$",
    re.IGNORECASE,
)
_EXPLICIT_UNVERIFIED_RESPONSE_RE = re.compile(
    r"\b(?:cannot|can't|can not|do not|don't|won't)\s+"
    r"(?:verify|confirm|substantiate|make factual claims)\b|"
    r"\bnot enough (?:evidence|sources|current evidence|retrieved context|information)\b|"
    r"\bunable to verify\b",
    re.IGNORECASE,
)
_CONTAINS_FINANCIAL_CLAIM_RE = re.compile(
    r"(?-i:\b[A-Z]{2,5}\b)|"                               # ASX tickers / company abbreviations
    r"\$[\d,]+(?:\.\d+)?[MBKmb]?\b|"                      # dollar amounts
    r"\b\d+(?:\.\d+)?%\b|"                                 # percentages
    r"\b(?:announced|reported|upgraded|downgraded|raised|cut|beat|missed|"
    r"acquired|merged|divested|appointed|resigned|flagged|guided|earnings|"
    r"revenue|profit|loss|EBIT|EBITDA|dividend|buyback|placement)\b",
    re.IGNORECASE,
)
_PURE_OPERATIONAL_NO_HIT_RE = re.compile(
    r"\bno (?:news )?(?:hits|results|articles|videos?) (?:were )?(?:returned|found)\b|"
    r"\breturned no (?:news )?(?:hits|results|articles|videos?)\b",
    re.IGNORECASE,
)
_SOURCE_CONTRACT_REFUSAL = (
    "I can't verify that from current evidence, and I won't make factual claims unless "
    "the supporting sources can be shown in the Sources dropdown. Please narrow the "
    "question or ask me to fetch the relevant news, announcements, financials, or price data first."
)
_COMMANDS_REQUIRING_SOURCES = (
    "/market-update",
    "/watch scan",
    "/alerts",
)
_OPERATIONAL_COMMAND_TOOLS_WITHOUT_VISIBLE_SOURCES = frozenset(
    {
        "watch_youtube_channel",
    }
)


def _message_requires_visible_sources(message: str) -> bool:
    text = str(message or "").strip()
    if not text:
        return False
    explicit_command = text.lower()
    if explicit_command.startswith("/"):
        return explicit_command.startswith(_COMMANDS_REQUIRING_SOURCES)
    rewritten = derive_conversational_command(text)
    if rewritten:
        command = rewritten.lower()
        return command.startswith(_COMMANDS_REQUIRING_SOURCES)
    if _NON_SUBSTANTIVE_CHAT_MESSAGE_RE.fullmatch(text):
        return False
    return True


def _is_operational_command_result(response: Any) -> bool:
    if str(getattr(response, "mode", "") or "").strip() != "command":
        return False
    evidence = getattr(response, "evidence", None) or []
    for entry in evidence:
        if not isinstance(entry, dict):
            continue
        tool = str(entry.get("tool") or "").strip()
        if tool in _OPERATIONAL_COMMAND_TOOLS_WITHOUT_VISIBLE_SOURCES:
            return True
    return False


def _only_operational_no_hit_sources(sources: list[dict[str, Any]]) -> bool:
    if not sources:
        return False
    for source in sources:
        source_id = str(source.get("source_id") or "")
        doc_type = str(source.get("doc_type") or "")
        if not source_id.startswith("search_news:no_hits:") and doc_type != "operational_no_hit":
            return False
    return True


def _enforce_visible_source_contract(message: str, response: Any) -> list[dict[str, Any]]:
    sources = _build_ui_sources(getattr(response, "evidence", None) or [])
    text = str(getattr(response, "text", "") or "").strip()

    if not text or getattr(response, "action_preview", None) is not None:
        return sources
    if _is_operational_command_result(response):
        return sources
    if not _message_requires_visible_sources(message):
        return sources
    if sources:
        if (
            _only_operational_no_hit_sources(sources)
            and _CONTAINS_FINANCIAL_CLAIM_RE.search(text)
            and not _PURE_OPERATIONAL_NO_HIT_RE.search(text)
        ):
            sources = []
        else:
            return sources
    # Only allow the "explicit unverified" bypass when the response is a PURE
    # statement of inability — no named tickers, monetary figures, events, or
    # company-specific claims. A hedged hallucination ("the evidence is
    # incomplete, but BHP reported...") must still be blocked.
    if _EXPLICIT_UNVERIFIED_RESPONSE_RE.search(text) and not _CONTAINS_FINANCIAL_CLAIM_RE.search(text):
        return sources

    meta = dict(getattr(response, "routing_metadata", None) or {})
    meta["grounding_guard"] = "missing_visible_sources"
    # Build a tool audit so the UI can surface "Searched X: 0 results"
    # rather than a completely empty sources panel.
    evidence = getattr(response, "evidence", None) or []
    tool_audit = []
    for ev in evidence:
        if not isinstance(ev, dict):
            continue
        tool = ev.get("tool", "")
        result = ev.get("result", {})
        if not isinstance(result, dict):
            continue
        hit_count = result.get("hit_count") or len(result.get("hits", [])) or len(result.get("results", []))
        audit_entry: dict[str, Any] = {"tool": tool, "hit_count": hit_count}
        if result.get("error"):
            audit_entry["error"] = str(result["error"])[:120]
        if tool:
            tool_audit.append(audit_entry)
    if tool_audit:
        meta["tool_audit"] = tool_audit
    response.routing_metadata = meta
    response.text = _SOURCE_CONTRACT_REFUSAL
    return []


def _maybe_auto_flag_chat_response(
    service: Any,
    *,
    session_id: str | None,
    ticker: str | None,
    response: Any,
) -> dict[str, Any] | None:
    auto_flag = getattr(service, "auto_flag_chat_response", None)
    if not callable(auto_flag):
        return None
    try:
        return auto_flag(session_id=session_id, ticker=ticker, response=response)
    except Exception:
        logger.exception("Cockpit auto diagnostic flag failed")
        return None


def _serialize_flag_handoff(flag_result: dict[str, Any] | None) -> dict[str, Any] | None:
    if not isinstance(flag_result, dict):
        return None
    keys = (
        "report_id",
        "feedback_type",
        "capture_kind",
        "report_dir",
        "read_api_path",
        "codex_prompt",
        "codex_prompt_path",
        "investigation_path",
        "investigation_status",
        "codex_cli_command",
        "analysis_summary",
    )
    payload = {
        key: flag_result.get(key)
        for key in keys
        if flag_result.get(key) is not None
    }
    return payload or None


# ---------------------------------------------------------------------------
# Helper: probe a single HTTP endpoint
# ---------------------------------------------------------------------------


def _probe_http(
    url: str, path: str, *, timeout: float = 3.0
) -> tuple[bool, float, str | None]:
    """Return (reachable, latency_ms, error_or_none)."""
    target = str(url or "").strip().rstrip("/")
    if not target:
        return False, 0.0, "not configured"
    try:
        start = time.monotonic()
        resp = httpx.get(f"{target}{path}", timeout=timeout)
        elapsed_ms = (time.monotonic() - start) * 1000
        resp.raise_for_status()
        return True, round(elapsed_ms, 1), None
    except Exception as exc:
        return False, 0.0, str(exc)


def _parse_float(raw: str) -> float | None:
    stripped = raw.strip()
    if not stripped or stripped.startswith("["):
        return None
    try:
        return float(stripped)
    except ValueError:
        return None


def _probe_gpu() -> ServiceHealthItem:
    """Return a GPU runtime summary including power, clocks, and memory bandwidth."""
    try:
        start = time.monotonic()
        result = subprocess.run(
            [
                "nvidia-smi",
                "--query-gpu=name,temperature.gpu,utilization.gpu,memory.used,memory.total,"
                "power.draw,power.limit,fan.speed,utilization.memory,clocks.gr,clocks.mem,pstate",
                "--format=csv,noheader,nounits",
            ],
            capture_output=True,
            text=True,
            timeout=3,
            check=False,
        )
        elapsed_ms = round((time.monotonic() - start) * 1000, 1)
    except FileNotFoundError:
        return ServiceHealthItem(
            name="gpu", status="unknown", error="nvidia-smi not installed"
        )
    except subprocess.TimeoutExpired:
        return ServiceHealthItem(
            name="gpu", status="degraded", error="nvidia-smi timed out"
        )
    except Exception as exc:
        return ServiceHealthItem(name="gpu", status="degraded", error=str(exc))

    if result.returncode != 0:
        stderr = (result.stderr or "").strip()
        return ServiceHealthItem(
            name="gpu",
            status="degraded",
            response_time_ms=elapsed_ms,
            error=stderr or f"nvidia-smi exited {result.returncode}",
        )

    lines = [
        line.strip() for line in (result.stdout or "").splitlines() if line.strip()
    ]
    if not lines:
        return ServiceHealthItem(
            name="gpu",
            status="unknown",
            response_time_ms=elapsed_ms,
            error="no GPU devices reported",
        )

    gpus: list[dict[str, Any]] = []
    for line in lines:
        parts = [part.strip() for part in line.split(",")]
        if len(parts) < 5:
            continue
        name = parts[0] or "GPU"
        pstate = parts[11].strip() if len(parts) > 11 else None
        gpus.append(
            {
                "name": name,
                "temp_c": _parse_float(parts[1]) if len(parts) > 1 else None,
                "util_percent": _parse_float(parts[2]) if len(parts) > 2 else None,
                "mem_used_mib": _parse_float(parts[3]) if len(parts) > 3 else None,
                "mem_total_mib": _parse_float(parts[4]) if len(parts) > 4 else None,
                "power_draw_w": _parse_float(parts[5]) if len(parts) > 5 else None,
                "power_limit_w": _parse_float(parts[6]) if len(parts) > 6 else None,
                "fan_speed_pct": _parse_float(parts[7]) if len(parts) > 7 else None,
                "mem_util_percent": _parse_float(parts[8]) if len(parts) > 8 else None,
                "clock_gr_mhz": _parse_float(parts[9]) if len(parts) > 9 else None,
                "clock_mem_mhz": _parse_float(parts[10]) if len(parts) > 10 else None,
                "pstate": pstate if pstate and not pstate.startswith("[") else None,
            }
        )

    if not gpus:
        return ServiceHealthItem(
            name="gpu",
            status="unknown",
            response_time_ms=elapsed_ms,
            error="unable to parse GPU status",
        )

    return ServiceHealthItem(
        name="gpu",
        status="healthy",
        response_time_ms=elapsed_ms,
        details={"gpus": gpus},
    )


def _probe_host() -> ServiceHealthItem:
    """Return CPU, memory, disk, and top-process metrics using psutil."""
    try:
        import psutil

        start = time.monotonic()

        load_1m, load_5m, load_15m = os.getloadavg()
        core_count = psutil.cpu_count(logical=False) or 1
        logical_count = psutil.cpu_count(logical=True) or 1
        normalized_load = round((load_1m / logical_count) * 100, 1)

        mem = psutil.virtual_memory()
        swap = psutil.swap_memory()

        disk_data: list[dict[str, Any]] = []
        for part in psutil.disk_partitions(all=False):
            try:
                usage = psutil.disk_usage(part.mountpoint)
                disk_data.append(
                    {
                        "mount": part.mountpoint,
                        "total_gib": round(usage.total / 1024**3, 2),
                        "used_gib": round(usage.used / 1024**3, 2),
                        "used_percent": round(usage.percent, 1),
                    }
                )
            except (PermissionError, OSError):
                continue

        top_procs: list[dict[str, Any]] = []
        try:
            procs = []
            for proc in psutil.process_iter(["pid", "name", "cmdline", "cpu_percent", "memory_percent", "memory_info"]):
                try:
                    info = proc.info
                    procs.append(info)
                except (psutil.NoSuchProcess, psutil.AccessDenied):
                    continue
            procs.sort(key=lambda p: float(p.get("cpu_percent") or 0), reverse=True)
            for info in procs[:10]:
                rss_mib = None
                mem_info = info.get("memory_info")
                if mem_info is not None:
                    rss_mib = round(mem_info.rss / 1024**2, 1)
                cmdline = info.get("cmdline") or []
                top_procs.append(
                    {
                        "pid": info.get("pid"),
                        "command_name": info.get("name") or "unknown",
                        "command": " ".join(str(c) for c in cmdline[:8]) if cmdline else (info.get("name") or ""),
                        "cpu_percent": round(float(info.get("cpu_percent") or 0), 1),
                        "mem_percent": round(float(info.get("memory_percent") or 0), 2),
                        "rss_mib": rss_mib,
                    }
                )
        except Exception:
            pass

        elapsed_ms = round((time.monotonic() - start) * 1000, 1)
        return ServiceHealthItem(
            name="host",
            status="healthy",
            response_time_ms=elapsed_ms,
            details={
                "cpu": {
                    "core_count": core_count,
                    "logical_count": logical_count,
                    "load_1m": round(load_1m, 2),
                    "load_5m": round(load_5m, 2),
                    "load_15m": round(load_15m, 2),
                    "normalized_load_percent": normalized_load,
                },
                "memory": {
                    "total_gib": round(mem.total / 1024**3, 2),
                    "used_gib": round(mem.used / 1024**3, 2),
                    "available_gib": round(mem.available / 1024**3, 2),
                    "used_percent": round(mem.percent, 1),
                    "swap_total_gib": round(swap.total / 1024**3, 2),
                    "swap_used_gib": round(swap.used / 1024**3, 2),
                    "swap_used_percent": round(swap.percent, 1),
                },
                "disks": disk_data,
                "top_processes": top_procs,
            },
        )
    except ImportError:
        return ServiceHealthItem(name="host", status="unknown", error="psutil not available")
    except Exception as exc:
        return ServiceHealthItem(name="host", status="degraded", error=str(exc))


# ---------------------------------------------------------------------------
# GET /api/cockpit/health
# ---------------------------------------------------------------------------


@router.get("/health", response_model=AggregatedHealthResponse)
def cockpit_health() -> AggregatedHealthResponse:
    """Aggregated health check probing backend, llama.cpp, Ollama, Qdrant, and Redis."""
    services: list[ServiceHealthItem] = []

    # 1. Backend (self-check — always healthy if this code runs)
    services.append(
        ServiceHealthItem(
            name="backend",
            status="healthy",
            endpoint="http://localhost:8000",
        )
    )

    # 2. llama.cpp
    llamacpp_url = str(settings.llamacpp_url or "").strip().rstrip("/")
    ok, latency, err = _probe_http(llamacpp_url, "/v1/models")
    services.append(
        ServiceHealthItem(
            name="llamacpp",
            status="healthy" if ok else "down",
            endpoint=llamacpp_url or None,
            response_time_ms=latency if ok else None,
            error=err,
        )
    )

    # 3. Ollama
    ollama_url = str(settings.ollama_url or "").strip().rstrip("/")
    if ollama_url:
        ok, latency, err = _probe_http(ollama_url, "/api/tags")
        services.append(
            ServiceHealthItem(
                name="ollama",
                status="healthy" if ok else "down",
                endpoint=ollama_url,
                response_time_ms=latency if ok else None,
                error=err,
            )
        )
    else:
        services.append(
            ServiceHealthItem(
                name="ollama",
                status="unknown",
                error="not configured",
            )
        )

    # 4. Qdrant
    qdrant_url = str(settings.qdrant_url or "").strip().rstrip("/")
    if settings.enable_qdrant and qdrant_url:
        ok, latency, err = _probe_http(qdrant_url, "/collections")
        services.append(
            ServiceHealthItem(
                name="qdrant",
                status="healthy" if ok else "down",
                endpoint=qdrant_url,
                response_time_ms=latency if ok else None,
                error=err,
            )
        )
    else:
        services.append(
            ServiceHealthItem(
                name="qdrant",
                status="unknown",
                error="disabled" if not settings.enable_qdrant else "not configured",
            )
        )

    # 5. Redis
    redis_ok = False
    redis_err: str | None = None
    try:
        import socket as _socket

        parsed = urlparse(str(settings.celery_broker_url or ""))
        host = str(parsed.hostname or "127.0.0.1").strip()
        port = int(parsed.port or 6379)
        start = time.monotonic()
        with _socket.create_connection((host, port), timeout=2.0):
            redis_latency = round((time.monotonic() - start) * 1000, 1)
            redis_ok = True
    except Exception as exc:
        redis_latency = 0.0
        redis_err = str(exc)

    services.append(
        ServiceHealthItem(
            name="redis",
            status="healthy" if redis_ok else "down",
            endpoint=str(settings.celery_broker_url or None),
            response_time_ms=redis_latency if redis_ok else None,
            error=redis_err,
        )
    )

    services.append(_probe_gpu())
    services.append(_probe_host())

    # 8. CockpitService initialization
    cs_ok = False
    cs_latency: float = 0.0
    cs_err: str | None = None
    try:
        cs_start = time.monotonic()
        CockpitService.get_instance()
        cs_latency = round((time.monotonic() - cs_start) * 1000, 1)
        cs_ok = True
    except Exception as exc:
        cs_err = str(exc)

    services.append(
        ServiceHealthItem(
            name="cockpit_service",
            status="healthy" if cs_ok else "down",
            response_time_ms=cs_latency if cs_ok else None,
            error=cs_err,
        )
    )

    # Derive overall status
    statuses = [s.status for s in services]
    if all(s == "healthy" for s in statuses):
        overall = "healthy"
    elif any(s == "down" for s in statuses if s != "unknown"):
        overall = "degraded"
    else:
        overall = "healthy"

    return AggregatedHealthResponse(status=overall, services=services)


# ---------------------------------------------------------------------------
# GET /api/cockpit/metrics/gpu   — fast GPU-only poll (no HTTP probes)
# GET /api/cockpit/metrics/host  — fast host-only poll (no HTTP probes)
# ---------------------------------------------------------------------------


@router.get("/metrics/gpu")
def cockpit_metrics_gpu() -> dict[str, Any]:
    """Fast GPU-only metrics endpoint for the dialog live-poll loop."""
    item = _probe_gpu()
    return {
        "status": item.status,
        "error": item.error,
        "response_time_ms": item.response_time_ms,
        "details": item.details or {},
    }


@router.get("/metrics/host")
def cockpit_metrics_host() -> dict[str, Any]:
    """Fast host-only metrics endpoint for the dialog live-poll loop."""
    item = _probe_host()
    return {
        "status": item.status,
        "error": item.error,
        "response_time_ms": item.response_time_ms,
        "details": item.details or {},
    }


# ---------------------------------------------------------------------------
# GET /api/cockpit/config
# ---------------------------------------------------------------------------


def _git_branch() -> str | None:
    try:
        result = subprocess.run(
            ["git", "rev-parse", "--abbrev-ref", "HEAD"],
            cwd=PROJECT_ROOT.parent,
            capture_output=True,
            text=True,
            timeout=2,
            check=False,
        )
        if result.returncode == 0 and result.stdout:
            return result.stdout.strip()
    except (FileNotFoundError, subprocess.TimeoutExpired):
        pass
    return None


@router.get("/config", response_model=CockpitConfigResponse)
def cockpit_config() -> CockpitConfigResponse:
    """Return system configuration for the cockpit settings screen."""
    from app.services.llamacpp_runtime import resolve_llm_runtime_config

    load_env(PROJECT_ROOT)

    llm_endpoint: str | None = None
    llm_model: str | None = None
    try:
        llm_url, llm_mdl = resolve_llm_runtime_config()
        llm_endpoint = llm_url
        llm_model = llm_mdl
    except Exception:
        llm_endpoint = str(settings.llamacpp_url or "").strip() or None
        llm_model = None

    server_models = _fetch_llama_server_models()
    preferred_loaded = _pick_preferred_loaded_model_id(server_models)
    if preferred_loaded:
        llm_model = preferred_loaded

    config_path_value = str(
        os.getenv("COCKPIT_CONFIG") or "config/cockpit.yaml"
    ).strip()
    config_path = Path(config_path_value)
    if not config_path.is_absolute():
        config_path = (PROJECT_ROOT / config_path).resolve()

    effective_cfg = compute_effective_cockpit_config(
        PROJECT_ROOT,
        str(config_path),
        profile=str(os.getenv("COCKPIT_PROFILE") or "default").strip() or "default",
        read_only=False,
        no_web=False,
    )
    cockpit_llm = (
        effective_cfg.get("cockpit_llm")
        if isinstance(effective_cfg.get("cockpit_llm"), dict)
        else {}
    )
    backend_cfg = (
        effective_cfg.get("backend")
        if isinstance(effective_cfg.get("backend"), dict)
        else {}
    )
    runtime_cfg = (
        effective_cfg.get("runtime")
        if isinstance(effective_cfg.get("runtime"), dict)
        else {}
    )
    anthropic_key_configured = bool(effective_anthropic_api_key(cockpit_llm))
    extraction_activity = get_extraction_activity_snapshot()

    return CockpitConfigResponse(
        llm_model=llm_model,
        llm_endpoint=llm_endpoint,
        anthropic_key_configured=anthropic_key_configured,
        extraction_active=bool(extraction_activity.get("active")),
        extraction_activity_source=str(extraction_activity.get("source") or "none"),
        extraction_activity_expires_in_seconds=int(
            extraction_activity.get("expires_in_seconds") or 0
        ),
        extraction_active_runs=[
            CockpitConfigResponse.ExtractionActivityRun(
                token=str(run.get("token") or "").strip(),
                run_id=str(run.get("run_id") or "").strip() or None,
                document_id=str(run.get("document_id") or "").strip() or None,
                requested_method=str(run.get("requested_method") or "").strip() or None,
                strict_method=(
                    bool(run.get("strict_method"))
                    if run.get("strict_method") is not None
                    else None
                ),
                ticker=str(run.get("ticker") or "").strip() or None,
                title=str(run.get("title") or "").strip() or None,
                expires_at=(
                    float(run.get("expires_at"))
                    if run.get("expires_at") is not None
                    else None
                ),
                expires_in_seconds=(
                    int(run.get("expires_in_seconds"))
                    if run.get("expires_in_seconds") is not None
                    else None
                ),
            )
            for run in (extraction_activity.get("active_runs") or [])
            if str(run.get("token") or "").strip()
        ],
        extract_model=str(settings.extract_model or "").strip() or None,
        embed_model=str(settings.embed_model or "").strip() or None,
        routing_policy=str(cockpit_llm.get("hybrid_router_policy") or "").strip() or None,
        backend_url=str(backend_cfg.get("api_base_url") or "").strip() or None,
        profile=(
            str(
                cockpit_llm.get("llm_profile_label")
                or runtime_cfg.get("profile")
                or os.environ.get("LOCAL_BACKEND_PROFILE")
                or ""
            ).strip()
            or None
        ),
        features=_effective_cockpit_feature_flags(effective_cfg),
        python_version=f"{sys.version_info.major}.{sys.version_info.minor}.{sys.version_info.micro}",
        git_branch=_git_branch(),
        data_root=str(settings.data_root or "").strip() or None,
    )


# ---------------------------------------------------------------------------
# GET /api/cockpit/models
# ---------------------------------------------------------------------------


def _parse_quantization(filename: str) -> str | None:
    """Extract quantization tag from GGUF filename (e.g. 'Q4_K_M' from '...-Q4_K_M.gguf')."""
    stem = filename.rsplit(".", 1)[0] if "." in filename else filename
    parts = stem.split("-")
    for part in reversed(parts):
        upper = part.upper()
        if upper.startswith("Q") and any(c.isdigit() for c in upper):
            return upper
        if upper in ("MXFP4", "FP16", "BF16", "F16", "F32"):
            return upper
    return None


def _extract_model_path(status_obj: dict[str, Any]) -> str:
    args_list = status_obj.get("args") or []
    for i, arg in enumerate(args_list):
        if arg == "--model" and i + 1 < len(args_list):
            return str(args_list[i + 1] or "").strip()

    preset_text = str(status_obj.get("preset") or "")
    for line in preset_text.splitlines():
        if "=" not in line:
            continue
        key, value = line.split("=", 1)
        if key.strip().lower() == "model":
            return value.strip()
    return ""


def _path_location_key(path_text: str) -> str:
    return path_text.replace("\\", "/").rstrip("/").lower()


def _classify_model_location(model_path: str) -> dict[str, str] | None:
    normalized = _path_location_key(model_path)
    if not normalized:
        return None

    for loc in _MODEL_LOCATIONS:
        dir_path = os.environ.get(loc["env_key"], "").strip() if loc["env_key"] else ""
        if not dir_path:
            dir_path = loc["default"]
        if dir_path and normalized.startswith(_path_location_key(dir_path) + "/"):
            return loc

    if "/.cache/llmfit/models/" in normalized:
        return next(loc for loc in _MODEL_LOCATIONS if loc["location"] == "ssd")
    if "/cold_storage/models/" in normalized:
        return next(loc for loc in _MODEL_LOCATIONS if loc["location"] == "hdd")
    return None


def _choose_preferred_model_id(path_stem: str, server_model_ids: list[str]) -> str:
    norm_stem = path_stem.strip().lower()
    if not server_model_ids:
        return path_stem

    normalized_ids = [
        model_id for model_id in server_model_ids if str(model_id).strip()
    ]
    alias_ids = [
        model_id for model_id in normalized_ids if model_id.startswith("model:")
    ]
    for model_id in alias_ids:
        short = model_id.split(":", 1)[1].strip().lower()
        if short and (
            norm_stem == short
            or norm_stem.startswith(short)
            or short.startswith(norm_stem)
        ):
            return model_id

    for model_id in normalized_ids:
        if model_id.strip().lower() == norm_stem:
            return model_id
    return path_stem


def _build_server_model_groups(
    server_models: dict[str, dict[str, Any]],
) -> list[ModelGroup]:
    grouped: dict[str, list[ModelInfo]] = {
        loc["location"]: [] for loc in _MODEL_LOCATIONS
    }
    by_path: dict[str, dict[str, Any]] = {}

    for model_id, info in server_models.items():
        model_path = str(info.get("model_path") or "").strip()
        if not model_path:
            continue
        location = _classify_model_location(model_path)
        if location is None:
            continue

        normalized_path = _path_location_key(model_path)
        entry = by_path.setdefault(
            normalized_path,
            {
                "path": model_path,
                "filename": Path(model_path).name,
                "stem": Path(model_path).stem,
                "location": location,
                "server_ids": set(),
            },
        )
        entry["server_ids"].add(model_id)

    all_server_ids = list(server_models.keys())
    for entry in sorted(
        by_path.values(), key=lambda item: str(item["filename"]).lower()
    ):
        filename = str(entry["filename"])
        path_stem = str(entry["stem"])
        model_path = str(entry["path"])
        model_file = Path(model_path)
        size_gb = 0.0
        if model_file.is_file():
            size_gb = round(model_file.stat().st_size / (1024**3), 1)

        location = entry["location"]
        grouped[location["location"]].append(
            ModelInfo(
                id=_choose_preferred_model_id(
                    path_stem,
                    list(entry["server_ids"]) + all_server_ids,
                ),
                filename=filename,
                size_gb=size_gb,
                quantization=_parse_quantization(filename),
                available=location["location"] != "hdd",
            )
        )

    result: list[ModelGroup] = []
    for loc in _MODEL_LOCATIONS:
        models = grouped.get(loc["location"], [])
        if models:
            result.append(
                ModelGroup(
                    location=loc["location"],
                    label=loc["label"],
                    models=models,
                )
            )
    return result


def _scan_model_directory(dir_path: str) -> list[ModelInfo]:
    """Scan a directory for .gguf files and return ModelInfo list."""
    results: list[ModelInfo] = []
    p = Path(dir_path)
    if not p.is_dir():
        return results
    for f in sorted(p.glob("*.gguf")):
        if not f.is_file():
            continue
        size_bytes = f.stat().st_size
        stem = f.stem
        results.append(
            ModelInfo(
                id=stem,
                filename=f.name,
                size_gb=round(size_bytes / (1024**3), 1),
                quantization=_parse_quantization(f.name),
                available=True,
            )
        )
    return results


_MODEL_LOCATIONS: list[dict[str, str]] = [
    {
        "env_key": "COCKPIT_MODELS_NVME_DIR",
        "default": "/mnt/nvme/tenn/models",
        "label": "NVMe (Fast)",
        "location": "nvme",
    },
    {
        "env_key": "COCKPIT_MODELS_SSD_DIR",
        "default": str(Path.home() / ".cache" / "llmfit" / "models"),
        "label": "SSD Cache",
        "location": "ssd",
    },
    {
        "env_key": "COCKPIT_MODELS_HDD_DIR",
        "default": str(Path.home() / "cold_storage" / "models"),
        "label": "HDD Cold Storage",
        "location": "hdd",
    },
]


def _fetch_llama_server_models() -> dict[str, dict[str, Any]]:
    """Query llama-server /v1/models and return {model_id: {status, path_stem}}."""
    return _fetch_runtime_models()


def _fetch_runtime_models(base_url: str | None = None) -> dict[str, dict[str, Any]]:
    """Query a llama.cpp runtime /v1/models and return {model_id: {status, path_stem}}."""
    llamacpp_url = str(base_url or settings.llamacpp_url or "").strip().rstrip("/")
    if not llamacpp_url:
        return {}
    headers: dict[str, str] = {}
    api_key = (
        str(os.getenv("LLM_API_KEY") or "").strip()
        or str(os.getenv("LLAMA_SERVER_API_KEY") or "").strip()
        or str(os.getenv("LLAMACPP_API_KEY") or "").strip()
    )
    if api_key:
        headers["Authorization"] = f"Bearer {api_key}"
    try:
        resp = httpx.get(f"{llamacpp_url}/v1/models", headers=headers, timeout=3.0)
        resp.raise_for_status()
        data = resp.json()
    except Exception:
        return {}

    result: dict[str, dict[str, Any]] = {}
    for entry in data.get("data", []):
        model_id = str(entry.get("id", "")).strip()
        if not model_id:
            continue
        status_obj = entry.get("status") or {}
        status_val = str(status_obj.get("value", "unknown"))
        model_path = _extract_model_path(status_obj)
        result[model_id] = {
            "status": status_val,
            "model_path": model_path,
            "path_stem": Path(model_path).stem if model_path else "",
        }
    return result


def _normalize_model_identifier(value: str | None) -> str:
    return str(value or "").strip().removeprefix("model:").lower()


def _model_alias_tokens(model_id: str, info: dict[str, Any] | None = None) -> set[str]:
    tokens = {_normalize_model_identifier(model_id)}
    if str(model_id).startswith("model:"):
        tokens.add(_normalize_model_identifier(str(model_id).split(":", 1)[1]))
    if info:
        tokens.add(_normalize_model_identifier(str(info.get("path_stem") or "").strip()))
    return {token for token in tokens if token}


def _find_matching_runtime_model(
    runtime_models: dict[str, dict[str, Any]],
    requested_model: str,
) -> str | None:
    requested_tokens = _model_alias_tokens(requested_model)
    if not requested_tokens:
        return None

    preferred_match: str | None = None
    for model_id, info in runtime_models.items():
        matches = any(
            req == token or req.startswith(token) or token.startswith(req)
            for req in requested_tokens
            for token in _model_alias_tokens(model_id, info)
        )
        if not matches:
            continue
        if str(model_id).startswith("model:"):
            return model_id
        if preferred_match is None:
            preferred_match = model_id

    return preferred_match


@router.get("/models", response_model=AvailableModelsResponse)
def cockpit_available_models() -> AvailableModelsResponse:
    """Return all discoverable GGUF models grouped by storage location.

    Model IDs are resolved from llama-server's registry so the UI sends
    exactly the ID that llama-server expects in chat requests.
    """
    server_models = _fetch_llama_server_models()
    all_server_ids = list(server_models.keys())

    # Build filename_stem → preferred server model ID lookup
    stem_to_server_id: dict[str, str] = {}
    active_model = _pick_preferred_loaded_model_id(server_models)
    for model_id, info in server_models.items():
        stem = info.get("path_stem", "")
        if stem:
            # Prefer model:alias form over bare filename stems
            existing = stem_to_server_id.get(stem, "")
            if not existing or model_id.startswith("model:"):
                stem_to_server_id[stem] = model_id

    # Map preset aliases that lack explicit --model paths
    for model_id in server_models:
        if model_id.startswith("model:"):
            short = model_id.split(":", 1)[1]
            if short not in stem_to_server_id:
                stem_to_server_id[short] = model_id

    groups: list[ModelGroup] = []
    seen_files: set[str] = set()

    for loc in _MODEL_LOCATIONS:
        dir_path = os.environ.get(loc["env_key"], "").strip() if loc["env_key"] else ""
        if not dir_path:
            dir_path = loc["default"]

        raw_models = _scan_model_directory(dir_path)
        unique_models: list[ModelInfo] = []
        for m in raw_models:
            if m.filename in seen_files:
                continue
            seen_files.add(m.filename)
            server_id = _choose_preferred_model_id(
                m.id,
                [stem_to_server_id.get(m.id, "")] + all_server_ids,
            )
            unique_models.append(
                ModelInfo(
                    id=server_id if server_id else m.id,
                    filename=m.filename,
                    size_gb=m.size_gb,
                    quantization=m.quantization,
                    available=loc["location"] != "hdd",
                )
            )

        if unique_models:
            groups.append(
                ModelGroup(
                    location=loc["location"],
                    label=loc["label"],
                    models=unique_models,
                )
            )

    server_groups = _build_server_model_groups(server_models) if server_models else []
    if not groups:
        groups = server_groups
    elif server_groups:
        local_by_location = {group.location: group for group in groups}
        server_by_location = {group.location: group for group in server_groups}
        merged_groups: list[ModelGroup] = []
        for loc in _MODEL_LOCATIONS:
            group = local_by_location.get(loc["location"]) or server_by_location.get(
                loc["location"]
            )
            if group is not None:
                merged_groups.append(group)
        groups = merged_groups

    groups = _hoist_manual_fallback_model_groups(groups)

    return AvailableModelsResponse(
        groups=groups,
        active_model=active_model,
    )


@router.post("/models/load", response_model=ModelLoadResponse)
def cockpit_load_model(payload: ModelLoadRequest) -> ModelLoadResponse:
    from cockpit.integrations.llamacpp_manager import load_model_api

    runtime_url, default_model = resolve_llm_runtime_config(model=payload.model_id)
    requested_model = str(payload.model_id or default_model or "").strip()
    if not requested_model:
        raise HTTPException(status_code=400, detail="model_id is required")

    runtime_models = _fetch_runtime_models(runtime_url)
    matched_model = _find_matching_runtime_model(runtime_models, requested_model)
    resolved_model = matched_model or requested_model

    if runtime_models and matched_model is None:
        available_models = sorted(runtime_models.keys())
        raise HTTPException(
            status_code=404,
            detail={
                "message": f"Model '{requested_model}' is not available on the configured chat runtime.",
                "requested_model": requested_model,
                "runtime_url": runtime_url,
                "available_models": available_models,
            },
        )

    if any(
        info.get("status") == "loaded"
        and (
            _normalize_model_identifier(resolved_model)
            in _normalize_model_identifier(model_id)
            or _normalize_model_identifier(resolved_model)
            in _normalize_model_identifier(str(info.get("path_stem") or ""))
        )
        for model_id, info in runtime_models.items()
    ):
        return ModelLoadResponse(
            ok=True,
            requested_model=requested_model,
            resolved_model=resolved_model,
            runtime_url=runtime_url,
            already_loaded=True,
            message=f"Model '{resolved_model}' is already loaded.",
        )

    parsed = urlparse(runtime_url)
    host = str(parsed.hostname or "").strip()
    port = parsed.port or (443 if parsed.scheme == "https" else 80)
    if not host:
        raise HTTPException(
            status_code=500,
            detail=f"Configured extraction runtime URL is invalid: {runtime_url}",
        )

    api_key = (
        str(os.getenv("LLM_API_KEY") or "").strip()
        or str(os.getenv("LLAMA_SERVER_API_KEY") or "").strip()
        or str(os.getenv("LLAMACPP_API_KEY") or "").strip()
        or "local-openai-key"
    )

    status_messages: list[str] = []
    ok = load_model_api(
        host=host,
        port=str(port),
        model_name=resolved_model,
        api_key=api_key,
        timeout=300.0,
        on_status=status_messages.append,
    )
    if not ok:
        detail = status_messages[-1] if status_messages else f"Failed to load model '{resolved_model}'."
        raise HTTPException(
            status_code=400,
            detail={
                "message": detail,
                "requested_model": requested_model,
                "resolved_model": resolved_model,
                "runtime_url": runtime_url,
                "status_messages": status_messages,
            },
        )

    return ModelLoadResponse(
        ok=True,
        requested_model=requested_model,
        resolved_model=resolved_model,
        runtime_url=runtime_url,
        already_loaded=False,
        message=status_messages[-1] if status_messages else f"Model '{resolved_model}' loaded.",
    )


# ---------------------------------------------------------------------------
# GET /api/cockpit/queue
# ---------------------------------------------------------------------------


@router.get("/queue", response_model=QueueStatusResponse)
def cockpit_queue_status() -> QueueStatusResponse:
    """Return queue statistics in the shape the cockpit UI expects.

    Maps the richer Redis-based queue probe into {pending, active, completed, failed}.
    Celery does not natively expose completed/failed counts via broker state alone,
    so those are reported as 0 unless a result backend is queryable.
    """
    total_queued = 0
    active_count = 0
    try:
        parsed = urlparse(str(settings.celery_broker_url or ""))
        host = str(parsed.hostname or "127.0.0.1").strip()
        port = int(parsed.port or 6379)

        import socket as _socket

        with _socket.create_connection((host, port), timeout=2.0):
            pass  # socket reachable

        import redis as redis_lib
        from app.celery_app import _SPECIALIZED_QUEUES

        db = int((parsed.path or "").lstrip("/") or "0")
        client = redis_lib.Redis(host=host, port=port, db=db, socket_timeout=2)
        for queue_name in _SPECIALIZED_QUEUES:
            depth = client.llen(queue_name) or 0
            total_queued += depth

        # Probe active tasks via Celery inspect (best-effort, short timeout).
        try:
            from app.celery_app import celery_app

            inspector = celery_app.control.inspect(timeout=1.0)
            active_tasks = inspector.active() or {}
            for worker_tasks in active_tasks.values():
                active_count += len(worker_tasks)
        except Exception:
            pass  # workers may be offline

        client.close()
    except Exception as exc:
        logger.debug("Queue status probe failed (non-fatal): %s", exc)

    return QueueStatusResponse(
        pending=total_queued,
        active=active_count,
        completed=0,
        failed=0,
    )


# ---------------------------------------------------------------------------
# GET /api/cockpit/docs
# ---------------------------------------------------------------------------


@router.get("/docs")
def cockpit_docs():
    """Return recent documents across all tickers for the cockpit history screen.

    Unlike the main /api/docs endpoint which requires a ticker parameter, this
    returns the most recent documents globally (capped at 200).
    """
    db = SessionLocal()
    try:
        rows = (
            db.query(Document)
            .order_by(Document.published_at.desc().nullslast())
            .limit(200)
            .all()
        )
        return [
            {
                "document_id": str(r.document_id),
                "ticker": r.ticker,
                "doc_class": r.doc_class,
                "doc_subtype": r.doc_subtype,
                "published_at": r.published_at,
                "title": r.title,
                "source_url": r.source_url,
                "pdf_path": r.pdf_path,
            }
            for r in rows
        ]
    finally:
        db.close()


def _extract_latest_price_from_payload(payload: dict[str, Any]) -> tuple[float | None, str | None]:
    current = payload.get("current") if isinstance(payload.get("current"), dict) else {}
    raw_price = current.get("price")
    market_time = current.get("market_time")
    try:
        if raw_price is not None:
            price = float(raw_price)
            if price > 0:
                return price, str(market_time or "") or None
    except (TypeError, ValueError):
        pass

    history = payload.get("history")
    if isinstance(history, list):
        for point in reversed(history):
            if not isinstance(point, dict):
                continue
            raw_close = point.get("close")
            try:
                if raw_close is None:
                    continue
                close_price = float(raw_close)
            except (TypeError, ValueError):
                continue
            if close_price > 0:
                ts = point.get("timestamp")
                return close_price, str(ts or "") or None
    return None, None


def _fetch_live_price_snapshot_for_holding(
    ticker: str,
    market_exchange: str | None,
) -> dict[str, Any] | None:
    symbol = str(ticker or "").strip().upper()
    if not symbol:
        return None
    exchange_hint = _normalize_market_exchange(market_exchange)
    cache_key = (symbol, exchange_hint)
    now = time.monotonic()
    with _HOLDINGS_PRICE_CACHE_LOCK:
        cached = _HOLDINGS_PRICE_CACHE.get(cache_key)
        if cached and now - cached[0] <= _HOLDINGS_PRICE_CACHE_TTL_SECONDS:
            return cached[1]

    provider = MarketPriceProvider(
        base_url=getattr(
            settings, "market_data_base_url", "https://query1.finance.yahoo.com"
        ),
        timeout=getattr(settings, "market_data_timeout_seconds", 20.0),
    )
    candidates: list[str] = []
    if exchange_hint:
        candidates.append(exchange_hint)
    for item in _HOLDINGS_PRICE_EXCHANGE_FALLBACKS:
        if item not in candidates:
            candidates.append(item)

    snapshot: dict[str, Any] | None = None
    for exchange in candidates:
        try:
            payload = provider.fetch(
                ticker=symbol,
                exchange=exchange,
                range_="5d",
                interval="1d",
            )
        except (ValueError, MarketPriceProviderError):
            continue
        price, price_as_of = _extract_latest_price_from_payload(payload)
        if price is None:
            continue
        snapshot = {
            "current_price": price,
            "price_currency": payload.get("currency"),
            "price_as_of": price_as_of,
            "market_exchange": exchange_hint or payload.get("exchange") or exchange,
        }
        break

    with _HOLDINGS_PRICE_CACHE_LOCK:
        _HOLDINGS_PRICE_CACHE[cache_key] = (now, snapshot)
    return snapshot


def _enrich_holdings_with_live_prices(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    keys: list[tuple[str, str | None]] = []
    seen: set[tuple[str, str | None]] = set()
    for row in rows:
        ticker = str(row.get("ticker") or "").strip().upper()
        if not ticker:
            continue
        key = (ticker, _normalize_market_exchange(row.get("market_exchange")))
        if key in seen:
            continue
        seen.add(key)
        keys.append(key)

    live_by_key: dict[tuple[str, str | None], dict[str, Any] | None] = {}
    if keys:
        with ThreadPoolExecutor(max_workers=max(1, min(8, len(keys)))) as pool:
            future_map = {
                pool.submit(_fetch_live_price_snapshot_for_holding, ticker, exchange): (ticker, exchange)
                for ticker, exchange in keys
            }
            for future in as_completed(future_map):
                key = future_map[future]
                try:
                    live_by_key[key] = future.result()
                except Exception as exc:
                    logger.debug("Live price enrichment failed for %s: %s", key[0], exc)
                    live_by_key[key] = None

    enriched: list[dict[str, Any]] = []
    for row in rows:
        item = dict(row)
        ticker = str(item.get("ticker") or "").strip().upper()
        exchange = _normalize_market_exchange(item.get("market_exchange"))
        live = live_by_key.get((ticker, exchange))
        if live:
            item["market_exchange"] = live.get("market_exchange") or exchange
            item["current_price"] = live.get("current_price")
            item["price_currency"] = live.get("price_currency")
            item["price_as_of"] = live.get("price_as_of")
        else:
            item["market_exchange"] = exchange
            item["current_price"] = None
            item["price_currency"] = None
            item["price_as_of"] = None

        quantity = item.get("quantity")
        current_price = item.get("current_price")
        avg_cost = item.get("avg_cost")
        cost_currency = str(item.get("cost_currency") or "").strip().upper() or None
        price_currency = str(item.get("price_currency") or "").strip().upper() or None
        try:
            qty_val = float(quantity) if quantity is not None else None
        except (TypeError, ValueError):
            qty_val = None
        try:
            price_val = float(current_price) if current_price is not None else None
        except (TypeError, ValueError):
            price_val = None
        try:
            avg_cost_val = float(avg_cost) if avg_cost is not None else None
        except (TypeError, ValueError):
            avg_cost_val = None

        item["market_value"] = (
            qty_val * price_val if qty_val is not None and price_val is not None else None
        )
        valuation_warning: str | None = None
        if qty_val is not None and price_val is not None and avg_cost_val is not None:
            if not cost_currency or not price_currency:
                item["unrealized_pnl"] = None
                valuation_warning = (
                    "Unrealized P&L unavailable until both cost currency and live price currency are known."
                )
            elif cost_currency != price_currency:
                item["unrealized_pnl"] = None
                valuation_warning = (
                    f"Unrealized P&L unavailable due to currency mismatch "
                    f"({cost_currency} cost vs {price_currency} price)."
                )
            else:
                item["unrealized_pnl"] = (price_val - avg_cost_val) * qty_val
        else:
            item["unrealized_pnl"] = None
        item["valuation_warning"] = valuation_warning
        enriched.append(item)
    return enriched


@router.get("/holdings", response_model=CockpitHoldingListResponse)
def cockpit_list_holdings(
    ticker: str | None = None,
    include_archived: bool = False,
) -> CockpitHoldingListResponse:
    try:
        service = CockpitService.get_instance()
        rows = service.state_store.list_holdings(
            ticker=ticker,
            include_archived=include_archived,
        )
        enriched_rows = _enrich_holdings_with_live_prices([dict(row) for row in rows])
        return CockpitHoldingListResponse(
            items=[CockpitHoldingRecord(**row) for row in enriched_rows]
        )
    except Exception as exc:
        logger.exception("Failed to list cockpit holdings")
        raise HTTPException(status_code=500, detail=f"Failed to list holdings: {str(exc)}") from exc


@router.post("/holdings", response_model=CockpitHoldingRecord)
def cockpit_add_holding(payload: CockpitHoldingCreateRequest) -> CockpitHoldingRecord:
    ticker = str(payload.ticker or "").strip().upper()
    if not ticker:
        raise HTTPException(status_code=400, detail="ticker is required")
    market_exchange = _normalize_market_exchange(payload.market_exchange)

    try:
        service = CockpitService.get_instance()
        state_store = service.state_store
        holding_id = state_store.add_holding(
            ticker=ticker,
            account_label=payload.account_label,
            market_exchange=market_exchange,
            thesis_bucket=payload.thesis_bucket,
            quantity=payload.quantity,
            avg_cost=payload.avg_cost,
            cost_currency=payload.cost_currency,
            opened_at=payload.opened_at,
            note=payload.note,
        )
        row = state_store.get_holding(holding_id)
    except Exception as exc:
        logger.exception("Failed to add cockpit holding")
        raise HTTPException(status_code=500, detail=f"Failed to add holding: {str(exc)}") from exc

    if row is None:
        raise HTTPException(status_code=500, detail="Holding was created but could not be reloaded")
    return CockpitHoldingRecord(**dict(row))


@router.patch("/holdings/{holding_id}", response_model=CockpitHoldingRecord)
def cockpit_update_holding(
    holding_id: str,
    payload: CockpitHoldingUpdateRequest,
) -> CockpitHoldingRecord:
    fields = payload.model_dump(exclude_unset=True)
    if not fields:
        raise HTTPException(status_code=400, detail="No fields supplied for update")

    if "ticker" in fields:
        ticker_value = str(fields.get("ticker") or "").strip().upper()
        if not ticker_value:
            raise HTTPException(status_code=400, detail="ticker cannot be blank")
        fields["ticker"] = ticker_value
    if "market_exchange" in fields:
        fields["market_exchange"] = _normalize_market_exchange(fields.get("market_exchange"))

    try:
        service = CockpitService.get_instance()
        state_store = service.state_store
        updated = state_store.update_holding(holding_id, **fields)
        if not updated:
            raise HTTPException(status_code=404, detail=f"Holding not found: {holding_id}")
        row = state_store.get_holding(holding_id)
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Failed to update cockpit holding")
        raise HTTPException(status_code=500, detail=f"Failed to update holding: {str(exc)}") from exc

    if row is None:
        raise HTTPException(status_code=404, detail=f"Holding not found: {holding_id}")
    return CockpitHoldingRecord(**dict(row))


@router.delete("/holdings/{holding_id}", response_model=CockpitHoldingMutationResponse)
def cockpit_remove_holding(holding_id: str) -> CockpitHoldingMutationResponse:
    try:
        service = CockpitService.get_instance()
        removed = service.state_store.remove_holding(holding_id)
    except Exception as exc:
        logger.exception("Failed to remove cockpit holding")
        raise HTTPException(status_code=500, detail=f"Failed to remove holding: {str(exc)}") from exc

    if not removed:
        raise HTTPException(status_code=404, detail=f"Holding not found: {holding_id}")
    return CockpitHoldingMutationResponse(ok=True, holding_id=holding_id)


@router.post(
    "/chat/attachments/upload",
    response_model=CockpitChatAttachmentUploadResponse,
)
def cockpit_upload_chat_attachment(
    payload: CockpitChatAttachmentUploadRequest,
) -> CockpitChatAttachmentUploadResponse:
    filename = Path(str(payload.filename or "").strip()).name
    if not filename:
        raise HTTPException(status_code=400, detail="filename is required")
    if not str(payload.content_base64 or "").strip():
        raise HTTPException(status_code=400, detail="content_base64 is required")

    try:
        content_bytes = base64.b64decode(payload.content_base64, validate=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="invalid base64 content") from exc

    content_type = str(payload.mime_type or "").strip().lower()
    suffix = Path(filename).suffix.lower()
    is_csv = suffix == ".csv" or "csv" in content_type
    is_xlsx = (
        suffix in {".xlsx", ".xlsm"}
        or "spreadsheetml" in content_type
        or "application/vnd.ms-excel" in content_type
    )
    is_pdf = suffix == ".pdf" or "pdf" in content_type or content_bytes.startswith(b"%PDF")

    try:
        service = CockpitService.get_instance()
    except Exception as exc:
        logger.exception("Failed to initialize CockpitService")
        raise HTTPException(
            status_code=500, detail=f"Service initialization failed: {str(exc)}"
        ) from exc

    if (is_csv or is_xlsx) and not is_pdf:
        try:
            if is_xlsx:
                rows, headers_by_key = _parse_xlsx_rows(content_bytes)
            else:
                text = content_bytes.decode("utf-8-sig", errors="replace")
                rows, headers_by_key = _parse_csv_rows(text)
            requested_profile = payload.csv_profile
            resolved_profile: Literal["holdings", "trades"]
            if requested_profile == "auto":
                resolved_profile = _detect_csv_profile(headers_by_key)
            else:
                resolved_profile = requested_profile

            if resolved_profile == "trades":
                records, parse_errors = _extract_trade_holdings_rows_from_csv(
                    rows, headers_by_key, strict=payload.csv_strict
                )
            else:
                records, parse_errors = _extract_holdings_rows_from_csv(
                    rows, headers_by_key, strict=payload.csv_strict
                )
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        except RuntimeError as exc:
            raise HTTPException(status_code=422, detail=str(exc)) from exc

        if payload.csv_strict and parse_errors:
            joined = "; ".join(parse_errors[:5])
            raise HTTPException(status_code=400, detail=f"CSV validation failed: {joined}")

        imported_count = 0
        import_errors: list[str] = []
        for index, row in enumerate(records, start=1):
            try:
                service.state_store.add_holding(**row)
                imported_count += 1
            except Exception as exc:
                import_errors.append(f"row {index}: {exc}")
                if len(import_errors) >= 25:
                    break

        if payload.csv_strict and import_errors:
            joined = "; ".join(import_errors[:5])
            raise HTTPException(status_code=400, detail=f"CSV import failed: {joined}")

        all_errors = [*parse_errors, *import_errors]
        skipped_count = max(0, len(records) - imported_count) + len(parse_errors)
        return CockpitChatAttachmentUploadResponse(
            file_kind="holdings_csv",
            message=(
                f"Imported {imported_count} holding"
                f"{'' if imported_count == 1 else 's'} from {filename}."
            ),
            imported_count=imported_count,
            skipped_count=skipped_count,
            errors=all_errors[:25],
        )

    if is_pdf:
        try:
            extracted_text = _extract_pdf_text(content_bytes)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        except RuntimeError as exc:
            raise HTTPException(status_code=422, detail=str(exc)) from exc

        published_at = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
        try:
            source_id, chunks_staged = _stage_uploaded_pdf_chunks(
                filename=filename,
                extracted_text=extracted_text,
                published_at=published_at,
            )
        except ValueError as exc:
            raise HTTPException(status_code=422, detail=f"PDF staging failed: {exc}") from exc
        except OSError as exc:
            raise HTTPException(status_code=500, detail=f"PDF staging failed: {exc}") from exc

        source_kind: Literal["ephemeral", "concat", "primary"] = (
            "ephemeral" if chunks_staged >= 24 else "concat"
        )
        key_points = _derive_key_points_from_text(extracted_text, limit=5)
        return CockpitChatAttachmentUploadResponse(
            file_kind="strategy_pdf",
            message=f"Attached {filename} for chat context.",
            source_id=source_id,
            source_kind=source_kind,
            chunks_staged=chunks_staged,
            key_points=key_points,
        )

    raise HTTPException(
        status_code=400,
        detail="unsupported file type: upload CSV/XLSX (holdings/trades) or PDF",
    )


@router.get("/pulse", response_model=IntelPulseResponse)
def cockpit_intel_pulse(ticker: str | None = None) -> IntelPulseResponse:
    """Return system population and quality metrics for Intel Pulse."""
    try:
        service = CockpitService.get_instance()
        return service.get_intel_pulse_stats(ticker)
    except Exception as exc:
        logger.exception("Failed to fetch intel pulse stats")
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/matrix", response_model=IntelPulseMatrixResponse)
def cockpit_intel_matrix(
    stage: str, ticker: str | None = None
) -> IntelPulseMatrixResponse:
    """Return diagnostic density matrix for Intel Pulse."""
    try:
        service = CockpitService.get_instance()
        return service.get_diagnostic_matrix(stage, ticker)
    except Exception as exc:
        logger.exception("Failed to fetch diagnostic matrix")
        raise HTTPException(status_code=500, detail=str(exc))


class CockpitChatRequest(BaseModel):
    class AttachedSource(BaseModel):
        source_id: str
        source_kind: Literal["ephemeral", "concat", "primary"]

    message: str
    mode: str = "analysis"
    ticker: str | None = None
    session_id: str | None = None
    stream: bool = True
    model: str | None = None
    web_search: bool | None = None
    rag: bool | None = None
    db_diagnostics: bool | None = None
    attached_sources: list[AttachedSource] = Field(default_factory=list)


class CockpitChatSessionSummary(BaseModel):
    session_id: str
    updated_at: str | None = None
    message_count: int = 0
    title: str | None = None
    last_message: str | None = None


class CockpitChatSessionListResponse(BaseModel):
    items: list[CockpitChatSessionSummary] = Field(default_factory=list)


class CockpitChatMessageRecord(BaseModel):
    id: int
    session_id: str
    role: str
    content: str
    created_at: str


class CockpitChatSessionMessagesResponse(BaseModel):
    session_id: str
    message_count: int = 0
    items: list[CockpitChatMessageRecord] = Field(default_factory=list)


class CockpitChatSessionDeleteResponse(BaseModel):
    ok: bool
    session_id: str
    deleted_count: int = 0


class CockpitChatSessionCreateRequest(BaseModel):
    session_id: str | None = None


class CockpitChatSessionCreateResponse(BaseModel):
    ok: bool = True
    session_id: str
    created: bool = False


class CockpitPreferencesResponse(BaseModel):
    api_default_enabled: bool = False
    marketplace_prefer_cloud_routing: bool = False


class CockpitPreferencesPatchRequest(BaseModel):
    api_default_enabled: bool | None = None
    marketplace_prefer_cloud_routing: bool | None = None


class CockpitActionExecuteRequest(BaseModel):
    action_id: str
    args: dict[str, Any] = Field(default_factory=dict)
    session_id: str | None = None
    wait: bool = True


class CockpitActionExecuteResponse(BaseModel):
    ok: bool = True
    action_id: str
    result: str = ""
    exit_code: int = 0
    job_id: str | None = None
    status: str | None = None
    queued: bool = False
    chart: dict[str, str] | None = None


class CockpitActionJobStatusResponse(BaseModel):
    ok: bool = True
    job_id: str
    action_id: str
    mission_id: str | None = None
    status: str
    started_at: str | None = None
    ended_at: str | None = None
    exit_code: int | None = None
    stdout_path: str | None = None
    stderr_path: str | None = None
    result: str | None = None
    progress_stage: str | None = None
    progress_pct: float | None = None


class MarketplaceMissionRecord(BaseModel):
    mission_id: str
    name: str
    status: str
    mission_type: str = "find_good_deals"
    brief: str
    user_goal: str | None = None
    category_hint: str | None = None
    hard_filters: dict[str, Any] = Field(default_factory=dict)
    soft_preferences: dict[str, Any] = Field(default_factory=dict)
    search_config: dict[str, Any] = Field(default_factory=dict)
    scan_config: dict[str, Any] = Field(default_factory=dict)
    benchmark_sources: list[str] = Field(default_factory=lambda: ["centre_com"])
    requirement_profile: dict[str, Any] | None = None
    candidate_products: list[dict[str, Any]] = Field(default_factory=list)
    primary_tracked_product: dict[str, Any] | None = None
    benchmark_state: dict[str, Any] | None = None
    deployment_args: dict[str, Any] = Field(default_factory=dict)
    last_error: str | None = None
    created_from_chat_message_id: str | None = None
    created_at: str
    updated_at: str
    last_scan_at: str | None = None


class MarketplaceMissionListResponse(BaseModel):
    items: list[MarketplaceMissionRecord] = Field(default_factory=list)


class MarketplaceMissionDeleteResponse(BaseModel):
    ok: bool = True
    mission_id: str
    status: str = "deleted"
    deleted_missions: int = 0
    deleted_seen_listings: int = 0
    deleted_matches: int = 0
    deleted_alerts: int = 0
    deleted_listing_product_matches: int = 0
    deleted_listing_benchmark_scores: int = 0
    deleted_mission_product_links: int = 0
    deleted_mission_candidate_products: int = 0
    deleted_match_value_assessments: int = 0


class MarketplaceMissionUpsertRequest(BaseModel):
    name: str | None = None
    status: str | None = None
    mission_type: str | None = None
    brief: str | None = None
    user_goal: str | None = None
    category_hint: str | None = None
    hard_filters: dict[str, Any] = Field(default_factory=dict)
    soft_preferences: dict[str, Any] = Field(default_factory=dict)
    search_config: dict[str, Any] = Field(default_factory=dict)
    scan_config: dict[str, Any] = Field(default_factory=dict)
    benchmark_sources: list[str] = Field(default_factory=list)
    deployment_args: dict[str, Any] = Field(default_factory=dict)
    last_error: str | None = None
    created_from_chat_message_id: str | None = None


class MarketplaceMissionProductLinkRequest(BaseModel):
    tracked_product_id: str


class MarketplaceBrowserHealthResponse(BaseModel):
    status: str
    cdp_url: str
    browser_family: str
    profile_path: str
    challenge_detected: bool
    last_checked_at: str
    detail: str | None = None
    final_url: str | None = None


class MarketplaceScanRequest(BaseModel):
    mission_id: str | None = None


class MarketplaceScanJobListResponse(BaseModel):
    items: list[CockpitActionJobStatusResponse] = Field(default_factory=list)


class MarketplaceMatchRecord(BaseModel):
    match_id: str
    mission_id: str
    mission_name: str | None = None
    listing_id: str
    listing_url: str
    title: str
    price: str | None = None
    price_value: float | None = None
    location: str | None = None
    seller_name: str | None = None
    captured_at: str
    score: int
    decision_band: str
    reasons_for: list[str] = Field(default_factory=list)
    reasons_against: list[str] = Field(default_factory=list)
    confidence: float | None = None
    raw_text_snapshot: str
    screenshot_path: str | None = None
    listing_media: list[str] = Field(default_factory=list)
    status: str
    metadata: dict[str, Any] = Field(default_factory=dict)
    benchmark: dict[str, Any] | None = None
    value_context: dict[str, Any] | None = None
    updated_at: str


class MarketplaceMatchListResponse(BaseModel):
    items: list[MarketplaceMatchRecord] = Field(default_factory=list)


class MarketplaceMatchStatusRequest(BaseModel):
    status: str


class MarketplaceBenchmarkReviewRequest(BaseModel):
    review_status: str
    note: str | None = None


class MarketplaceAlertRecord(BaseModel):
    alert_id: str
    mission_id: str
    mission_name: str | None = None
    match_id: str
    match_title: str | None = None
    listing_url: str | None = None
    price: str | None = None
    location: str | None = None
    decision_band: str | None = None
    status: str
    created_at: str
    updated_at: str
    trigger_reason: str
    metadata: dict[str, Any] = Field(default_factory=dict)


class MarketplaceAlertListResponse(BaseModel):
    items: list[MarketplaceAlertRecord] = Field(default_factory=list)


class MarketplaceAlertStatusRequest(BaseModel):
    status: str


class CockpitFeedbackFlagRequest(BaseModel):
    session_id: str | None = None
    ticker: str | None = None
    feedback_type: Literal["good", "poor"] = "poor"
    capture_kind: Literal["chat_feedback", "ui_issue", "auto_diagnostic"] = "chat_feedback"
    note: str | None = None
    flagged_message: dict[str, Any] = Field(default_factory=dict)
    transcript: list[dict[str, Any]] = Field(default_factory=list)
    frontend_context: dict[str, Any] = Field(default_factory=dict)
    screenshot: dict[str, Any] | None = None


class CockpitFeedbackFlagResponse(BaseModel):
    ok: bool = True
    report_id: str
    feedback_type: Literal["good", "poor"]
    capture_kind: Literal["chat_feedback", "ui_issue", "auto_diagnostic"] = "chat_feedback"
    report_dir: str
    bundle_path: str
    summary_path: str
    analysis_path: str | None = None
    read_api_path: str
    codex_prompt: str
    codex_prompt_path: str | None = None
    investigation_path: str | None = None
    investigation_status: (
        Literal["queued", "running", "completed", "failed", "not_requested"] | None
    ) = None
    codex_cli_command: str | None = None
    analysis_summary: str | None = None
    resolution_status: Literal["open", "resolved"] = "open"
    resolved_at: str | None = None
    resolution_commit_sha: str | None = None


class CockpitFlaggedReportListItem(BaseModel):
    report_id: str
    feedback_type: Literal["good", "poor"]
    capture_kind: Literal["chat_feedback", "ui_issue", "auto_diagnostic"] = "chat_feedback"
    session_id: str
    ticker: str | None = None
    saved_at: str | None = None
    note: str | None = None
    flagged_response_excerpt: str | None = None
    read_api_path: str
    resolution_status: Literal["open", "resolved"] = "open"
    resolved_at: str | None = None
    resolution_commit_sha: str | None = None


class CockpitFlaggedReportListResponse(BaseModel):
    items: list[CockpitFlaggedReportListItem] = Field(default_factory=list)


class CockpitFlaggedReportResponse(BaseModel):
    report_id: str
    feedback_type: Literal["good", "poor"]
    capture_kind: Literal["chat_feedback", "ui_issue", "auto_diagnostic"] = "chat_feedback"
    report_dir: str
    bundle_path: str
    summary_path: str
    analysis_path: str | None = None
    read_api_path: str
    bundle: dict[str, Any] = Field(default_factory=dict)
    summary_markdown: str = ""
    analysis: dict[str, Any] | None = None
    investigation: dict[str, Any] | None = None
    codex_prompt_path: str | None = None
    investigation_path: str | None = None
    resolution_status: Literal["open", "resolved"] = "open"
    resolved_at: str | None = None
    resolution_commit_sha: str | None = None
    resolved_by: str | None = None


class CockpitFlagResolutionRequest(BaseModel):
    commit_sha: str
    resolved_by: str | None = None
    note: str | None = None


class CockpitFlagResolutionResponse(BaseModel):
    ok: bool = True
    report_id: str
    resolution_status: Literal["open", "resolved"] = "resolved"
    resolved_at: str | None = None
    resolution_commit_sha: str | None = None
    resolved_by: str | None = None
    summary_path: str
    read_api_path: str


def _coerce_float(raw: Any) -> float | None:
    text = str(raw or "").strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _coerce_int(raw: Any) -> int | None:
    value = _coerce_float(raw)
    if value is None:
        return None
    return int(value)


def _clip_action_output(value: str | None, limit: int = 12000) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    return text[:limit]


def _read_job_output(path: str | None, limit: int = 12000) -> str:
    raw_path = str(path or "").strip()
    if not raw_path:
        return ""
    try:
        return _clip_action_output(Path(raw_path).read_text(encoding="utf-8"), limit)
    except OSError:
        return ""


def _run_action_subprocess(
    *,
    normalized_command: list[str],
    repo_root: Path,
    action_env: dict[str, str],
    timeout_seconds: int,
) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        normalized_command,
        cwd=str(repo_root),
        capture_output=True,
        text=True,
        timeout=timeout_seconds,
        check=False,
        env=action_env,
    )


def _run_action_subprocess_streaming(
    *,
    job_id: str,
    normalized_command: list[str],
    repo_root: Path,
    action_env: dict[str, str],
    timeout_seconds: int,
    stdout_path: Path,
    stderr_path: Path,
    on_stdout_line: Callable[[str], None] | None = None,
) -> tuple[int, str, str]:
    """Run a subprocess, streaming stdout/stderr to log files line-by-line.

    Returns ``(exit_code, stdout_text, stderr_text)``.
    """

    def _pump(
        pipe: IO[str],
        dest: IO[str],
        callback: Callable[[str], None] | None,
    ) -> None:
        """Read lines from *pipe*, write to *dest* (flushed), optionally call *callback*."""
        try:
            for raw_line in pipe:
                dest.write(raw_line)
                dest.flush()
                if callback is not None:
                    callback(raw_line)
        except ValueError:
            pass  # pipe closed

    with (
        stdout_path.open("w", encoding="utf-8") as out_f,
        stderr_path.open("w", encoding="utf-8") as err_f,
    ):
        proc = subprocess.Popen(
            normalized_command,
            cwd=str(repo_root),
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            bufsize=1,
            env=action_env,
        )
        with _ACTION_JOB_PROC_LOCK:
            _ACTION_JOB_PROCS[job_id] = proc
        t_out = threading.Thread(
            target=_pump, args=(proc.stdout, out_f, on_stdout_line), daemon=True
        )
        t_err = threading.Thread(
            target=_pump, args=(proc.stderr, err_f, None), daemon=True
        )
        t_out.start()
        t_err.start()

        try:
            proc.wait(timeout=timeout_seconds)
        except subprocess.TimeoutExpired:
            proc.kill()
            proc.wait(timeout=5)
            t_out.join(timeout=2)
            t_err.join(timeout=2)
            with _ACTION_JOB_PROC_LOCK:
                _ACTION_JOB_PROCS.pop(job_id, None)
            return 124, "", f"Action timed out after {timeout_seconds}s\n"

        t_out.join(timeout=5)
        t_err.join(timeout=5)
        with _ACTION_JOB_PROC_LOCK:
            _ACTION_JOB_PROCS.pop(job_id, None)

    stdout_text = stdout_path.read_text(encoding="utf-8", errors="replace")
    stderr_text = stderr_path.read_text(encoding="utf-8", errors="replace")
    return proc.returncode, stdout_text, stderr_text


def _read_job_output_tail(raw_path: str | None, max_bytes: int = 64000) -> str:
    """Read the last *max_bytes* of a log file (for tailing in-progress jobs)."""
    if not raw_path:
        return ""
    p = Path(raw_path)
    if not p.is_file():
        return ""
    try:
        size = p.stat().st_size
        with p.open("r", encoding="utf-8", errors="replace") as f:
            if size > max_bytes:
                f.seek(size - max_bytes)
                f.readline()  # skip partial first line
            return f.read()
    except OSError:
        return ""


def _serialize_action_job_status(
    service: CockpitService, job_id: str, *, tail: int = 0
) -> dict[str, Any]:
    job = service.state_store.get_job(job_id)
    if job is None:
        raise FileNotFoundError(job_id)

    status = str(job.get("status") or "unknown")
    result_text = ""
    if status == "success":
        result_text = _read_job_output(job.get("stdout_path"))
    elif status == "failed":
        result_text = _read_job_output(job.get("stderr_path")) or _read_job_output(
            job.get("stdout_path")
        )
    elif status == "running":
        raw = _read_job_output_tail(job.get("stdout_path"))
        if tail > 0:
            result_text = "\n".join(raw.splitlines()[-tail:])
        else:
            result_text = raw

    return {
        "ok": True,
        "job_id": str(job.get("job_id") or job_id),
        "action_id": str(job.get("action_id") or ""),
        "mission_id": (
            str((job.get("args") or {}).get("mission_id") or "").strip() or None
            if isinstance(job.get("args"), dict)
            else None
        ),
        "status": status,
        "started_at": job.get("started_at"),
        "ended_at": job.get("ended_at"),
        "exit_code": job.get("exit_code"),
        "stdout_path": job.get("stdout_path"),
        "stderr_path": job.get("stderr_path"),
        "result": result_text or None,
        "progress_stage": job.get("progress_stage"),
        "progress_pct": job.get("progress_pct"),
    }


def _marketplace_mission_service(service: CockpitService) -> MarketplaceMissionService:
    _ensure_marketplace_scan_scheduler(service)
    return MarketplaceMissionService(service.state_store)


def _marketplace_benchmark_service(service: CockpitService) -> MarketplaceBenchmarkService:
    return MarketplaceBenchmarkService(service.state_store)


def _marketplace_price_intelligence_service(
    service: CockpitService,
) -> MarketplacePriceIntelligenceService:
    return MarketplacePriceIntelligenceService(service.state_store)


def _marketplace_requirement_profile(mission: dict[str, Any]) -> dict[str, Any] | None:
    return marketplace_requirement_profile(mission)


def _marketplace_candidate_contexts(
    mission_service: MarketplaceMissionService,
    price_service: MarketplacePriceIntelligenceService,
    mission_id: str,
) -> list[dict[str, Any]]:
    return marketplace_candidate_contexts(mission_service, price_service, mission_id)


def _marketplace_candidate_products_payload(
    mission_service: MarketplaceMissionService,
    price_service: MarketplacePriceIntelligenceService,
    mission_id: str,
) -> list[dict[str, Any]]:
    return marketplace_candidate_products_payload(
        mission_service,
        price_service,
        mission_id,
    )


def _prepare_marketplace_requirement_candidates(
    mission_service: MarketplaceMissionService,
    price_service: MarketplacePriceIntelligenceService,
    mission: dict[str, Any],
) -> dict[str, Any]:
    return prepare_requirement_driven_mission(mission_service, price_service, mission)


def _candidate_unmatched_value_context(
    *,
    profile: dict[str, Any] | None,
    resolution: dict[str, Any],
) -> dict[str, Any]:
    warnings = [
        "Requirement-driven value scoring only uses a matched candidate benchmark.",
    ]
    if resolution.get("warning"):
        warnings.append(str(resolution["warning"]))
    return {
        "state": "candidate_unmatched",
        "value_score": None,
        "value_label": "unclear",
        "value_confidence": "low",
        "benchmark_snapshot_id": None,
        "fair_low": None,
        "fair_high": None,
        "used_median": None,
        "retail_anchor_price": None,
        "price_movement_summary": None,
        "explanation": (
            "Listing did not resolve to a single requirement candidate, so no "
            "single-product benchmark was applied."
        ),
        "warnings": warnings,
        "notes": [],
        "mission_mode": "requirement_driven",
        "value_source": "none",
        "requirement_category": profile.get("category") if isinstance(profile, dict) else None,
        "candidate_match_confidence": resolution.get("candidate_match_confidence"),
    }


def _enrich_marketplace_mission_with_price_context(
    mission_service: MarketplaceMissionService,
    price_service: MarketplacePriceIntelligenceService,
    mission: dict[str, Any],
) -> dict[str, Any]:
    mission_id = str(mission.get("mission_id") or "")
    requirement_profile = _marketplace_requirement_profile(mission)
    candidate_products = _marketplace_candidate_products_payload(
        mission_service,
        price_service,
        mission_id,
    )
    link = mission_service.get_primary_tracked_product_link(str(mission.get("mission_id") or ""))
    if link is None:
        return {
            **mission,
            "requirement_profile": requirement_profile,
            "candidate_products": candidate_products,
            "primary_tracked_product": None,
            "benchmark_state": None,
        }
    product = price_service.get_tracked_product(str(link.get("tracked_product_id") or ""))
    if product is None:
        return {
            **mission,
            "requirement_profile": requirement_profile,
            "candidate_products": candidate_products,
            "primary_tracked_product": {
                **link,
                "tracked_product": None,
                "warning": "Linked tracked product was not found.",
            },
            "benchmark_state": None,
        }
    snapshot = price_service.latest_benchmark_snapshot(product["tracked_product_id"])
    return {
        **mission,
        "requirement_profile": requirement_profile,
        "candidate_products": candidate_products,
        "primary_tracked_product": {
            **link,
            "tracked_product": product,
        },
        "benchmark_state": price_service.build_benchmark_state(product, snapshot),
    }


def _enrich_marketplace_match_with_value_context(
    mission_service: MarketplaceMissionService,
    price_service: MarketplacePriceIntelligenceService,
    match: dict[str, Any],
) -> dict[str, Any]:
    mission_id = str(match.get("mission_id") or "")
    mission = mission_service.get_mission(mission_id) if mission_id else None
    profile = _marketplace_requirement_profile(mission or {})
    if isinstance(profile, dict) and profile.get("mode") == "requirement_driven":
        contexts = _marketplace_candidate_contexts(mission_service, price_service, mission_id)
        if not contexts and mission is not None:
            mission = _prepare_marketplace_requirement_candidates(
                mission_service,
                price_service,
                mission,
            )
            contexts = _marketplace_candidate_contexts(mission_service, price_service, mission_id)
        resolution = price_service.resolve_match_candidate(match, contexts)
        if not resolution.get("matched"):
            return {
                **match,
                "value_context": _candidate_unmatched_value_context(
                    profile=profile,
                    resolution=resolution,
                ),
            }
        product = resolution.get("tracked_product")
        candidate = resolution.get("candidate") if isinstance(resolution.get("candidate"), dict) else {}
        snapshot = resolution.get("benchmark_snapshot")
        if not isinstance(product, dict):
            return {
                **match,
                "value_context": _candidate_unmatched_value_context(
                    profile=profile,
                    resolution={
                        **resolution,
                        "warning": "Matched candidate tracked product was not found.",
                    },
                ),
            }
        try:
            value_context = price_service.upsert_match_value_assessment(
                match=match,
                tracked_product=product,
                snapshot=snapshot if isinstance(snapshot, dict) else None,
                context={
                    "mission_mode": "requirement_driven",
                    "value_source": "matched_candidate_benchmark",
                    "matched_candidate_tracked_product_id": product.get("tracked_product_id"),
                    "matched_candidate_name": product.get("canonical_key"),
                    "candidate_match_confidence": resolution.get("candidate_match_confidence"),
                    "requirement_fit_score": candidate.get("fit_score"),
                    "requirement_fit_label": candidate.get("fit_label"),
                    "requirement_explanation": candidate.get("explanation"),
                },
            )
        except Exception:
            logger.exception(
                "Marketplace requirement-driven value assessment failed for %s",
                match.get("match_id"),
            )
            value_context = {
                **_candidate_unmatched_value_context(profile=profile, resolution=resolution),
                "explanation": "Value assessment failed for the matched requirement candidate.",
                "warnings": ["Backend could not compute value context for the matched candidate."],
            }
        return {**match, "value_context": value_context}

    link = mission_service.get_primary_tracked_product_link(str(match.get("mission_id") or ""))
    if link is None:
        return {**match, "value_context": None}
    product = price_service.get_tracked_product(str(link.get("tracked_product_id") or ""))
    if product is None:
        return {
            **match,
            "value_context": {
                "state": "value_unavailable",
                "value_score": None,
                "value_label": "unclear",
                "value_confidence": "low",
                "benchmark_snapshot_id": None,
                "fair_low": None,
                "fair_high": None,
                "used_median": None,
                "retail_anchor_price": None,
                "price_movement_summary": None,
                "explanation": "Linked tracked product was not found.",
                "warnings": ["Relink the mission to an existing tracked product."],
                "notes": [],
            },
        }
    snapshot = price_service.latest_benchmark_snapshot(product["tracked_product_id"])
    try:
        value_context = price_service.upsert_match_value_assessment(
            match=match,
            tracked_product=product,
            snapshot=snapshot,
        )
    except Exception:
        logger.exception(
            "Marketplace value assessment failed for %s",
            match.get("match_id"),
        )
        value_context = {
            "state": "value_unavailable",
            "value_score": None,
            "value_label": "unclear",
            "value_confidence": "low",
            "benchmark_snapshot_id": snapshot.get("snapshot_id") if snapshot else None,
            "fair_low": None,
            "fair_high": None,
            "used_median": None,
            "retail_anchor_price": None,
            "price_movement_summary": None,
            "explanation": "Value assessment failed.",
            "warnings": ["Backend could not compute value context for this match."],
            "notes": [],
        }
    return {**match, "value_context": value_context}


def _list_marketplace_scan_jobs(
    service: CockpitService, *, limit: int = 50
) -> list[dict[str, Any]]:
    rows = service.state_store.list_jobs(limit=max(limit * 5, 100))
    items: list[dict[str, Any]] = []
    for row in rows:
        if str(row.get("action_id") or "") != "marketplace_scan":
            continue
        items.append(
            {
                "ok": True,
                "job_id": str(row.get("job_id") or ""),
                "action_id": "marketplace_scan",
                "mission_id": (
                    str((row.get("args") or {}).get("mission_id") or "").strip() or None
                    if isinstance(row.get("args"), dict)
                    else None
                ),
                "status": str(row.get("status") or "unknown"),
                "started_at": row.get("started_at"),
                "ended_at": row.get("ended_at"),
                "exit_code": row.get("exit_code"),
                "stdout_path": row.get("stdout_path"),
                "stderr_path": row.get("stderr_path"),
                "result": None,
                "progress_stage": row.get("progress_stage"),
                "progress_pct": row.get("progress_pct"),
            }
        )
        if len(items) >= limit:
            break
    return items


def _write_marketplace_job_line(handle: IO[str], message: str) -> None:
    handle.write(message.rstrip() + "\n")
    handle.flush()


def _marketplace_scan_in_progress(service: CockpitService) -> bool:
    for job in _list_marketplace_scan_jobs(service, limit=50):
        if str(job.get("status") or "") in {"queued", "running"}:
            job_id = str(job.get("job_id") or "")
            if job_id and _get_queued_action_job(job_id) is None:
                # Orphaned from a previous server session — no live runtime exists.
                logger.warning(
                    "Auto-cancelling orphaned marketplace scan job %s (no live runtime after restart)",
                    job_id,
                )
                service.state_store.update_job_status(
                    job_id,
                    status="cancelled",
                    exit_code=-1,
                    ended_at=datetime.now(timezone.utc).isoformat(),
                )
                continue
            return True
    return False


def _run_marketplace_scan_job(
    *,
    service: CockpitService,
    mission_id: str | None,
    job_id: str,
    stdout_path: Path,
    stderr_path: Path,
    tracker: Any | None,
    stop_event: threading.Event,
) -> None:
    started_at = datetime.now(timezone.utc).isoformat()
    mission_service = _marketplace_mission_service(service)
    title = "Marketplace scan"
    if mission_id:
        mission = mission_service.get_mission(mission_id)
        if mission is not None:
            title = f"Marketplace scan: {mission['name']}"

    _persist_action_job_row(
        service,
        job_id=job_id,
        action_id="marketplace_scan",
        args={"mission_id": mission_id},
        started_at=started_at,
        status="running",
        stdout_path=stdout_path,
        stderr_path=stderr_path,
        progress_stage="Starting Marketplace scan",
        progress_pct=0.0,
    )
    if tracker is not None:
        _best_effort_tracker_call("start_job", job_id)
        _best_effort_tracker_call(
            "change_phase", job_id, "marketplace_scan", message=f"Started {title}"
        )

    with stdout_path.open("a", encoding="utf-8") as stdout_handle, stderr_path.open(
        "a", encoding="utf-8"
    ) as stderr_handle:
        scanner = MarketplaceScanner(mission_service)

        def progress(stage: str, pct: float | None) -> None:
            service.state_store.update_job_progress(job_id, stage, pct)
            if tracker is not None:
                _best_effort_tracker_call("change_phase", job_id, stage, message=stage)
                if pct is not None:
                    bounded = max(0, min(100, int(pct)))
                    # Update tracker metrics so JobDetailPanel shows the progress bar
                    try:
                        tracker.store.update_job_run(
                            job_id,
                            total_items=100,
                            succeeded_items=bounded,
                            current_item_label=stage,
                        )
                    except Exception:
                        pass
                    _best_effort_tracker_call(
                        "record_progress",
                        job_id,
                        current=bounded,
                        total=100,
                        message=stage,
                    )

        def log(message: str) -> None:
            _write_marketplace_job_line(stdout_handle, message)

        try:
            result = scanner.run_sync(
                mission_id=mission_id,
                progress=progress,
                log=log,
                cancel_requested=stop_event.is_set,
            )
            _write_marketplace_job_line(stdout_handle, json.dumps(result, indent=2))
            ended_at = datetime.now(timezone.utc).isoformat()
            service.state_store.update_job_status(
                job_id,
                status="success",
                exit_code=0,
                ended_at=ended_at,
            )
            if tracker is not None:
                _best_effort_tracker_call("complete_job", job_id, result.get("summary"))
        except MarketplaceScanCancelled as exc:
            _write_marketplace_job_line(stdout_handle, str(exc))
            ended_at = datetime.now(timezone.utc).isoformat()
            service.state_store.update_job_status(
                job_id,
                status="cancelled",
                exit_code=130,
                ended_at=ended_at,
            )
            if tracker is not None:
                _best_effort_tracker_call("cancel_job", job_id, reason=str(exc))
        except Exception as exc:
            _write_marketplace_job_line(stderr_handle, str(exc))
            ended_at = datetime.now(timezone.utc).isoformat()
            service.state_store.update_job_status(
                job_id,
                status="failed",
                exit_code=1,
                ended_at=ended_at,
            )
            if tracker is not None:
                _best_effort_tracker_call("fail_job", job_id, str(exc))
        finally:
            _forget_queued_action_job(job_id)


def _launch_marketplace_scan_job(
    service: CockpitService,
    *,
    mission_id: str | None,
    trigger_source: str = "cockpit",
) -> dict[str, Any]:
    tracker = _get_backend_job_tracker()
    job_id = uuid.uuid4().hex
    logs_dir = Path(service.artifact_store.logs_dir)
    logs_dir.mkdir(parents=True, exist_ok=True)
    stdout_path = logs_dir / f"{job_id}.out.log"
    stderr_path = logs_dir / f"{job_id}.err.log"
    started_at = datetime.now(timezone.utc).isoformat()
    title = "Marketplace scan"

    if tracker is not None:
        _best_effort_tracker_call(
            "create_job",
            job_id=job_id,
            job_type="marketplace_scan",
            job_family="marketplace",
            title=title,
            trigger_source=trigger_source,
            entity_scope=mission_id or "all",
            metadata={"mission_id": mission_id, "trigger_source": trigger_source},
        )
        _best_effort_tracker_call(
            "add_artifact",
            job_id,
            artifact_type="log",
            artifact_label="stdout log",
            artifact_path=str(stdout_path),
        )
        _best_effort_tracker_call(
            "add_artifact",
            job_id,
            artifact_type="log",
            artifact_label="stderr log",
            artifact_path=str(stderr_path),
        )

    _persist_action_job_row(
        service,
        job_id=job_id,
        action_id="marketplace_scan",
        args={"mission_id": mission_id},
        started_at=started_at,
        status="queued",
        stdout_path=stdout_path,
        stderr_path=stderr_path,
        progress_stage="Queued",
        progress_pct=0.0,
    )
    runtime = QueuedActionJobRuntime(
        job_id=job_id,
        action_id="marketplace_scan",
        started_at=started_at,
    )
    _register_queued_action_job(runtime)

    worker = threading.Thread(
        target=_run_marketplace_scan_job,
        kwargs={
            "service": service,
            "mission_id": mission_id,
            "job_id": job_id,
            "stdout_path": stdout_path,
            "stderr_path": stderr_path,
            "tracker": tracker,
            "stop_event": runtime.stop_event,
        },
        daemon=True,
        name=f"marketplace-scan-{job_id[:8]}",
    )
    worker.start()

    return {
        "ok": True,
        "action_id": "marketplace_scan",
        "result": f"Queued marketplace scan job ({trigger_source})",
        "exit_code": 0,
        "job_id": job_id,
        "status": "queued",
        "queued": True,
    }


def _run_marketplace_scheduler_tick(service: CockpitService) -> list[dict[str, Any]]:
    mission_service = MarketplaceMissionService(service.state_store)
    due_missions = mission_service.due_missions()
    if not due_missions or _marketplace_scan_in_progress(service):
        return []

    health = check_marketplace_browser_health()
    if not marketplace_scan_health_allows_execution(health):
        return []

    launched: list[dict[str, Any]] = []
    mission = due_missions[0]
    queued = _launch_marketplace_scan_job(
        service,
        mission_id=str(mission.get("mission_id") or "") or None,
        trigger_source="scheduler",
    )
    launched.append(
        {
            "mission_id": mission.get("mission_id"),
            "job_id": queued.get("job_id"),
        }
    )
    return launched


def _marketplace_scheduler_loop(service: CockpitService) -> None:
    while True:
        try:
            launched = _run_marketplace_scheduler_tick(service)
            if launched:
                logger.info("Marketplace scheduler queued %d scan(s)", len(launched))
        except Exception:
            logger.exception("Marketplace scheduler tick failed")
        time.sleep(_MARKETPLACE_SCHEDULER_INTERVAL_SECONDS)


def _cancel_orphaned_marketplace_scan_jobs(service: CockpitService) -> None:
    """Cancel any marketplace_scan jobs left in running/queued state from a previous session."""
    for job in _list_marketplace_scan_jobs(service, limit=100):
        if str(job.get("status") or "") in {"queued", "running"}:
            job_id = str(job.get("job_id") or "")
            if not job_id:
                continue
            logger.warning(
                "Startup cleanup: cancelling orphaned marketplace scan job %s", job_id
            )
            service.state_store.update_job_status(
                job_id,
                status="cancelled",
                exit_code=-1,
                ended_at=datetime.now(timezone.utc).isoformat(),
            )


def _ensure_marketplace_scan_scheduler(service: CockpitService) -> None:
    global _MARKETPLACE_SCHEDULER_STARTED
    with _MARKETPLACE_SCHEDULER_LOCK:
        if _MARKETPLACE_SCHEDULER_STARTED:
            return
        _cancel_orphaned_marketplace_scan_jobs(service)
        worker = threading.Thread(
            target=_marketplace_scheduler_loop,
            kwargs={"service": service},
            daemon=True,
            name="marketplace-scheduler",
        )
        worker.start()
        _MARKETPLACE_SCHEDULER_STARTED = True


def _launch_action_job(
    *,
    service: CockpitService,
    action_id: str,
    args: dict[str, Any],
    normalized_command: list[str],
    action_env: dict[str, str],
    timeout_seconds: int,
) -> dict[str, Any]:
    from app.services.job_tracker import get_tracker

    job_id = uuid.uuid4().hex
    logs_dir = Path(service.artifact_store.logs_dir)
    logs_dir.mkdir(parents=True, exist_ok=True)
    stdout_path = logs_dir / f"{job_id}.out.log"
    stderr_path = logs_dir / f"{job_id}.err.log"
    started_at = datetime.now(timezone.utc).isoformat()
    tracker = get_tracker()
    ticker = str(args.get("ticker") or "").strip().upper() or None
    tickers = str(args.get("tickers") or "").strip().upper()
    entity_scope = tickers or action_id
    if not ticker and tickers and "," not in tickers:
        ticker = tickers
    action_label = action_id
    try:
        action_label = str(service.action_registry.get(action_id).label or action_id)
    except Exception:
        action_label = action_id

    if tracker is not None:
        try:
            tracker.create_job(
                job_id=job_id,
                job_type=action_id,
                job_family="cockpit_action",
                title=action_label,
                trigger_source="cockpit",
                entity_scope=entity_scope,
                ticker=ticker,
                metadata={"args": args},
            )
            tracker.add_artifact(
                job_id,
                artifact_type="log",
                artifact_label="stdout log",
                artifact_path=str(stdout_path),
            )
            tracker.add_artifact(
                job_id,
                artifact_type="log",
                artifact_label="stderr log",
                artifact_path=str(stderr_path),
            )
        except Exception:
            logger.warning(
                "ops tracker init for cockpit action failed (non-fatal)",
                exc_info=True,
            )

    service.state_store.add_job(
        {
            "job_id": job_id,
            "action_id": action_id,
            "args": args,
            "started_at": started_at,
            "ended_at": None,
            "status": "queued",
            "exit_code": None,
            "stdout_path": str(stdout_path),
            "stderr_path": str(stderr_path),
            "artifacts": [],
        }
    )

    def _worker() -> None:
        from app.services.progress_parser import parse_progress_line

        if tracker is not None:
            try:
                tracker.start_job(job_id)
            except Exception:
                logger.warning(
                    "ops tracker start for cockpit action failed (non-fatal)",
                    exc_info=True,
                )
        service.state_store.add_job(
            {
                "job_id": job_id,
                "action_id": action_id,
                "args": args,
                "started_at": started_at,
                "ended_at": None,
                "status": "running",
                "exit_code": None,
                "stdout_path": str(stdout_path),
                "stderr_path": str(stderr_path),
                "artifacts": [],
            }
        )

        _last_stage: str | None = None

        def _on_stdout_line(line: str) -> None:
            nonlocal _last_stage
            info = parse_progress_line(line)
            if info is None:
                return
            # Debounce: only write to DB when stage or pct changes
            if info.stage != _last_stage or info.pct is not None:
                _last_stage = info.stage
                service.state_store.update_job_progress(
                    job_id, info.stage, info.pct
                )
                if tracker is not None:
                    try:
                        tracker.change_phase(
                            job_id, info.stage, message=info.detail or info.stage
                        )
                        if info.current is not None and info.total is not None:
                            # Update tracker metrics so JobDetailPanel shows the progress bar
                            tracker.store.update_job_run(
                                job_id,
                                total_items=info.total,
                                succeeded_items=info.current,
                                current_item_label=info.detail,
                            )
                            tracker.record_progress(
                                job_id,
                                current=info.current,
                                total=info.total,
                                message=info.detail
                                or f"{info.current}/{info.total}",
                            )
                        elif info.percent_override is not None:
                            tracker.record_progress(
                                job_id,
                                current=int(info.percent_override),
                                total=100,
                                message=info.detail or f"{info.percent_override}%",
                            )
                    except Exception:
                        logger.warning(
                            "ops tracker progress for cockpit action failed (non-fatal)",
                            exc_info=True,
                        )

        exit_code: int | None = None
        status = "failed"
        stdout_text = ""
        stderr_text = ""
        try:
            exit_code, stdout_text, stderr_text = _run_action_subprocess_streaming(
                job_id=job_id,
                normalized_command=normalized_command,
                repo_root=Path(service.repo_root),
                action_env=action_env,
                timeout_seconds=timeout_seconds,
                stdout_path=stdout_path,
                stderr_path=stderr_path,
                on_stdout_line=_on_stdout_line,
            )
            with _ACTION_JOB_PROC_LOCK:
                was_cancelled = job_id in _ACTION_JOB_CANCEL_REQUESTS
                if was_cancelled:
                    _ACTION_JOB_CANCEL_REQUESTS.discard(job_id)
            status = (
                "cancelled" if was_cancelled else "success" if exit_code == 0 else "failed"
            )
        except Exception as exc:
            stderr_path.write_text(
                f"Action execution failed: {exc}\n", encoding="utf-8"
            )
            stderr_text = f"Action execution failed: {exc}\n"
            status = "failed"

        # Update the job status without overwriting the progress metrics
        service.state_store.update_job_status(
            job_id,
            status=status,
            exit_code=exit_code,
            ended_at=datetime.now(timezone.utc).isoformat(),
        )
        if tracker is not None:
            try:
                if status == "success":
                    tracker.complete_job(
                        job_id,
                        summary=_clip_action_output(
                            stdout_text or f"Action {action_id} completed successfully",
                            4000,
                        ),
                    )
                elif status == "cancelled":
                    tracker.cancel_job(job_id, reason="Cockpit action cancelled")
                else:
                    tracker.fail_job(
                        job_id,
                        _clip_action_output(
                            stderr_text or stdout_text or f"Action {action_id} failed",
                            4000,
                        ),
                    )
            except Exception:
                logger.warning(
                    "ops tracker finalize for cockpit action failed (non-fatal)",
                    exc_info=True,
                )

    thread = threading.Thread(
        target=_worker,
        daemon=True,
        name=f"cockpit-action-{job_id[:8]}",
    )
    thread.start()
    return {
        "ok": True,
        "action_id": action_id,
        "result": f"Queued action {action_id}",
        "exit_code": 0,
        "job_id": job_id,
        "status": "queued",
        "queued": True,
    }


def _resolve_chart_csv_path(repo_root: Path, raw_path: str) -> Path:
    raw_text = str(raw_path or "").strip()
    if not raw_text:
        raise HTTPException(status_code=400, detail="Chart CSV path is required")

    candidate = Path(raw_text)
    if not candidate.is_absolute():
        candidate = (repo_root / candidate).resolve()
    else:
        candidate = candidate.resolve()

    allowed_root = (repo_root / "reports" / "candles").resolve()
    try:
        candidate.relative_to(allowed_root)
    except ValueError as exc:
        raise HTTPException(
            status_code=400,
            detail="Chart CSV must be under reports/candles",
        ) from exc

    if not candidate.exists():
        raise HTTPException(
            status_code=404, detail=f"Chart CSV not found: {candidate.name}"
        )
    return candidate


def _read_chart_rows_from_csv(csv_path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with csv_path.open("r", encoding="utf-8", newline="") as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            rows.append(
                {
                    "timestamp": str(row.get("timestamp") or ""),
                    "open": _coerce_float(row.get("open")),
                    "high": _coerce_float(row.get("high")),
                    "low": _coerce_float(row.get("low")),
                    "close": _coerce_float(row.get("close")),
                    "volume": _coerce_int(row.get("volume")),
                }
            )
    return rows


def _build_candlestick_chart_response(
    service: CockpitService,
    action_id: str,
    args: dict[str, Any],
) -> CockpitActionExecuteResponse:
    from cockpit.core.plotly_html import build_candlestick_dashboard_html

    ticker = str(args.get("ticker") or "UNKNOWN").strip().upper() or "UNKNOWN"
    timeframe = str(args.get("timeframe") or "1d").strip() or "1d"
    mode_flag = str(args.get("mode_flag") or "").strip()

    price_state: dict[str, Any] = {}
    recent_history: list[dict[str, Any]] = []
    csv_error: HTTPException | None = None

    if mode_flag == "-f":
        try:
            csv_path = _resolve_chart_csv_path(
                service.repo_root, str(args.get("mode_value") or "")
            )
            recent_history = _read_chart_rows_from_csv(csv_path)
            if recent_history:
                latest_close = recent_history[-1].get("close")
                price_state = {
                    "current": {"close": latest_close},
                    "metrics": {"sample_count": len(recent_history)},
                }
        except HTTPException as exc:
            csv_error = exc

    if not recent_history:
        try:
            bundle = service.tool_router.get_price_context_for_window(
                ticker,
                range_="1y",
                interval=timeframe,
                max_history_rows=260,
            )
        except RuntimeError as exc:
            raise HTTPException(status_code=502, detail=str(exc)) from exc

        if not isinstance(bundle, dict):
            raise HTTPException(
                status_code=502, detail=f"Chart data unavailable for {ticker}"
            )

        price = bundle.get("price") if isinstance(bundle.get("price"), dict) else {}
        recent_history = (
            price.get("recent_history")
            if isinstance(price.get("recent_history"), list)
            else []
        )
        if not recent_history:
            if csv_error is not None:
                raise csv_error
            raise HTTPException(
                status_code=404, detail=f"No OHLC data available for {ticker}"
            )
        price_state = (
            bundle.get("price_state")
            if isinstance(bundle.get("price_state"), dict)
            else {}
        )

    html = build_candlestick_dashboard_html(
        {
            "ticker": ticker,
            "window": "1y",
            "price_state": price_state,
            "recent_history": recent_history,
            "generated_at": datetime.now(timezone.utc).isoformat(),
        }
    )
    if not isinstance(html, str) or not html.strip():
        raise HTTPException(
            status_code=502,
            detail=f"Candlestick chart render failed for {ticker}: empty chart payload",
        )
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S_%f")
    service.artifact_store.write_text(
        f"reports/cockpit/{ticker}_{ts}_candlestick_dashboard.html",
        html,
    )
    return CockpitActionExecuteResponse(
        ok=True,
        action_id=action_id,
        result=f"Candlestick chart rendered for {ticker} ({timeframe}).",
        exit_code=0,
        chart={
            "title": f"{ticker} candlestick chart",
            "html": html,
        },
    )


def _build_filestats_chart_from_chat_response(
    response: Any,
) -> dict[str, str] | None:
    try:
        from cockpit.core.plotly_html import build_filestats_dashboard_html
    except Exception as exc:
        logger.debug("Filestats dashboard builder unavailable: %s", exc)
        return None

    evidence = getattr(response, "evidence", None)
    if not isinstance(evidence, list):
        return None

    for item in evidence:
        if not isinstance(item, dict):
            continue
        if str(item.get("type") or "") != "company_dump":
            continue
        details = item.get("details")
        if not isinstance(details, dict):
            continue

        backend = details.get("backend")
        if not isinstance(backend, dict):
            continue

        ticker = (
            str(details.get("ticker") or backend.get("ticker") or "UNKNOWN")
            .strip()
            .upper()
        )
        cockpit_local_memory = details.get("cockpit_local_memory")
        if not isinstance(cockpit_local_memory, dict):
            cockpit_local_memory = {}

        payload = {
            **backend,
            "ticker": ticker,
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "cockpit_local_memory": cockpit_local_memory,
        }
        try:
            html = build_filestats_dashboard_html(payload)
        except Exception as exc:
            logger.warning("Failed to build filestats dashboard HTML: %s", exc)
            return None
        return {
            "title": f"{ticker} filestats dashboard",
            "html": html,
        }

    return None


def _build_holdings_chart_from_chat_response(
    response: Any,
) -> dict[str, str] | None:
    evidence = getattr(response, "evidence", None)
    if not isinstance(evidence, list):
        return None

    for item in evidence:
        if not isinstance(item, dict):
            continue
        if str(item.get("type") or "") != "holdings":
            continue
        details = item.get("details")
        if not isinstance(details, list):
            continue

        positions: list[dict[str, Any]] = []
        for row in details:
            if not isinstance(row, dict):
                continue
            ticker = str(row.get("ticker") or "").strip().upper()
            if not ticker:
                continue

            market_value = _coerce_float(row.get("market_value"))
            price_currency = str(row.get("price_currency") or "").strip().upper() or None
            if market_value is not None and market_value > 0 and price_currency:
                positions.append(
                    {
                        "label": ticker,
                        "value": market_value,
                        "currency": price_currency,
                    }
                )
                continue

            quantity = _coerce_float(row.get("quantity"))
            avg_cost = _coerce_float(row.get("avg_cost"))
            cost_currency = str(row.get("cost_currency") or "").strip().upper() or None
            if (
                quantity is not None
                and avg_cost is not None
                and quantity > 0
                and avg_cost > 0
                and cost_currency
            ):
                positions.append(
                    {
                        "label": ticker,
                        "value": quantity * avg_cost,
                        "currency": cost_currency,
                    }
                )

        if not positions:
            return None

        totals_by_currency: dict[str, float] = {}
        for pos in positions:
            ccy = str(pos.get("currency") or "").strip().upper()
            totals_by_currency[ccy] = totals_by_currency.get(ccy, 0.0) + float(
                pos["value"]
            )
        if not totals_by_currency:
            return None

        dominant_currency = max(totals_by_currency.items(), key=lambda item: item[1])[0]
        total_value = totals_by_currency.get(dominant_currency, 0.0)
        if total_value <= 0:
            return None

        dominant_positions = [
            pos for pos in positions if pos.get("currency") == dominant_currency
        ]
        dominant_positions.sort(key=lambda pos: float(pos.get("value") or 0.0), reverse=True)
        top_positions = dominant_positions[:12]
        displayed_total = sum(float(pos.get("value") or 0.0) for pos in top_positions)
        other_value = max(total_value - displayed_total, 0.0)

        rows_html: list[str] = []
        for pos in top_positions:
            value = float(pos.get("value") or 0.0)
            if value <= 0:
                continue
            pct = (value / total_value) * 100 if total_value > 0 else 0.0
            rows_html.append(
                "<div class='row'>"
                f"<div class='label'>{html.escape(str(pos.get('label') or '?'))}</div>"
                "<div class='track'><div class='fill' "
                f"style='width:{pct:.1f}%'></div></div>"
                f"<div class='value'>{dominant_currency} {value:,.2f} ({pct:.1f}%)</div>"
                "</div>"
            )

        if other_value > 0:
            pct = (other_value / total_value) * 100
            rows_html.append(
                "<div class='row'>"
                "<div class='label'>OTHER</div>"
                "<div class='track'><div class='fill other' "
                f"style='width:{pct:.1f}%'></div></div>"
                f"<div class='value'>{dominant_currency} {other_value:,.2f} ({pct:.1f}%)</div>"
                "</div>"
            )

        if not rows_html:
            return None

        generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
        html_payload = f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <title>Portfolio Allocation</title>
  <style>
    :root {{
      color-scheme: dark;
      --bg: #071019;
      --surface: #0d1724;
      --line: #1f3247;
      --text: #d6e4f5;
      --muted: #90abc8;
      --accent: #2dd4bf;
      --other: #7dd3fc;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      padding: 20px;
      font-family: "IBM Plex Sans", "Segoe UI", sans-serif;
      background: radial-gradient(circle at 0% 0%, #10253d, var(--bg) 60%);
      color: var(--text);
    }}
    .card {{
      max-width: 1040px;
      margin: 0 auto;
      border: 1px solid var(--line);
      background: linear-gradient(170deg, rgba(255, 255, 255, 0.04), transparent 60%), var(--surface);
      border-radius: 14px;
      padding: 18px 18px 14px;
    }}
    h1 {{
      margin: 0;
      font-size: 18px;
      letter-spacing: 0.02em;
    }}
    .meta {{
      margin: 6px 0 16px;
      color: var(--muted);
      font-size: 12px;
    }}
    .row {{
      display: grid;
      grid-template-columns: 120px 1fr 220px;
      gap: 10px;
      align-items: center;
      padding: 6px 0;
    }}
    .label {{
      font-size: 12px;
      letter-spacing: 0.04em;
      color: #e2ecf8;
      text-transform: uppercase;
      white-space: nowrap;
      overflow: hidden;
      text-overflow: ellipsis;
    }}
    .track {{
      height: 10px;
      border: 1px solid var(--line);
      border-radius: 999px;
      overflow: hidden;
      background: rgba(255, 255, 255, 0.05);
    }}
    .fill {{
      height: 100%;
      background: linear-gradient(90deg, rgba(45, 212, 191, 0.55), rgba(45, 212, 191, 0.95));
    }}
    .fill.other {{
      background: linear-gradient(90deg, rgba(125, 211, 252, 0.55), rgba(125, 211, 252, 0.95));
    }}
    .value {{
      text-align: right;
      font-variant-numeric: tabular-nums;
      font-size: 12px;
      color: #dce9f7;
    }}
    @media (max-width: 820px) {{
      body {{ padding: 10px; }}
      .row {{
        grid-template-columns: 88px 1fr;
        gap: 8px;
      }}
      .value {{
        grid-column: 1 / -1;
        text-align: left;
      }}
    }}
  </style>
</head>
<body>
  <div class="card">
    <h1>Portfolio Allocation ({dominant_currency})</h1>
    <div class="meta">Total {dominant_currency} {total_value:,.2f} • generated {generated_at}</div>
    {"".join(rows_html)}
  </div>
</body>
</html>"""
        return {
            "title": f"Holdings allocation ({dominant_currency})",
            "html": html_payload,
        }

    return None


def _build_chart_from_chat_response(response: Any) -> dict[str, str] | None:
    filestats_chart = _build_filestats_chart_from_chat_response(response)
    if filestats_chart is not None:
        return filestats_chart
    return _build_holdings_chart_from_chat_response(response)


class CockpitActionPreviewRequest(BaseModel):
    action_id: str
    args: dict[str, Any] = Field(default_factory=dict)


class CockpitActionPreviewResponse(BaseModel):
    action_id: str
    command: list[str]
    summary: str
    estimated_impact: str
    timeout_seconds: int
    guard_message: str | None = None


@router.post("/action/preview", response_model=CockpitActionPreviewResponse)
async def cockpit_preview_action(payload: CockpitActionPreviewRequest):
    """Preview an action: show the command that would run, impact, and guard warnings."""
    try:
        service = CockpitService.get_instance()
    except Exception as exc:
        logger.exception("Failed to initialize CockpitService for action preview")
        raise HTTPException(
            status_code=500, detail=f"Service initialization failed: {str(exc)}"
        ) from exc

    action_id = str(payload.action_id or "").strip()
    if not action_id:
        raise HTTPException(status_code=400, detail="action_id is required")

    args = payload.args if isinstance(payload.args, dict) else {}

    try:
        preview = service.action_registry.preview(action_id, args)
    except KeyError as exc:
        raise HTTPException(
            status_code=404, detail=f"Unknown action_id: {action_id}"
        ) from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    return CockpitActionPreviewResponse(
        action_id=preview.action_id,
        command=preview.command,
        summary=preview.summary,
        estimated_impact=preview.estimated_impact,
        timeout_seconds=preview.timeout_seconds,
        guard_message=preview.guard_message,
    )


def _normalize_action_command(command: list[str], repo_root: Path) -> list[str]:
    """Best-effort command normalization for container/runtime differences."""
    if not command:
        return command

    normalized = list(command)

    # Python launcher normalization: ActionRegistry may include /.venv/bin/python which
    # does not exist in some backend container setups.
    python_bin = Path(normalized[0])
    if not python_bin.exists():
        normalized[0] = sys.executable

    # Script path fallback: resolve against known mount points.
    if len(normalized) > 1:
        raw_script = normalized[1]
        script_path = Path(raw_script)
        if script_path.is_absolute() and not script_path.exists():
            shared_root_override = str(
                os.getenv("COCKPIT_SHARED_SCRIPTS_ROOT") or ""
            ).strip()
            candidate_roots = [
                Path(shared_root_override) if shared_root_override else None,
                Path("/workspace/scripts"),
                Path("/workspace-scripts"),
            ]
            for root in candidate_roots:
                if root is None:
                    continue
                candidate = (root / script_path.name).resolve()
                if candidate.exists():
                    normalized[1] = str(candidate)
                    script_path = candidate
                    break

        if not script_path.is_absolute():
            candidates: list[Path] = [
                (repo_root / script_path).resolve(),
                (Path("/app") / script_path).resolve(),
                (Path("/scripts") / script_path.name).resolve(),
                (Path("/workspace/scripts") / script_path.name).resolve(),
                (Path("/workspace-scripts") / script_path.name).resolve(),
            ]
            if raw_script.startswith("../scripts/"):
                candidates.append(
                    (Path("/scripts") / raw_script.split("/")[-1]).resolve()
                )

            for candidate in candidates:
                if candidate.exists():
                    normalized[1] = str(candidate)
                    break

    return normalized


def _build_action_env(repo_root: Path) -> dict[str, str]:
    """Ensure action subprocesses can import backend/cockpit modules."""
    env = os.environ.copy()
    existing = env.get("PYTHONPATH", "")
    shared_scripts_root = str(
        os.getenv("COCKPIT_SHARED_SCRIPTS_ROOT") or ""
    ).strip()
    candidates = [
        str((repo_root / "backend").resolve()),
        str((repo_root / "cockpit").resolve()),
        str((repo_root.parent / "scripts").resolve()),
        str((repo_root / "scripts").resolve()),
        shared_scripts_root,
        "/app",
        "/app/cockpit",
        "/scripts",
        "/workspace/scripts",
    ]
    merged = [p for p in candidates if p]
    if existing:
        merged.append(existing)
    env["PYTHONPATH"] = ":".join(merged)
    return env


def _execute_user_thesis_action(
    action_id: str,
    args: dict[str, Any],
) -> CockpitActionExecuteResponse:
    from app.services.user_thesis_memory import UserThesisMemoryStore

    store = UserThesisMemoryStore()
    ticker = str(args.get("ticker") or "").strip().upper()
    if not ticker:
        raise HTTPException(status_code=400, detail="ticker is required")

    if action_id == "create_thesis":
        statement = str(args.get("thesis") or "").strip()
        signal = str(args.get("signal") or "HOLD").strip().upper()
        if not statement:
            raise HTTPException(status_code=400, detail="thesis is required")
        proposal = store.create_proposal(
            ticker=ticker,
            proposal_type="create_thesis",
            statement=statement,
            signal=signal,
            confidence=0.7,
            metadata={"source_action_id": action_id},
            requested_by="cockpit_user",
        )
    elif action_id == "add_thesis_evidence":
        statement = str(args.get("finding") or "").strip()
        if not statement:
            raise HTTPException(status_code=400, detail="finding is required")
        is_supporting = bool(args.get("is_supporting", True))
        proposal = store.create_proposal(
            ticker=ticker,
            proposal_type="add_evidence",
            statement=statement,
            confidence=0.7,
            metadata={
                "source_action_id": action_id,
                "is_supporting": is_supporting,
            },
            requested_by="cockpit_user",
        )
    else:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported strategy action: {action_id}",
        )

    proposal_id = str(proposal.get("proposal_id") or "")
    store.confirm_proposal(
        proposal_id,
        note="confirmed via cockpit action execute",
    )
    applied = store.apply_confirmed_proposal(proposal_id)
    entry = dict(applied.get("entry") or {})
    statement = str(entry.get("statement") or proposal.get("statement") or "").strip()
    if len(statement) > 180:
        statement = statement[:177] + "..."

    return CockpitActionExecuteResponse(
        ok=True,
        action_id=action_id,
        result=(
            f"Recorded user thesis memory for {ticker}: {statement}"
            f" (proposal_id={proposal_id})"
        ),
        exit_code=0,
        status="success",
    )


@router.post("/action/execute", response_model=CockpitActionExecuteResponse)
async def cockpit_execute_action(payload: CockpitActionExecuteRequest):
    """Execute a confirmed cockpit action command and return output."""
    try:
        service = CockpitService.get_instance()
    except Exception as exc:
        logger.exception("Failed to initialize CockpitService for action execution")
        raise HTTPException(
            status_code=500, detail=f"Service initialization failed: {str(exc)}"
        ) from exc

    action_id = str(payload.action_id or "").strip()
    if not action_id:
        raise HTTPException(status_code=400, detail="action_id is required")

    args = payload.args if isinstance(payload.args, dict) else {}
    if action_id in {"create_thesis", "add_thesis_evidence"}:
        try:
            return await asyncio.to_thread(
                _execute_user_thesis_action,
                action_id,
                args,
            )
        except HTTPException:
            raise
        except Exception as exc:
            logger.exception("Strategy memory action execution failed: %s", action_id)
            raise HTTPException(
                status_code=500,
                detail=f"Strategy memory action failed: {str(exc)}",
            ) from exc

    try:
        preview = service.action_registry.preview(action_id, args)
    except KeyError as exc:
        raise HTTPException(
            status_code=404, detail=f"Unknown action_id: {action_id}"
        ) from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    if action_id == "show_candlestick":
        return _build_candlestick_chart_response(service, action_id, args)

    timeout_seconds = max(1, int(preview.timeout_seconds or 300))
    normalized_command = _normalize_action_command(
        preview.command, Path(service.repo_root)
    )
    action_env = _build_action_env(Path(service.repo_root))

    if not payload.wait:
        queued = _launch_action_job(
            service=service,
            action_id=action_id,
            args=args,
            normalized_command=normalized_command,
            action_env=action_env,
            timeout_seconds=timeout_seconds,
        )
        return CockpitActionExecuteResponse(**queued)

    try:
        proc = await asyncio.to_thread(
            _run_action_subprocess,
            normalized_command=normalized_command,
            repo_root=Path(service.repo_root),
            action_env=action_env,
            timeout_seconds=timeout_seconds,
        )
    except subprocess.TimeoutExpired as exc:
        raise HTTPException(
            status_code=504,
            detail=f"Action timed out after {timeout_seconds}s: {action_id}",
        ) from exc
    except Exception as exc:
        logger.exception("Action execution failed: %s", action_id)
        raise HTTPException(
            status_code=500,
            detail=f"Action execution failed: {str(exc)}",
        ) from exc

    stdout = (proc.stdout or "").strip()
    stderr = (proc.stderr or "").strip()

    if proc.returncode != 0:
        detail = stderr or stdout or f"Action failed with exit code {proc.returncode}"
        raise HTTPException(status_code=500, detail=detail[:4000])

    output = stdout or stderr or f"Action {action_id} completed successfully"
    return CockpitActionExecuteResponse(
        ok=True,
        action_id=action_id,
        result=output[:12000],
        exit_code=proc.returncode,
        status="success",
    )


@router.get("/action/jobs/{job_id}", response_model=CockpitActionJobStatusResponse)
async def cockpit_get_action_job(job_id: str, tail: int = 0):
    """Return persisted status for a queued cockpit action."""
    try:
        service = CockpitService.get_instance()
    except Exception as exc:
        logger.exception("Failed to initialize CockpitService for action job read")
        raise HTTPException(
            status_code=500, detail=f"Service initialization failed: {str(exc)}"
        ) from exc

    try:
        result = await asyncio.to_thread(
            _serialize_action_job_status, service, job_id, tail=tail
        )
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except Exception as exc:
        logger.exception("Cockpit action job read failed")
        raise HTTPException(
            status_code=500,
            detail=f"Action job read failed: {str(exc)}",
        ) from exc

    return CockpitActionJobStatusResponse(**result)


@router.post("/action/jobs/{job_id}/stop")
async def cockpit_stop_action_job(job_id: str):
    """Stop a running queued cockpit action job."""
    try:
        service = CockpitService.get_instance()
    except Exception as exc:
        logger.exception("Failed to initialize CockpitService for action stop")
        raise HTTPException(
            status_code=500, detail=f"Service initialization failed: {str(exc)}"
        ) from exc

    runtime = _get_queued_action_job(job_id)
    if runtime is not None:
        runtime.stop_event.set()

    with _ACTION_JOB_PROC_LOCK:
        proc = _ACTION_JOB_PROCS.get(job_id)
        if proc is not None:
            _ACTION_JOB_CANCEL_REQUESTS.add(job_id)

    if proc is not None:
        try:
            proc.terminate()
            try:
                proc.wait(timeout=3)
            except Exception:
                proc.kill()
        except Exception as exc:
            raise HTTPException(
                status_code=500,
                detail=f"Failed to stop action job: {str(exc)}",
            ) from exc
        return {"ok": True, "job_id": job_id, "status": "cancelling"}

    if runtime is not None:
        try:
            service.state_store.update_job_progress(job_id, "Cancelling", None)
        except Exception:
            logger.debug("Could not update cancelling progress for %s", job_id, exc_info=True)
        return {"ok": True, "job_id": job_id, "status": "cancelling"}

    tracker = _get_backend_job_tracker()
    tracker_job: dict[str, Any] | None = None
    if tracker is not None:
        try:
            tracker_job = tracker.store.get_job_run(job_id)
        except Exception as exc:
            logger.debug("Tracker job lookup failed for %s: %s", job_id, exc)

    if tracker_job is not None:
        tracker_status = str(tracker_job.get("status") or "unknown")
        if tracker_status in {"succeeded", "failed", "cancelled"}:
            return {"ok": True, "job_id": job_id, "status": tracker_status}

        tracker_job_type = str(tracker_job.get("job_type") or "").strip().lower()
        tracker_job_family = (
            str(tracker_job.get("job_family") or "").strip().lower()
        )
        tracker_metadata = dict(tracker_job.get("metadata") or {})
        tracker_supports_cancellation = bool(
            tracker_metadata.get("supports_cancellation")
        )
        if (
            tracker_supports_cancellation
            and tracker_job_type in {"extraction", "backfill"}
            and tracker_job_family in {"pipeline", "celery"}
        ):
            try:
                tracker.request_cancellation(
                    job_id, reason="Cancellation requested from Cockpit."
                )
            except Exception as exc:
                raise HTTPException(
                    status_code=500,
                    detail=f"Failed to request operation cancellation: {str(exc)}",
                ) from exc
            return {"ok": True, "job_id": job_id, "status": "cancelling"}

    job = service.state_store.get_job(job_id)
    if job is None:
        raise HTTPException(status_code=404, detail=f"Action job not found: {job_id}")

    status = str(job.get("status") or "unknown")
    if status in {"success", "failed", "cancelled"}:
        return {"ok": True, "job_id": job_id, "status": status}

    # Marketplace scan jobs that survived a backend restart have no live runtime or
    # process — force-cancel them in the state store so new scans are not blocked.
    if str(job.get("action_id") or "") == "marketplace_scan" and status in {"running", "queued"}:
        service.state_store.update_job_status(
            job_id,
            status="cancelled",
            exit_code=-1,
            ended_at=datetime.now(timezone.utc).isoformat(),
        )
        return {"ok": True, "job_id": job_id, "status": "cancelled"}

    raise HTTPException(
        status_code=409,
        detail=f"Action job is not currently stoppable: {job_id}",
    )


@router.get(
    "/marketplace/browser-health",
    response_model=MarketplaceBrowserHealthResponse,
)
async def cockpit_marketplace_browser_health():
    try:
        health = await asyncio.to_thread(check_marketplace_browser_health)
    except Exception as exc:
        logger.exception("Marketplace browser health check failed")
        raise HTTPException(
            status_code=500,
            detail=f"Marketplace browser health check failed: {str(exc)}",
        ) from exc
    return MarketplaceBrowserHealthResponse(**health)


@router.get(
    "/marketplace/missions",
    response_model=MarketplaceMissionListResponse,
)
async def cockpit_list_marketplace_missions(status: str | None = None):
    try:
        service = CockpitService.get_instance()
        mission_service = _marketplace_mission_service(service)
        price_service = _marketplace_price_intelligence_service(service)
        statuses = [item.strip().lower() for item in str(status or "").split(",") if item.strip()]
        items = await asyncio.to_thread(mission_service.list_missions, statuses=statuses or None)
        items = await asyncio.to_thread(
            lambda: [
                _enrich_marketplace_mission_with_price_context(
                    mission_service,
                    price_service,
                    item,
                )
                for item in items
            ]
        )
    except Exception as exc:
        logger.exception("Marketplace mission listing failed")
        raise HTTPException(
            status_code=500,
            detail=f"Marketplace mission listing failed: {str(exc)}",
        ) from exc
    return MarketplaceMissionListResponse(items=items)


@router.post(
    "/marketplace/missions",
    response_model=MarketplaceMissionRecord,
)
async def cockpit_create_marketplace_mission(payload: MarketplaceMissionUpsertRequest):
    try:
        service = CockpitService.get_instance()
        mission_service = _marketplace_mission_service(service)
        price_service = _marketplace_price_intelligence_service(service)
        mission = await asyncio.to_thread(mission_service.create_mission, payload.model_dump())
        mission = await asyncio.to_thread(
            _prepare_marketplace_requirement_candidates,
            mission_service,
            price_service,
            mission,
        )
        mission = await asyncio.to_thread(
            _enrich_marketplace_mission_with_price_context,
            mission_service,
            price_service,
            mission,
        )
    except MarketplaceMissionError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        logger.exception("Marketplace mission creation failed")
        raise HTTPException(
            status_code=500,
            detail=f"Marketplace mission creation failed: {str(exc)}",
        ) from exc
    return MarketplaceMissionRecord(**mission)


@router.get(
    "/marketplace/missions/{mission_id}",
    response_model=MarketplaceMissionRecord,
)
async def cockpit_get_marketplace_mission(mission_id: str):
    try:
        service = CockpitService.get_instance()
        mission_service = _marketplace_mission_service(service)
        price_service = _marketplace_price_intelligence_service(service)
        mission = await asyncio.to_thread(mission_service.get_mission, mission_id)
    except Exception as exc:
        logger.exception("Marketplace mission read failed")
        raise HTTPException(
            status_code=500,
            detail=f"Marketplace mission read failed: {str(exc)}",
        ) from exc
    if mission is None:
        raise HTTPException(status_code=404, detail=f"Marketplace mission not found: {mission_id}")
    mission = await asyncio.to_thread(
        _prepare_marketplace_requirement_candidates,
        mission_service,
        price_service,
        mission,
    )
    mission = await asyncio.to_thread(
        _enrich_marketplace_mission_with_price_context,
        mission_service,
        price_service,
        mission,
    )
    return MarketplaceMissionRecord(**mission)


@router.patch(
    "/marketplace/missions/{mission_id}",
    response_model=MarketplaceMissionRecord,
)
async def cockpit_update_marketplace_mission(
    mission_id: str,
    payload: MarketplaceMissionUpsertRequest,
):
    try:
        service = CockpitService.get_instance()
        mission_service = _marketplace_mission_service(service)
        price_service = _marketplace_price_intelligence_service(service)
        mission = await asyncio.to_thread(
            mission_service.update_mission,
            mission_id,
            payload.model_dump(exclude_none=True, exclude_unset=True),
        )
        mission = await asyncio.to_thread(
            _prepare_marketplace_requirement_candidates,
            mission_service,
            price_service,
            mission,
        )
        mission = await asyncio.to_thread(
            _enrich_marketplace_mission_with_price_context,
            mission_service,
            price_service,
            mission,
        )
    except MarketplaceMissionNotFound as exc:
        raise HTTPException(status_code=404, detail=f"Marketplace mission not found: {exc}") from exc
    except MarketplaceMissionError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        logger.exception("Marketplace mission update failed")
        raise HTTPException(
            status_code=500,
            detail=f"Marketplace mission update failed: {str(exc)}",
        ) from exc
    return MarketplaceMissionRecord(**mission)


@router.post(
    "/marketplace/missions/{mission_id}/link-product",
    response_model=MarketplaceMissionRecord,
)
async def cockpit_link_marketplace_mission_product(
    mission_id: str,
    payload: MarketplaceMissionProductLinkRequest,
):
    try:
        service = CockpitService.get_instance()
        mission_service = _marketplace_mission_service(service)
        price_service = _marketplace_price_intelligence_service(service)
        product = await asyncio.to_thread(
            price_service.get_tracked_product,
            payload.tracked_product_id,
        )
        if product is None:
            raise HTTPException(status_code=404, detail="tracked product not found")
        await asyncio.to_thread(
            mission_service.link_primary_tracked_product,
            mission_id,
            payload.tracked_product_id,
        )
        mission = await asyncio.to_thread(mission_service.get_mission, mission_id)
    except HTTPException:
        raise
    except MarketplaceMissionNotFound as exc:
        raise HTTPException(status_code=404, detail=f"Marketplace mission not found: {exc}") from exc
    except MarketplaceMissionError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        logger.exception("Marketplace mission product link failed")
        raise HTTPException(
            status_code=500,
            detail=f"Marketplace mission product link failed: {str(exc)}",
        ) from exc
    if mission is None:
        raise HTTPException(status_code=404, detail=f"Marketplace mission not found: {mission_id}")
    mission = await asyncio.to_thread(
        _enrich_marketplace_mission_with_price_context,
        mission_service,
        price_service,
        mission,
    )
    return MarketplaceMissionRecord(**mission)


@router.delete(
    "/marketplace/missions/{mission_id}/link-product",
    response_model=MarketplaceMissionRecord,
)
async def cockpit_unlink_marketplace_mission_product(mission_id: str):
    try:
        service = CockpitService.get_instance()
        mission_service = _marketplace_mission_service(service)
        price_service = _marketplace_price_intelligence_service(service)
        await asyncio.to_thread(mission_service.unlink_primary_tracked_product, mission_id)
        mission = await asyncio.to_thread(mission_service.get_mission, mission_id)
    except MarketplaceMissionNotFound as exc:
        raise HTTPException(status_code=404, detail=f"Marketplace mission not found: {exc}") from exc
    except Exception as exc:
        logger.exception("Marketplace mission product unlink failed")
        raise HTTPException(
            status_code=500,
            detail=f"Marketplace mission product unlink failed: {str(exc)}",
        ) from exc
    if mission is None:
        raise HTTPException(status_code=404, detail=f"Marketplace mission not found: {mission_id}")
    mission = await asyncio.to_thread(
        _enrich_marketplace_mission_with_price_context,
        mission_service,
        price_service,
        mission,
    )
    return MarketplaceMissionRecord(**mission)


@router.delete(
    "/marketplace/missions/{mission_id}",
    response_model=MarketplaceMissionDeleteResponse,
)
async def cockpit_delete_marketplace_mission(mission_id: str):
    try:
        service = CockpitService.get_instance()
        mission_service = _marketplace_mission_service(service)
        mission = await asyncio.to_thread(mission_service.get_mission, mission_id)
    except Exception as exc:
        logger.exception("Marketplace mission delete precheck failed")
        raise HTTPException(
            status_code=500,
            detail=f"Marketplace mission delete precheck failed: {str(exc)}",
        ) from exc

    if mission is None:
        raise HTTPException(status_code=404, detail=f"Marketplace mission not found: {mission_id}")

    try:
        scan_jobs = await asyncio.to_thread(_list_marketplace_scan_jobs, service, limit=100)
    except Exception as exc:
        logger.exception("Marketplace mission delete scan lookup failed")
        raise HTTPException(
            status_code=500,
            detail=f"Marketplace mission delete scan lookup failed: {str(exc)}",
        ) from exc

    active_scan = next(
        (
            job
            for job in scan_jobs
            if str(job.get("mission_id") or "") == mission_id
            and str(job.get("status") or "") in {"queued", "running"}
        ),
        None,
    )
    if active_scan is not None:
        raise HTTPException(
            status_code=409,
            detail="Mission has an active scan job. Cancel it before deleting the mission.",
        )

    try:
        deleted = await asyncio.to_thread(mission_service.delete_mission, mission_id)
    except MarketplaceMissionNotFound as exc:
        raise HTTPException(status_code=404, detail=f"Marketplace mission not found: {exc}") from exc
    except Exception as exc:
        logger.exception("Marketplace mission delete failed")
        raise HTTPException(
            status_code=500,
            detail=f"Marketplace mission delete failed: {str(exc)}",
        ) from exc
    return MarketplaceMissionDeleteResponse(**deleted)


@router.get(
    "/marketplace/scans",
    response_model=MarketplaceScanJobListResponse,
)
async def cockpit_list_marketplace_scan_jobs(limit: int = 50):
    try:
        service = CockpitService.get_instance()
        items = await asyncio.to_thread(_list_marketplace_scan_jobs, service, limit=limit)
    except Exception as exc:
        logger.exception("Marketplace scan listing failed")
        raise HTTPException(
            status_code=500,
            detail=f"Marketplace scan listing failed: {str(exc)}",
        ) from exc
    return MarketplaceScanJobListResponse(
        items=[CockpitActionJobStatusResponse(**item) for item in items]
    )


@router.post(
    "/marketplace/scans",
    response_model=CockpitActionExecuteResponse,
)
async def cockpit_trigger_marketplace_scan(payload: MarketplaceScanRequest):
    try:
        service = CockpitService.get_instance()
        mission_service = _marketplace_mission_service(service)
    except Exception as exc:
        logger.exception("Failed to initialize CockpitService for marketplace scan")
        raise HTTPException(
            status_code=500,
            detail=f"Service initialization failed: {str(exc)}",
        ) from exc

    mission_id = str(payload.mission_id or "").strip() or None
    if mission_id:
        mission = await asyncio.to_thread(mission_service.get_mission, mission_id)
        if mission is None:
            raise HTTPException(status_code=404, detail=f"Marketplace mission not found: {mission_id}")
    else:
        active_missions = await asyncio.to_thread(mission_service.list_missions, statuses=["active"])
        if not active_missions:
            raise HTTPException(status_code=400, detail="No active Marketplace missions are available to scan")

    if await asyncio.to_thread(_marketplace_scan_in_progress, service):
        raise HTTPException(
            status_code=409,
            detail="A Marketplace scan is already in progress. Please wait for it to finish or stop it first.",
        )

    health = await asyncio.to_thread(check_marketplace_browser_health)
    if not marketplace_scan_health_allows_execution(health):
        raise HTTPException(status_code=503, detail=str(health.get("detail") or health.get("status")))

    try:
        queued = await asyncio.to_thread(
            _launch_marketplace_scan_job,
            service,
            mission_id=mission_id,
        )
    except Exception as exc:
        logger.exception("Marketplace scan launch failed")
        raise HTTPException(
            status_code=500,
            detail=f"Marketplace scan launch failed: {str(exc)}",
        ) from exc
    return CockpitActionExecuteResponse(**queued)


@router.get(
    "/marketplace/scans/{job_id}",
    response_model=CockpitActionJobStatusResponse,
)
async def cockpit_get_marketplace_scan_job(job_id: str, tail: int = 0):
    try:
        service = CockpitService.get_instance()
        result = await asyncio.to_thread(
            _serialize_action_job_status,
            service,
            job_id,
            tail=tail,
        )
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=f"Marketplace scan job not found: {exc}") from exc
    except Exception as exc:
        logger.exception("Marketplace scan job read failed")
        raise HTTPException(
            status_code=500,
            detail=f"Marketplace scan job read failed: {str(exc)}",
        ) from exc
    if str(result.get("action_id") or "") != "marketplace_scan":
        raise HTTPException(status_code=404, detail=f"Marketplace scan job not found: {job_id}")
    return CockpitActionJobStatusResponse(**result)


@router.get(
    "/marketplace/matches",
    response_model=MarketplaceMatchListResponse,
)
async def cockpit_list_marketplace_matches(
    mission_id: str | None = None,
    status: str | None = None,
    decision_band: str | None = None,
    limit: int = 100,
):
    try:
        service = CockpitService.get_instance()
        mission_service = _marketplace_mission_service(service)
        benchmark_service = _marketplace_benchmark_service(service)
        price_service = _marketplace_price_intelligence_service(service)
        items = await asyncio.to_thread(
            mission_service.list_matches,
            mission_id=mission_id,
            status=status,
            decision_band=decision_band,
            limit=limit,
        )
        enriched_items: list[dict[str, Any]] = []
        for item in items:
            try:
                enriched = benchmark_service.enrich_match(item)
                enriched_items.append(
                    _enrich_marketplace_match_with_value_context(
                        mission_service,
                        price_service,
                        enriched,
                    )
                )
            except Exception:
                logger.exception(
                    "Marketplace benchmark enrichment failed for %s",
                    item.get("match_id"),
                )
                enriched_items.append({**item, "benchmark": None})
    except Exception as exc:
        logger.exception("Marketplace match listing failed")
        raise HTTPException(
            status_code=500,
            detail=f"Marketplace match listing failed: {str(exc)}",
        ) from exc
    return MarketplaceMatchListResponse(items=enriched_items)


@router.get(
    "/marketplace/matches/{match_id}",
    response_model=MarketplaceMatchRecord,
)
async def cockpit_get_marketplace_match(match_id: str):
    try:
        service = CockpitService.get_instance()
        mission_service = _marketplace_mission_service(service)
        benchmark_service = _marketplace_benchmark_service(service)
        price_service = _marketplace_price_intelligence_service(service)
        match = await asyncio.to_thread(mission_service.get_match, match_id)
    except Exception as exc:
        logger.exception("Marketplace match read failed")
        raise HTTPException(
            status_code=500,
            detail=f"Marketplace match read failed: {str(exc)}",
        ) from exc
    if match is None:
        raise HTTPException(status_code=404, detail=f"Marketplace match not found: {match_id}")
    try:
        match = benchmark_service.enrich_match(match)
        match = _enrich_marketplace_match_with_value_context(
            mission_service,
            price_service,
            match,
        )
    except Exception:
        logger.exception("Marketplace benchmark enrichment failed for %s", match_id)
        match = {**match, "benchmark": None, "value_context": None}
    return MarketplaceMatchRecord(**match)


@router.patch(
    "/marketplace/matches/{match_id}",
    response_model=MarketplaceMatchRecord,
)
async def cockpit_update_marketplace_match(
    match_id: str,
    payload: MarketplaceMatchStatusRequest,
):
    try:
        service = CockpitService.get_instance()
        mission_service = _marketplace_mission_service(service)
        benchmark_service = _marketplace_benchmark_service(service)
        price_service = _marketplace_price_intelligence_service(service)
        match = await asyncio.to_thread(
            mission_service.update_match_status,
            match_id,
            payload.status,
        )
    except MarketplaceMissionNotFound as exc:
        raise HTTPException(status_code=404, detail=f"Marketplace match not found: {exc}") from exc
    except MarketplaceMissionError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        logger.exception("Marketplace match update failed")
        raise HTTPException(
            status_code=500,
            detail=f"Marketplace match update failed: {str(exc)}",
        ) from exc
    try:
        match = benchmark_service.enrich_match(match)
        match = _enrich_marketplace_match_with_value_context(
            mission_service,
            price_service,
            match,
        )
    except Exception:
        logger.exception("Marketplace benchmark enrichment failed for %s", match_id)
        match = {**match, "benchmark": None, "value_context": None}
    return MarketplaceMatchRecord(**match)


@router.post("/marketplace/benchmarks/refresh")
async def cockpit_refresh_marketplace_benchmarks():
    try:
        service = CockpitService.get_instance()
        benchmark_service = _marketplace_benchmark_service(service)
        summary = await asyncio.to_thread(benchmark_service.refresh_centre_com_benchmarks)
    except Exception as exc:
        logger.exception("Marketplace benchmark refresh failed")
        raise HTTPException(
            status_code=500,
            detail=f"Marketplace benchmark refresh failed: {str(exc)}",
        ) from exc
    return {"ok": True, **summary}


@router.patch(
    "/marketplace/matches/{match_id}/benchmark-review",
    response_model=MarketplaceMatchRecord,
)
async def cockpit_update_marketplace_benchmark_review(
    match_id: str,
    payload: MarketplaceBenchmarkReviewRequest,
):
    review_status = str(payload.review_status or "").strip().lower()
    if review_status not in REVIEW_STATUSES:
        raise HTTPException(status_code=400, detail=f"invalid review_status: {payload.review_status}")

    try:
        service = CockpitService.get_instance()
        mission_service = _marketplace_mission_service(service)
        benchmark_service = _marketplace_benchmark_service(service)
        price_service = _marketplace_price_intelligence_service(service)
        await asyncio.to_thread(
            benchmark_service.set_review_status,
            match_id=match_id,
            review_status=review_status,
            note=payload.note,
        )
        match = await asyncio.to_thread(mission_service.get_match, match_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail=f"Marketplace match not found: {exc}") from exc
    except Exception as exc:
        logger.exception("Marketplace benchmark review update failed")
        raise HTTPException(
            status_code=500,
            detail=f"Marketplace benchmark review update failed: {str(exc)}",
        ) from exc

    if match is None:
        raise HTTPException(status_code=404, detail=f"Marketplace match not found: {match_id}")

    try:
        match = benchmark_service.enrich_match(match)
        match = _enrich_marketplace_match_with_value_context(
            mission_service,
            price_service,
            match,
        )
    except Exception:
        logger.exception("Marketplace benchmark enrichment failed for %s", match_id)
        match = {**match, "benchmark": None, "value_context": None}
    return MarketplaceMatchRecord(**match)


@router.get(
    "/marketplace/alerts",
    response_model=MarketplaceAlertListResponse,
)
async def cockpit_list_marketplace_alerts(
    mission_id: str | None = None,
    status: str | None = None,
    limit: int = 100,
):
    try:
        service = CockpitService.get_instance()
        mission_service = _marketplace_mission_service(service)
        items = await asyncio.to_thread(
            mission_service.list_alerts,
            mission_id=mission_id,
            status=status,
            limit=limit,
        )
    except Exception as exc:
        logger.exception("Marketplace alert listing failed")
        raise HTTPException(
            status_code=500,
            detail=f"Marketplace alert listing failed: {str(exc)}",
        ) from exc
    return MarketplaceAlertListResponse(items=items)


@router.patch(
    "/marketplace/alerts/{alert_id}",
    response_model=MarketplaceAlertRecord,
)
async def cockpit_update_marketplace_alert(
    alert_id: str,
    payload: MarketplaceAlertStatusRequest,
):
    try:
        service = CockpitService.get_instance()
        mission_service = _marketplace_mission_service(service)
        alert = await asyncio.to_thread(
            mission_service.update_alert_status,
            alert_id,
            payload.status,
        )
    except MarketplaceMissionNotFound as exc:
        raise HTTPException(status_code=404, detail=f"Marketplace alert not found: {exc}") from exc
    except MarketplaceMissionError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        logger.exception("Marketplace alert update failed")
        raise HTTPException(
            status_code=500,
            detail=f"Marketplace alert update failed: {str(exc)}",
        ) from exc
    return MarketplaceAlertRecord(**alert)


@router.post("/feedback/flag", response_model=CockpitFeedbackFlagResponse)
async def cockpit_flag_feedback(payload: CockpitFeedbackFlagRequest):
    """Persist a flagged cockpit chat turn with relevant backend diagnostics."""
    try:
        service = CockpitService.get_instance()
    except Exception as exc:
        logger.exception("Failed to initialize CockpitService for feedback capture")
        raise HTTPException(
            status_code=500, detail=f"Service initialization failed: {str(exc)}"
        ) from exc

    try:
        result = await asyncio.to_thread(
            service.flag_chat_feedback,
            session_id=payload.session_id,
            ticker=payload.ticker,
            feedback_type=payload.feedback_type,
            capture_kind=payload.capture_kind,
            note=payload.note,
            flagged_message=payload.flagged_message,
            transcript=payload.transcript,
            frontend_context=payload.frontend_context,
            screenshot=payload.screenshot,
        )
    except Exception as exc:
        logger.exception("Cockpit feedback capture failed")
        raise HTTPException(
            status_code=500,
            detail=f"Feedback capture failed: {str(exc)}",
        ) from exc

    return CockpitFeedbackFlagResponse(**result)


@router.get("/feedback/flags", response_model=CockpitFlaggedReportListResponse)
async def cockpit_list_flagged_feedback(
    limit: int = 25,
    status: Literal["open", "resolved", "all"] = "open",
):
    """List recent flagged cockpit chat reports."""
    try:
        service = CockpitService.get_instance()
    except Exception as exc:
        logger.exception("Failed to initialize CockpitService for feedback listing")
        raise HTTPException(
            status_code=500, detail=f"Service initialization failed: {str(exc)}"
        ) from exc

    try:
        items = await asyncio.to_thread(service.list_flagged_reports, limit, status)
    except Exception as exc:
        logger.exception("Cockpit feedback listing failed")
        raise HTTPException(
            status_code=500,
            detail=f"Feedback listing failed: {str(exc)}",
        ) from exc

    return CockpitFlaggedReportListResponse(items=items)


@router.get("/feedback/flags/{report_id}", response_model=CockpitFlaggedReportResponse)
async def cockpit_get_flagged_feedback(report_id: str):
    """Return one flagged cockpit chat report by report_id."""
    try:
        service = CockpitService.get_instance()
    except Exception as exc:
        logger.exception("Failed to initialize CockpitService for feedback read")
        raise HTTPException(
            status_code=500, detail=f"Service initialization failed: {str(exc)}"
        ) from exc

    try:
        result = await asyncio.to_thread(service.get_flagged_report, report_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except Exception as exc:
        logger.exception("Cockpit feedback read failed")
        raise HTTPException(
            status_code=500,
            detail=f"Feedback read failed: {str(exc)}",
        ) from exc

    return CockpitFlaggedReportResponse(**result)


@router.post(
    "/feedback/flags/{report_id}/resolve",
    response_model=CockpitFlagResolutionResponse,
)
async def cockpit_resolve_flagged_feedback(
    report_id: str,
    payload: CockpitFlagResolutionRequest,
):
    """Mark a flagged cockpit report as resolved and attach fix commit metadata."""
    try:
        service = CockpitService.get_instance()
    except Exception as exc:
        logger.exception("Failed to initialize CockpitService for feedback resolve")
        raise HTTPException(
            status_code=500, detail=f"Service initialization failed: {str(exc)}"
        ) from exc

    try:
        result = await asyncio.to_thread(
            service.resolve_flagged_report,
            report_id,
            commit_sha=payload.commit_sha,
            resolved_by=payload.resolved_by,
            note=payload.note,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except Exception as exc:
        logger.exception("Cockpit feedback resolve failed")
        raise HTTPException(
            status_code=500,
            detail=f"Feedback resolve failed: {str(exc)}",
        ) from exc

    return CockpitFlagResolutionResponse(**result)


def _normalize_session_id(raw: str) -> str:
    session_id = str(raw or "").strip()
    if not session_id:
        raise HTTPException(status_code=400, detail="session_id is required")
    return session_id[:128]


def _derive_chat_title(
    *,
    title_seed: str | None,
    last_message: str | None,
) -> str:
    seed = str(title_seed or "").strip() or str(last_message or "").strip()
    if not seed:
        return "Untitled Chat"
    normalized = " ".join(seed.split())
    return (normalized[:80] + "...") if len(normalized) > 80 else normalized


def _parse_preference_bool(raw: str | None, default: bool = False) -> bool:
    text = str(raw or "").strip().lower()
    if not text:
        return default
    return text in {"1", "true", "yes", "on"}


@router.get("/preferences", response_model=CockpitPreferencesResponse)
def cockpit_get_preferences() -> CockpitPreferencesResponse:
    try:
        service = CockpitService.get_instance()
        api_default_enabled = _parse_preference_bool(
            service.state_store.get_preference("api_default_enabled", "false"),
            default=False,
        )
        marketplace_prefer_cloud_routing = _parse_preference_bool(
            service.state_store.get_preference(
                "marketplace_prefer_cloud_routing",
                "false",
            ),
            default=False,
        )
    except Exception as exc:
        logger.exception("Failed to load cockpit preferences")
        raise HTTPException(
            status_code=500, detail=f"Failed to load cockpit preferences: {str(exc)}"
        ) from exc

    return CockpitPreferencesResponse(
        api_default_enabled=api_default_enabled,
        marketplace_prefer_cloud_routing=marketplace_prefer_cloud_routing,
    )


@router.patch("/preferences", response_model=CockpitPreferencesResponse)
def cockpit_patch_preferences(
    payload: CockpitPreferencesPatchRequest,
) -> CockpitPreferencesResponse:
    try:
        service = CockpitService.get_instance()
        if payload.api_default_enabled is not None:
            service.state_store.set_preference(
                "api_default_enabled",
                "true" if payload.api_default_enabled else "false",
            )
        if payload.marketplace_prefer_cloud_routing is not None:
            service.state_store.set_preference(
                "marketplace_prefer_cloud_routing",
                "true" if payload.marketplace_prefer_cloud_routing else "false",
            )
    except Exception as exc:
        logger.exception("Failed to update cockpit preferences")
        raise HTTPException(
            status_code=500, detail=f"Failed to update cockpit preferences: {str(exc)}"
        ) from exc

    return cockpit_get_preferences()


@router.get("/chat/sessions", response_model=CockpitChatSessionListResponse)
def cockpit_list_chat_sessions(limit: int = 100) -> CockpitChatSessionListResponse:
    try:
        service = CockpitService.get_instance()
        rows = service.state_store.list_chat_sessions(limit=limit)
    except Exception as exc:
        logger.exception("Failed to list cockpit chat sessions")
        raise HTTPException(
            status_code=500, detail=f"Failed to list chat sessions: {str(exc)}"
        ) from exc

    items = [
        CockpitChatSessionSummary(
            session_id=str(row.get("thread_id") or ""),
            updated_at=str(row.get("updated_at") or "").strip() or None,
            message_count=int(row.get("message_count") or 0),
            title=_derive_chat_title(
                title_seed=str(row.get("title_seed") or "").strip() or None,
                last_message=str(row.get("last_message") or "").strip() or None,
            ),
            last_message=str(row.get("last_message") or "").strip() or None,
        )
        for row in rows
        if str(row.get("thread_id") or "").strip()
    ]
    return CockpitChatSessionListResponse(items=items)


@router.post("/chat/sessions", response_model=CockpitChatSessionCreateResponse)
def cockpit_create_chat_session(
    payload: CockpitChatSessionCreateRequest,
) -> CockpitChatSessionCreateResponse:
    raw_session_id = str(payload.session_id or "").strip()
    thread_id = _normalize_session_id(raw_session_id or str(uuid.uuid4()))
    try:
        service = CockpitService.get_instance()
        created = service.state_store.ensure_chat_session(thread_id)
    except Exception as exc:
        logger.exception("Failed to create cockpit chat session")
        raise HTTPException(
            status_code=500, detail=f"Failed to create chat session: {str(exc)}"
        ) from exc

    return CockpitChatSessionCreateResponse(
        ok=True,
        session_id=thread_id,
        created=bool(created),
    )


@router.get(
    "/chat/sessions/{session_id}",
    response_model=CockpitChatSessionMessagesResponse,
)
def cockpit_get_chat_session_messages(
    session_id: str,
    limit: int = 400,
) -> CockpitChatSessionMessagesResponse:
    thread_id = _normalize_session_id(session_id)
    try:
        service = CockpitService.get_instance()
        rows = service.state_store.get_chat_messages_with_ids(thread_id, limit=limit)
    except Exception as exc:
        logger.exception("Failed to load cockpit chat session messages")
        raise HTTPException(
            status_code=500, detail=f"Failed to load chat session: {str(exc)}"
        ) from exc

    items = [
        CockpitChatMessageRecord(
            id=int(row.get("id") or 0),
            session_id=thread_id,
            role=str(row.get("role") or "system"),
            content=str(row.get("content") or ""),
            created_at=str(row.get("created_at") or ""),
        )
        for row in rows
    ]
    return CockpitChatSessionMessagesResponse(
        session_id=thread_id,
        message_count=len(items),
        items=items,
    )


@router.delete(
    "/chat/sessions/{session_id}",
    response_model=CockpitChatSessionDeleteResponse,
)
def cockpit_delete_chat_session(session_id: str) -> CockpitChatSessionDeleteResponse:
    thread_id = _normalize_session_id(session_id)
    try:
        service = CockpitService.get_instance()
        existed = service.state_store.has_chat_session(thread_id)
        deleted_count = service.state_store.delete_chat_session(thread_id)
    except Exception as exc:
        logger.exception("Failed to delete cockpit chat session")
        raise HTTPException(
            status_code=500, detail=f"Failed to delete chat session: {str(exc)}"
        ) from exc

    return CockpitChatSessionDeleteResponse(
        ok=bool(existed),
        session_id=thread_id,
        deleted_count=deleted_count,
    )


@router.post("/chat")
async def cockpit_chat(payload: CockpitChatRequest, request: Request):
    """
    Unified cockpit chat endpoint supporting SSE streaming.
    Matches the TUI's ChatController logic but exposed for the Web UI.
    """
    try:
        service = CockpitService.get_instance()
    except Exception as exc:
        logger.exception("Failed to initialize CockpitService")
        raise HTTPException(
            status_code=500, detail=f"Service initialization failed: {str(exc)}"
        ) from exc

    if not payload.stream:
        # Blocking implementation if requested (rare for this UI)
        try:
            attached_sources = [item.model_dump() for item in payload.attached_sources]
            response = await asyncio.to_thread(
                service.chat_stream,
                message=payload.message,
                ticker=payload.ticker,
                session_id=payload.session_id,
                enable_web=payload.web_search,
                model=payload.model,
                rag=payload.rag,
                db_diagnostics=payload.db_diagnostics,
                ui_mode=payload.mode,
                attached_sources=attached_sources,
            )
            sources = _enforce_visible_source_contract(payload.message, response)
            auto_flag = _maybe_auto_flag_chat_response(
                service,
                session_id=payload.session_id,
                ticker=payload.ticker,
                response=response,
            )
            rendered_chart = _build_chart_from_chat_response(response)
            return {
                "type": "done",
                "data": {
                    "text": response.text,
                    "model": response.routing_metadata.get("model")
                    if response.routing_metadata
                    else "local",
                    "latency_ms": response.routing_metadata.get("latency_ms")
                    if response.routing_metadata
                    else 0,
                    "cost_usd": response.routing_metadata.get("cost_usd")
                    if response.routing_metadata
                    else 0,
                    "source": response.routing_metadata.get("source")
                    if response.routing_metadata
                    else "local",
                    "provider_error": response.routing_metadata.get("provider_error")
                    if response.routing_metadata
                    else None,
                    "action_preview": response.action_preview,
                    "chart": rendered_chart,
                    "sources": sources,
                    "auto_flag": _serialize_flag_handoff(auto_flag),
                },
            }
        except Exception as exc:
            logger.exception("Cockpit chat non-streaming error")
            raise HTTPException(
                status_code=500, detail=f"Chat processing failed: {str(exc)}"
            ) from exc

    async def event_generator() -> AsyncGenerator[str, None]:
        queue = asyncio.Queue()
        loop = asyncio.get_running_loop()

        def on_chunk(chunk: str):
            # This runs in the LLM thread (from ChatController)
            loop.call_soon_threadsafe(
                queue.put_nowait, {"type": "chunk", "data": {"text": chunk}}
            )

        def on_status(stage: str):
            loop.call_soon_threadsafe(
                queue.put_nowait, {"type": "status", "data": {"stage": stage}}
            )

        async def run_chat():
            try:
                await queue.put(
                    {"type": "status", "data": {"stage": "Resolving request context"}}
                )
                attached_sources = [item.model_dump() for item in payload.attached_sources]
                # ChatController.build_chat_response is synchronous and blocking.
                # We run it in a thread to keep the event loop free.
                response = await asyncio.to_thread(
                    service.chat_stream,
                    message=payload.message,
                    ticker=payload.ticker,
                    session_id=payload.session_id,
                    on_chunk=on_chunk,
                    on_status=on_status,
                    enable_web=payload.web_search,
                    model=payload.model,
                    rag=payload.rag,
                    db_diagnostics=payload.db_diagnostics,
                    ui_mode=payload.mode,
                    attached_sources=attached_sources,
                )
                sources = _enforce_visible_source_contract(payload.message, response)
                auto_flag = _maybe_auto_flag_chat_response(
                    service,
                    session_id=payload.session_id,
                    ticker=payload.ticker,
                    response=response,
                )

                # After streaming finishes, send metadata and final state
                if response.tool_traces:
                    for trace in response.tool_traces:
                        await queue.put({"type": "tool_trace", "data": trace})

                if sources:
                    await queue.put({"type": "sources", "data": {"items": sources}})

                if response.action_preview:
                    await queue.put(
                        {"type": "action_preview", "data": response.action_preview}
                    )

                rendered_chart = _build_chart_from_chat_response(response)
                if rendered_chart:
                    await queue.put({"type": "chart", "data": rendered_chart})

                # Final 'done' event with metrics
                meta = response.routing_metadata or {}
                await queue.put(
                    {
                        "type": "done",
                        "data": {
                            "text": response.text,
                            "model": meta.get("model", "local"),
                            "latency_ms": meta.get("latency_ms", 0),
                            "cost_usd": meta.get("cost_usd", 0),
                            "source": meta.get("source", "local"),
                            "provider_error": meta.get("provider_error"),
                            "chart": rendered_chart,
                            "sources": sources,
                            "auto_flag": _serialize_flag_handoff(auto_flag),
                        },
                    }
                )
            except asyncio.CancelledError:
                logger.info("Cockpit chat stream cancelled by client disconnect")
            except Exception as exc:
                logger.exception("Cockpit chat streaming error")
                await queue.put({"type": "error", "data": str(exc)})
            finally:
                # Signal end of stream
                await queue.put(None)

        # Emit an immediate status so UI leaves "Preparing request" quickly.
        yield f"data: {json.dumps({'type': 'status', 'data': {'stage': 'Request accepted'}})}\n\n"

        # Start the chat worker
        worker_task = asyncio.create_task(run_chat())

        # SSE keepalive: yield a comment line if no real event has been sent
        # for this long. Prevents intermediaries (nginx, corporate proxies) from
        # silently tearing down the connection during long LLM passes, and gives
        # the client a signal that the server is still alive.
        keepalive_interval = SSE_KEEPALIVE_INTERVAL_SECONDS
        last_yield_monotonic = time.monotonic()

        while True:
            # Check for client disconnect
            if await request.is_disconnected():
                worker_task.cancel()
                break

            try:
                item = await asyncio.wait_for(queue.get(), timeout=0.25)
            except asyncio.TimeoutError:
                if time.monotonic() - last_yield_monotonic >= keepalive_interval:
                    yield ": keepalive\n\n"
                    last_yield_monotonic = time.monotonic()
                continue

            if item is None:
                break

            yield f"data: {json.dumps(item)}\n\n"
            last_yield_monotonic = time.monotonic()

        yield "event: end\ndata: {}\n\n"

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "X-Accel-Buffering": "no",
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
        },
    )


# -------------------------------------------------------------------
# TradingView Pine Script webhook endpoints
# -------------------------------------------------------------------


class TvAlertPayload(BaseModel):
    model_config = ConfigDict(extra="allow")

    ticker: str
    action: str = "neutral"
    price: float | None = None
    message: str | None = None
    timestamp: str | None = None


_TV_ALERTS_LOCK = threading.Lock()
_TV_ALERTS_MAX = 200


def _tv_alerts_path() -> Path:
    data_root = Path(settings.data_root) if hasattr(settings, "data_root") else Path("/tmp")
    return data_root / "tv_alerts.json"


def _load_tv_alerts() -> list[dict]:
    p = _tv_alerts_path()
    if not p.exists():
        return []
    try:
        return json.loads(p.read_text())
    except Exception:
        return []


def _save_tv_alerts(alerts: list[dict]) -> None:
    p = _tv_alerts_path()
    tmp = p.with_suffix(".tmp")
    tmp.write_text(json.dumps(alerts, default=str))
    tmp.replace(p)


@router.post("/tv/alert", tags=["tradingview"])
async def receive_tv_alert(
    payload: TvAlertPayload,
    request: Request,
) -> dict:
    """Receive a Pine Script webhook alert from TradingView."""
    token = os.environ.get("TV_WEBHOOK_TOKEN", "")
    if token:
        incoming = request.headers.get("X-TradingView-Webhook-Token", "")
        if incoming != token:
            raise HTTPException(status_code=403, detail="Invalid webhook token")

    entry: dict = {
        "received_at": datetime.now(timezone.utc).isoformat(),
        **payload.model_dump(),
    }
    with _TV_ALERTS_LOCK:
        alerts = _load_tv_alerts()
        alerts.append(entry)
        if len(alerts) > _TV_ALERTS_MAX:
            alerts = alerts[-_TV_ALERTS_MAX:]
        _save_tv_alerts(alerts)

    logger.info("TV alert received: %s %s @ %s", payload.ticker, payload.action, payload.price)
    return {"ok": True, "received": entry}


@router.get("/tv/alerts", tags=["tradingview"])
async def get_tv_alerts(limit: int = 50) -> dict:
    """Return recent TradingView Pine Script alerts."""
    limit = max(1, min(limit, 200))
    with _TV_ALERTS_LOCK:
        alerts = _load_tv_alerts()
    return {"ok": True, "count": len(alerts[-limit:]), "alerts": list(reversed(alerts[-limit:]))}
